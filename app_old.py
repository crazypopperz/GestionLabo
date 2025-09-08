import sqlite3
import os
import re
import sys
import math
import uuid
import unicodedata
import hashlib
import logging
from logging.handlers import RotatingFileHandler
from flask_wtf.csrf import CSRFProtect
from flask import (Flask, render_template, request, redirect, url_for,
                   send_file, jsonify, g, flash, session, send_from_directory)
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
import shutil
from datetime import timedelta, datetime, date
from functools import wraps
import traceback

# Importations pour l'export
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
app = Flask(__name__)

# --- CONFIGURATION DE L'APPLICATION ---
# 1. Clé secrète et session (DOIT ÊTRE DÉFINI EN PREMIER)
app.config['SECRET_KEY'] = 'une-cle-secrete-tres-difficile-a-deviner'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

# 2. Protection CSRF (initialisée APRÈS la clé secrète)
csrf = CSRFProtect(app)

# 3. Chemins des données utilisateur
if sys.platform == "win32":
    # Chemin pour Windows (%APPDATA%)
    appdata = os.getenv('APPDATA')
elif sys.platform == "darwin":
    # Chemin pour macOS (~/Library/Application Support)
    appdata = os.path.join(os.path.expanduser('~'), 'Library', 'Application Support')
else:
    # Chemin pour Linux et autres (~/.local/share)
    appdata = os.path.join(os.path.expanduser('~'), '.local', 'share')

# Si pour une raison quelconque le chemin de base n'est pas trouvé, on se rabat sur le répertoire du script
if not appdata:
    appdata = os.path.dirname(os.path.abspath(__file__))

USER_DATA_PATH = os.path.join(appdata, 'GMLCL')
os.makedirs(USER_DATA_PATH, exist_ok=True)
app.config['UPLOAD_FOLDER'] = os.path.join(USER_DATA_PATH, 'uploads', 'images')
app.config['FDS_UPLOAD_FOLDER'] = os.path.join(USER_DATA_PATH, 'uploads', 'fds')
DATABASE = os.path.join(USER_DATA_PATH, 'base.db')

# 4. Création des dossiers nécessaires
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['FDS_UPLOAD_FOLDER'], exist_ok=True)

# 5. Configuration du logging
if not app.debug:
    log_file_path = os.path.join(USER_DATA_PATH, 'app.log')
    logging.basicConfig(
        level=logging.ERROR,
        format='%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]',
        handlers=[
            RotatingFileHandler(
                log_file_path, maxBytes=1024000, backupCount=5
            )
        ]
    )
    app.logger.info('Logging configuré pour écrire dans %s', log_file_path)

ITEMS_PER_PAGE = 10
CLE_PRO_SECRETE = "LABO-PRO-2025-X@v14211825!S@cha14211825!Quentin14211825!"

# --- GESTION CENTRALE DE LA BASE DE DONNÉES ---
def strip_accents(text):
    """Fonction pour retirer les accents d'une chaîne."""
    return ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        db.create_function("unaccent", 1, strip_accents)
    return db


@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()


# --- DÉCORATEURS DE SÉCURITÉ ---
def login_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash("Veuillez vous connecter pour accéder à cette page.",
                  "error")
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('user_role') != 'admin':
            flash("Accès réservé aux administrateurs.", "error")
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated_function


def pro_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):
        db = get_db()
        try:
            licence_row = db.execute(
                "SELECT valeur FROM parametres WHERE cle = ?",
                ('licence_statut', )).fetchone()
            is_pro = licence_row and licence_row['valeur'] == 'PRO'
        except sqlite3.Error:
            is_pro = False
        if not is_pro:
            flash("Cette fonctionnalité est réservée à la version Pro.",
                  "warning")
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated_function


def limit_objets_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):
        db = get_db()
        licence_row = db.execute("SELECT valeur FROM parametres WHERE cle = ?",
                                 ('licence_statut', )).fetchone()
        is_pro = licence_row and licence_row['valeur'] == 'PRO'

        if not is_pro:
            count = db.execute("SELECT COUNT(id) FROM objets").fetchone()[0]
            if count >= 50:
                flash(
                    "La version gratuite est limitée à 50 objets. "
                    "Passez à la version Pro pour en ajouter davantage.",
                    "warning")
                return redirect(request.referrer or url_for('index'))

        return f(*args, **kwargs)

    return decorated_function


# --- FILTRES JINJA2 PERSONNALISÉS ---
def format_datetime(value, fmt='%d/%m/%Y %H:%M'):
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except (ValueError, TypeError):
            try:
                value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S.%f')
            except (ValueError, TypeError):
                try:
                    value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    return value
    if isinstance(value, (datetime, date)):
        return value.strftime(fmt)
    return value


app.jinja_env.filters['strftime'] = format_datetime

def format_datetime_fr(value, fmt):
    """
    Un filtre strftime personnalisé qui traduit les jours et mois en français.
    """
    # --- BLOC AJOUTÉ POUR LA ROBUSTESSE ---
    # D'abord, on s'assure qu'on a bien un objet date/datetime
    if isinstance(value, str):
        try:
            # Tente de convertir la chaîne en objet datetime
            value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            # En cas d'échec, on essaie un autre format ou on renvoie la chaîne originale
            try:
                value = datetime.fromisoformat(value)
            except (ValueError, TypeError):
                return value
    if not isinstance(value, (datetime, date)):
        return value

    jours = ["lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"]
    mois = ["janvier", "février", "mars", "avril", "mai", "juin", "juillet",
            "août", "septembre", "octobre", "novembre", "décembre"]

    # Remplace les codes de format par les chaînes françaises
    # Note : capitalize() met la première lettre en majuscule
    format_fr = fmt.replace('%A', jours[value.weekday()].capitalize())
    format_fr = format_fr.replace('%B', mois[value.month - 1])

    return value.strftime(format_fr)

app.jinja_env.filters['strftime_fr'] = format_datetime_fr

def annee_scolaire_format(year):
    """Transforme une année de début (ex: 2025) en format année scolaire (ex: "2025-2026")."""
    if isinstance(year, int):
        return f"{year}-{year + 1}"
    return year

app.jinja_env.filters['annee_scolaire'] = annee_scolaire_format

# --- GESTION DE L'INITIALISATION AU PREMIER LANCEMENT ---
def is_setup_needed():
    with app.app_context():
        db = get_db()
        user = db.execute("SELECT id FROM utilisateurs LIMIT 1").fetchone()
        return user is None


@app.before_request
def check_setup():
    if not os.path.exists(DATABASE):
        return
    allowed_endpoints = ['static', 'setup', 'login', 'register']
    if request.endpoint and request.endpoint not in allowed_endpoints:
        if is_setup_needed():
            return redirect(url_for('setup'))


@app.route("/setup", methods=['GET', 'POST'])
def setup():
    if not is_setup_needed():
        flash("L'application est déjà configurée.", "error")
        return redirect(url_for('login'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password')
        password_confirm = request.form.get('password_confirm')
        email = request.form.get('email', '').strip()
        if not all([username, password, password_confirm, email]):
            flash("Tous les champs sont requis.", "error")
            return redirect(url_for('setup'))
        if len(password) < 12 or \
           not re.search(r"[a-z]", password) or \
           not re.search(r"[A-Z]", password) or \
           not re.search(r"[0-9]", password) or \
           not re.search(r"[!@#$%^&*(),.?:{}|<>]", password):
            flash("Le mot de passe doit contenir au moins 12 caractères, "
                  "incluant une majuscule, une minuscule, un chiffre et "
                  "un caractère spécial.", "error")
            return redirect(url_for('setup'))
        if password != password_confirm: # <-- NOUVELLE VÉRIFICATION
            flash("Les mots de passe ne correspondent pas.", "error")
            return redirect(url_for('setup'))
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_regex, email):
            flash("L'adresse e-mail fournie n'est pas valide.", "error")
            return redirect(url_for('setup'))
        db = get_db()
        db.execute(
            "INSERT INTO utilisateurs (nom_utilisateur, mot_de_passe, role, "
            "email) VALUES (?, ?, 'admin', ?)",
            (username, generate_password_hash(password,
                                              method='scrypt'), email))
        instance_id = str(uuid.uuid4())
        db.execute(
            "INSERT OR IGNORE INTO parametres (cle, valeur) VALUES (?, ?)",
            ('instance_id', instance_id))
        db.commit()
        flash(
            f"Administrateur '{username}' créé avec succès ! "
            "Vous pouvez maintenant vous connecter.", "success")
        return redirect(url_for('login'))
    return render_template('setup.html')


# --- FONCTIONS COMMUNES ET PROCESSEUR DE CONTEXTE ---
def get_alerte_info(db):
    date_limite = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
    count_stock = db.execute(
        "SELECT COUNT(*) FROM objets WHERE quantite <= seuil AND en_commande = 0"
    ).fetchone()[0]
    count_peremption = db.execute(
        "SELECT COUNT(*) FROM objets WHERE date_peremption IS NOT NULL AND "
        "date_peremption < ? AND traite = 0", (date_limite, )).fetchone()[0]
    total_alertes = count_stock + count_peremption
    return {
        "alertes_stock": count_stock,
        "alertes_peremption": count_peremption,
        "alertes_total": total_alertes
    }


@app.context_processor
def inject_alert_info():
    if 'user_id' in session:
        db = get_db()
        return get_alerte_info(db)
    return {'alertes_total': 0, 'alertes_stock': 0, 'alertes_peremption': 0}


@app.context_processor
def inject_licence_info():
    """
    Injecte le statut de la licence dans le contexte de tous les templates.
    Rend la variable 'licence' disponible globalement.
    """
    licence_info = {'statut': 'FREE', 'is_pro': False, 'instance_id': 'N/A'}
    if 'user_id' in session and g.get('_database', None):
        try:
            db = get_db()
            params = db.execute(
                "SELECT cle, valeur FROM parametres "
                "WHERE cle IN ('licence_statut', 'instance_id')").fetchall()
            params_dict = {row['cle']: row['valeur'] for row in params}

            if params_dict.get('licence_statut') == 'PRO':
                licence_info['statut'] = 'PRO'
                licence_info['is_pro'] = True

            if params_dict.get('instance_id'):
                licence_info['instance_id'] = params_dict.get('instance_id')

        except sqlite3.Error as e:
            app.logger.warning(
                f"Impossible de lire les informations de licence. Erreur : {e}"
            )
    return {'licence': licence_info}


def get_paginated_objets(db,
                         page,
                         sort_by='nom',
                         direction='asc',
                         search_query=None,
                         armoire_id=None,
                         categorie_id=None,
                         etat=None,
                         filter_field=None,
                         filter_id=None):
    offset = (page - 1) * ITEMS_PER_PAGE

    valid_sort_columns = {
        'nom': 'o.nom',
        'quantite': 'o.quantite',
        'seuil': 'o.seuil',
        'date_peremption': 'o.date_peremption',
        'categorie': 'c.nom',
        'armoire': 'a.nom'
    }
    sort_column = valid_sort_columns.get(sort_by, 'o.nom')
    sort_direction = 'DESC' if direction == 'desc' else 'ASC'

    base_query = """
    SELECT o.id, o.nom, o.quantite, o.seuil, o.armoire_id, o.categorie_id,
           o.fds_nom_original, o.fds_nom_securise,
           a.nom AS armoire, c.nom AS categorie, o.image, o.en_commande,
           o.date_peremption
    FROM objets o
    JOIN armoires a ON o.armoire_id = a.id
    JOIN categories c ON o.categorie_id = c.id
    """
    count_query = ("SELECT COUNT(*) FROM objets o "
                   "JOIN armoires a ON o.armoire_id = a.id "
                   "JOIN categories c ON o.categorie_id = c.id")

    conditions = []
    params = []

    if filter_field and filter_id:
        conditions.append(f"o.{filter_field} = ?")
        params.append(filter_id)

    if search_query:
        conditions.append("unaccent(LOWER(o.nom)) LIKE unaccent(LOWER(?))")
        params.append(f"%{search_query}%")

    if armoire_id:
        conditions.append("o.armoire_id = ?")
        params.append(armoire_id)

    if categorie_id:
        conditions.append("o.categorie_id = ?")
        params.append(categorie_id)

    if etat:
        now_str = datetime.now().strftime('%Y-%m-%d')
        date_limite = (datetime.now() +
                       timedelta(days=30)).strftime('%Y-%m-%d')
        if etat == 'perime':
            conditions.append("o.date_peremption < ?")
            params.append(now_str)
        elif etat == 'bientot':
            conditions.append(
                "o.date_peremption >= ? AND o.date_peremption < ?")
            params.extend([now_str, date_limite])
        elif etat == 'stock':
            conditions.append("o.quantite < o.seuil")
        elif etat == 'ok':
            conditions.append(
                "o.quantite >= o.seuil AND (o.date_peremption IS NULL OR "
                "o.date_peremption >= ?)")
            params.append(date_limite)

    if conditions:
        where_clause = " WHERE " + " AND ".join(conditions)
        base_query += where_clause
        count_query += where_clause

    total_objets = db.execute(count_query, params).fetchone()[0]
    total_pages = math.ceil(total_objets / ITEMS_PER_PAGE)
    if ITEMS_PER_PAGE <= 0:
        total_pages = 0

    base_query += f" ORDER BY {sort_column} {sort_direction}, o.nom ASC"
    base_query += " LIMIT ? OFFSET ?"
    params.extend([ITEMS_PER_PAGE, offset])

    objets = db.execute(base_query, params).fetchall()
    return objets, total_pages


def enregistrer_action(objet_id, action, details=""):
    if 'user_id' in session:
        db = get_db()
        try:
            db.execute(
                """INSERT INTO historique (objet_id, utilisateur_id, action,
                   details, timestamp)
                   VALUES (?, ?, ?, ?, ?)""",
                (objet_id, session['user_id'], action, details,
                 datetime.now()))
            db.commit()
        except sqlite3.Error as e:
            print(f"ERREUR LORS DE L'ENREGISTREMENT DE L'HISTORIQUE : {e}")
            db.rollback()

# --- ROUTES PRINCIPALES ---
@app.route("/")
@login_required
def index():
    db = get_db()
    dashboard_data = {}

    if session.get('user_role') == 'admin':
        dashboard_data['stats'] = {
            'total_objets':
            db.execute("SELECT COUNT(*) FROM objets").fetchone()[0],
            'total_utilisateurs':
            db.execute("SELECT COUNT(*) FROM utilisateurs").fetchone()[0],
            'reservations_actives':
            db.execute(
                "SELECT COUNT(DISTINCT groupe_id) FROM reservations "
                "WHERE debut_reservation >= ?",
                (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), )).fetchone()[0]
        }
    
    now = datetime.now()
    # CORRECTION : Le basculement se fait maintenant en août (mois 8)
    annee_scolaire_actuelle = now.year if now.month >= 8 else now.year - 1

    budget_actuel = db.execute(
        "SELECT * FROM budgets WHERE annee = ? AND cloture = 0",
        (annee_scolaire_actuelle, )).fetchone()

    solde_actuel = None
    if budget_actuel:
        total_depenses_result = db.execute(
            "SELECT SUM(montant) as total FROM depenses WHERE budget_id = ?",
            (budget_actuel['id'], )).fetchone()
        total_depenses = (total_depenses_result['total']
                          if total_depenses_result['total'] is not None else 0)
        solde_actuel = budget_actuel['montant_initial'] - total_depenses

    dashboard_data['solde_budget'] = solde_actuel

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    dashboard_data['reservations'] = db.execute(
        """
        SELECT
            groupe_id,
            debut_reservation,
            COUNT(objet_id) as item_count
        FROM reservations
        WHERE utilisateur_id = ? AND datetime(debut_reservation) >= datetime(?)
        GROUP BY groupe_id
        ORDER BY debut_reservation ASC
        LIMIT 5
        """, (session['user_id'], now_str)).fetchall()

    dashboard_data['alertes_widget'] = get_alerte_info(db)

    dashboard_data['historique_recent'] = db.execute(
        """
        SELECT h.action, h.timestamp, o.nom as objet_nom, o.id as objet_id
        FROM historique h JOIN objets o ON h.objet_id = o.id
        WHERE h.utilisateur_id = ? ORDER BY h.timestamp DESC LIMIT 5
        """, (session['user_id'], )).fetchall()

        vingt_quatre_heures_avant = datetime.now() - timedelta(hours=24)

    dashboard_data['objets_recents'] = db.execute(
        """
        SELECT o.id, o.nom FROM objets o
        WHERE o.id IN (
            SELECT objet_id FROM historique
            WHERE (action = 'Création' OR (action = 'Modification'
                                           AND details LIKE '%Quantité%'))
            AND timestamp >= ?
            GROUP BY objet_id ORDER BY MAX(timestamp) DESC
        ) LIMIT 10
        """, (vingt_quatre_heures_avant.strftime('%Y-%m-%d %H:%M:%S'), )).fetchall()

    admin_user = db.execute("SELECT nom_utilisateur, email FROM utilisateurs "
                            "WHERE role = 'admin' LIMIT 1").fetchone()
    if admin_user and admin_user['email']:
        dashboard_data['admin_contact'] = admin_user['email']
    elif admin_user:
        dashboard_data['admin_contact'] = admin_user['nom_utilisateur']
    else:
        dashboard_data['admin_contact'] = "Non défini"

    date_limite_echeances = datetime.now().date() + timedelta(days=30)
    date_aujourdhui = datetime.now().date()

    echeances_brutes = db.execute(
        """
        SELECT id, intitule, date_echeance
        FROM echeances
        WHERE traite = 0 AND date_echeance >= ? AND date_echeance <= ?
        ORDER BY date_echeance ASC
        LIMIT 5
        """, (date_aujourdhui.strftime('%Y-%m-%d'), date_limite_echeances.strftime('%Y-%m-%d'))).fetchall()

    prochaines_echeances_calculees = []
    for echeance in echeances_brutes:
        echeance_dict = dict(echeance)
        date_echeance_obj = datetime.strptime(echeance['date_echeance'],
                                              '%Y-%m-%d').date()
        jours_restants = (date_echeance_obj - date_aujourdhui).days

        echeance_dict['date_echeance_obj'] = date_echeance_obj
        echeance_dict['jours_restants'] = jours_restants
        prochaines_echeances_calculees.append(echeance_dict)

    dashboard_data['prochaines_echeances'] = prochaines_echeances_calculees

    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()

    return render_template("index.html",
                           data=dashboard_data,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/inventaire")
@login_required
def inventaire():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort_by', 'nom')
    direction = request.args.get('direction', 'asc')
    search_query = request.args.get('q', None)
    armoire_id = request.args.get('armoire', None)
    categorie_id = request.args.get('categorie', None)
    etat = request.args.get('etat', None)

    objets, total_pages = get_paginated_objets(db,
                                               page,
                                               sort_by=sort_by,
                                               direction=direction,
                                               search_query=search_query,
                                               armoire_id=armoire_id,
                                               categorie_id=categorie_id,
                                               etat=etat)

    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()

    pagination = {
        'page': page,
        'total_pages': total_pages,
        'endpoint': 'inventaire',
        'id': None
    }

    return render_template("inventaire.html",
                           armoires=armoires,
                           categories=categories,
                           objets=objets,
                           pagination=pagination,
                           date_actuelle=datetime.now(),
                           now=datetime.now,
                           sort_by=sort_by,
                           direction=direction)


@app.route("/armoire/<int:id>")
@login_required
def voir_armoire(id):
    db = get_db()
    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort_by', 'nom')
    direction = request.args.get('direction', 'asc')

    armoire = db.execute("SELECT * FROM armoires WHERE id = ?",
                         (id, )).fetchone()
    if not armoire:
        flash("Armoire non trouvée.", "error")
        return redirect(url_for('index'))

    objets, total_pages = get_paginated_objets(db,
                                               page,
                                               sort_by=sort_by,
                                               direction=direction,
                                               filter_field='armoire_id',
                                               filter_id=id)

    armoires_list = db.execute(
        "SELECT * FROM armoires ORDER BY nom").fetchall()
    categories_list = db.execute(
        "SELECT * FROM categories ORDER BY nom").fetchall()

    pagination = {
        'page': page,
        'total_pages': total_pages,
        'endpoint': 'voir_armoire',
        'id': id
    }

    return render_template("armoire.html",
                           armoire=armoire,
                           objets=objets,
                           armoires=armoires_list,
                           categories=categories_list,
                           armoires_list=armoires_list,
                           categories_list=categories_list,
                           pagination=pagination,
                           date_actuelle=datetime.now(),
                           now=datetime.now,
                           sort_by=sort_by,
                           direction=direction)


@app.route("/categorie/<int:id>")
@login_required
def voir_categorie(id):
    db = get_db()
    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort_by', 'nom')
    direction = request.args.get('direction', 'asc')

    categorie = db.execute("SELECT * FROM categories WHERE id = ?",
                           (id, )).fetchone()
    if not categorie:
        flash("Catégorie non trouvée.", "error")
        return redirect(url_for('index'))

    objets, total_pages = get_paginated_objets(db,
                                               page,
                                               sort_by,
                                               direction,
                                               filter_field='categorie_id',
                                               filter_id=id)

    armoires_list = db.execute(
        "SELECT * FROM armoires ORDER BY nom").fetchall()
    categories_list = db.execute(
        "SELECT * FROM categories ORDER BY nom").fetchall()

    pagination = {
        'page': page,
        'total_pages': total_pages,
        'endpoint': 'voir_categorie',
        'id': id
    }

    return render_template("categorie.html",
                           categorie=categorie,
                           objets=objets,
                           armoires=armoires_list,
                           categories=categories_list,
                           armoires_list=armoires_list,
                           categories_list=categories_list,
                           pagination=pagination,
                           date_actuelle=datetime.now(),
                           now=datetime.now,
                           sort_by=sort_by,
                           direction=direction)


@app.route("/objet/<int:objet_id>")
@login_required
def voir_objet(objet_id):
    db = get_db()
    objet = db.execute(
        """SELECT o.*, a.nom as armoire_nom, c.nom as categorie_nom
           FROM objets o
           JOIN armoires a ON o.armoire_id = a.id
           JOIN categories c ON o.categorie_id = c.id
           WHERE o.id = ?""", (objet_id, )).fetchone()
    if not objet:
        flash("Objet non trouvé.", "error")
        return redirect(url_for('index'))

    historique = db.execute(
        "SELECT h.*, u.nom_utilisateur FROM historique h "
        "JOIN utilisateurs u ON h.utilisateur_id = u.id "
        "WHERE h.objet_id = ? ORDER BY h.timestamp DESC",
        (objet_id, )).fetchall()

    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()

    return render_template("objet_details.html",
                           objet=objet,
                           historique=historique,
                           armoires=armoires,
                           categories=categories,
                           date_actuelle=datetime.now(),
                           now=datetime.now)


@app.route("/jour/<string:date_str>")
@login_required
def vue_jour(date_str):
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash("Format de date invalide.", "error")
        return redirect(url_for('calendrier'))

    db = get_db()
    start_of_day = f"{date_str} 00:00:00"
    end_of_day = f"{date_str} 23:59:59"

    reservations_brutes = db.execute(
        """
        SELECT
            r.groupe_id, r.debut_reservation, r.fin_reservation,
            r.utilisateur_id, u.nom_utilisateur, o.id as objet_id, o.nom as nom_objet,
            r.quantite_reservee, r.kit_id, k.nom as kit_nom
        FROM reservations r
        JOIN utilisateurs u ON r.utilisateur_id = u.id
        JOIN objets o ON r.objet_id = o.id
        LEFT JOIN kits k ON r.kit_id = k.id
        WHERE r.debut_reservation >= ? AND r.debut_reservation <= ?
        ORDER BY r.debut_reservation, r.groupe_id
        """, (start_of_day, end_of_day)).fetchall()

    # --- NOUVELLE LOGIQUE ROBUSTE ---
    reservations_groupees = {}
    for row in reservations_brutes:
        groupe_id = row['groupe_id']
        if groupe_id not in reservations_groupees:
            debut_dt = datetime.fromisoformat(row['debut_reservation'])
            fin_dt = datetime.fromisoformat(row['fin_reservation'])
            duree_minutes = (fin_dt - debut_dt).total_seconds() / 60
            reservations_groupees[groupe_id] = {
                'groupe_id': groupe_id,
                'debut_reservation': row['debut_reservation'],
                'fin_reservation': row['fin_reservation'],
                'utilisateur_id': row['utilisateur_id'],
                'nom_utilisateur': row['nom_utilisateur'],
                'duree_minutes': duree_minutes,
                'lignes_reservation': [] # Stockage temporaire des lignes brutes
            }
        reservations_groupees[groupe_id]['lignes_reservation'].append(row)

    # Traitement final pour agréger les données de chaque groupe
    reservations_finales = []
    for groupe_id, data in reservations_groupees.items():
        # On réinitialise les compteurs pour CHAQUE réservation
        data['kits'] = {}
        data['objets_manuels'] = []
        
        objets_manuels_calcul = {r['objet_id']: dict(r) for r in data['lignes_reservation'] if r['kit_id'] is None}
        objets_kits_reserves = [r for r in data['lignes_reservation'] if r['kit_id'] is not None]

        kits_comptes = {}
        for r in objets_kits_reserves:
            if r['kit_id'] not in kits_comptes:
                kits_comptes[r['kit_id']] = {'nom': r['kit_nom'], 'objets_reserves': {}}
            kits_comptes[r['kit_id']]['objets_reserves'][r['objet_id']] = r['quantite_reservee']

        for kit_id, kit_data in kits_comptes.items():
            objets_base_du_kit = db.execute("SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?", (kit_id,)).fetchall()
            if not objets_base_du_kit: continue
            
            id_objet_calcul, quantite_par_kit = next(((obj['objet_id'], obj['quantite']) for obj in objets_base_du_kit if obj['objet_id'] in kit_data['objets_reserves']), (None, 0))

            if id_objet_calcul and quantite_par_kit > 0:
                quantite_reelle_reservee = kit_data['objets_reserves'][id_objet_calcul]
                nombre_de_kits = quantite_reelle_reservee // quantite_par_kit
                
                data['kits'][kit_data['nom']] = data['kits'].get(kit_data['nom'], 0) + nombre_de_kits

                for obj_base in objets_base_du_kit:
                    if obj_base['objet_id'] in objets_manuels_calcul:
                        objets_manuels_calcul[obj_base['objet_id']]['quantite_reservee'] -= nombre_de_kits * obj_base['quantite']

        objets_manuels_agreges = {}
        for obj_id, obj_data in objets_manuels_calcul.items():
            if obj_data['quantite_reservee'] > 0:
                nom = obj_data['nom_objet']
                objets_manuels_agreges[nom] = objets_manuels_agreges.get(nom, 0) + obj_data['quantite_reservee']
        
        data['objets_manuels'] = [f"{qty} x {name}" for name, qty in objets_manuels_agreges.items()]
        
        del data['lignes_reservation'] # Nettoyage
        reservations_finales.append(data)

    return render_template("vue_jour.html",
                           date_concernee=date_obj,
                           reservations=reservations_finales)


# --- ROUTES SÉCURISÉES ET DE GESTION ---
@app.route("/calendrier")
@login_required
def calendrier():
    db = get_db()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("calendrier.html",
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/panier")
@login_required
def panier():
    return render_template("panier.html")


@app.route("/profil", methods=['GET', 'POST'])
@login_required
def profil():
    db = get_db()
    user_id = session['user_id']
    if request.method == 'POST':
        ancien_mdp = request.form.get('ancien_mot_de_passe')
        nouveau_mdp = request.form.get('nouveau_mot_de_passe')
        confirmation_mdp = request.form.get('confirmation_mot_de_passe')
        user = db.execute("SELECT mot_de_passe FROM utilisateurs WHERE id = ?",
                          (user_id, )).fetchone()
        if not user or not check_password_hash(user['mot_de_passe'],
                                               ancien_mdp):
            flash("Votre ancien mot de passe est incorrect.", "error")
            return redirect(url_for('profil'))
        if nouveau_mdp != confirmation_mdp:
            flash(
                "Le nouveau mot de passe et sa confirmation ne correspondent "
                "pas.", "error")
            return redirect(url_for('profil'))
        if len(nouveau_mdp) < 4:
            flash(
                "Le nouveau mot de passe doit contenir au moins 4 caractères.",
                "error")
            return redirect(url_for('profil'))
        try:
            db.execute("UPDATE utilisateurs SET mot_de_passe = ? WHERE id = ?",
                       (generate_password_hash(nouveau_mdp,
                                               method='scrypt'), user_id))
            db.commit()
            flash("Votre mot de passe a été mis à jour avec succès.",
                  "success")
            return redirect(url_for('index'))
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données : {e}", "error")
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("profil.html",
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/alertes")
@login_required
def alertes():
    db = get_db()
    objets_stock = db.execute(
        """
        SELECT o.id, o.nom, o.quantite, o.seuil, a.nom AS armoire,
            c.nom AS categorie, o.image, o.en_commande,
            o.date_peremption, o.traite
        FROM objets o JOIN armoires a ON o.armoire_id = a.id
        JOIN categories c ON o.categorie_id = c.id
        WHERE o.quantite <= seuil ORDER BY o.nom
        """).fetchall()
    date_limite = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
    objets_peremption = db.execute(
        """
        SELECT o.id, o.nom, o.quantite, o.seuil, a.nom AS armoire,
               c.nom AS categorie, o.image, o.en_commande, o.date_peremption,
               o.traite
        FROM objets o JOIN armoires a ON o.armoire_id = a.id
        JOIN categories c ON o.categorie_id = c.id
        WHERE o.date_peremption IS NOT NULL AND o.date_peremption < ?
        ORDER BY o.date_peremption ASC
        """, (date_limite, )).fetchall()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("alertes.html",
                           objets_stock=objets_stock,
                           objets_peremption=objets_peremption,
                           date_actuelle=datetime.now(),
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/gestion_armoires")
@login_required
def gestion_armoires():
    db = get_db()
    armoires = db.execute("""
        SELECT a.id, a.nom, COUNT(o.id) as count FROM armoires a
        LEFT JOIN objets o ON a.id = o.armoire_id
        GROUP BY a.id, a.nom ORDER BY a.nom
        """).fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("gestion_armoires.html",
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/gestion_categories")
@login_required
def gestion_categories():
    db = get_db()
    categories = db.execute("""
        SELECT c.id, c.nom, COUNT(o.id) as count FROM categories c
        LEFT JOIN objets o ON c.id = o.categorie_id
        GROUP BY c.id, c.nom ORDER BY c.nom
        """).fetchall()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    return render_template("gestion_categories.html",
                           categories=categories,
                           armoires=armoires,
                           now=datetime.now)


@app.route("/admin")
@admin_required
def admin():
    db = get_db()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("admin.html",
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)

@app.route("/a-propos")
@login_required
def a_propos():
    return render_template("a_propos.html")
    
@app.route("/admin/activer_licence", methods=["POST"])
@admin_required
def activer_licence():
    licence_cle_fournie = request.form.get('licence_cle', '').strip()
    db = get_db()

    instance_id_row = db.execute(
        "SELECT valeur FROM parametres WHERE cle = 'instance_id'").fetchone()
    if not instance_id_row:
        flash(
            "Erreur critique : Identifiant d'instance manquant. "
            "Impossible de vérifier la licence.", "error")
        return redirect(url_for('admin'))

    instance_id = instance_id_row['valeur']

    chaine_a_verifier = f"{instance_id}-{CLE_PRO_SECRETE}"
    cle_valide_calculee = hashlib.sha256(
        chaine_a_verifier.encode('utf-8')).hexdigest()

    if licence_cle_fournie == cle_valide_calculee[:16]:
        try:
            db.execute("UPDATE parametres SET valeur = 'PRO' "
                       "WHERE cle = 'licence_statut'")
            db.execute(
                "UPDATE parametres SET valeur = ? WHERE cle = 'licence_cle'",
                (licence_cle_fournie, ))
            db.commit()
            flash(
                "Licence Pro activée avec succès ! Toutes les "
                "fonctionnalités sont maintenant débloquées.", "success")
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données lors de l'activation : {e}",
                  "error")
    else:
        flash(
            "La clé de licence fournie est invalide ou ne correspond pas à "
            "cette installation.", "error")

    return redirect(url_for('admin'))


@app.route("/admin/reset_licence", methods=["POST"])
@admin_required
def reset_licence():
    admin_password = request.form.get('admin_password')
    db = get_db()

    admin_user = db.execute(
        "SELECT mot_de_passe FROM utilisateurs WHERE id = ?",
        (session['user_id'], )).fetchone()

    if not admin_user or not check_password_hash(admin_user['mot_de_passe'],
                                                 admin_password):
        flash(
            "Mot de passe administrateur incorrect. "
            "La réinitialisation a été annulée.", "error")
        return redirect(url_for('admin'))

    try:
        db.execute(
            "UPDATE parametres SET valeur = 'FREE' WHERE cle = 'licence_statut'"
        )
        db.execute(
            "UPDATE parametres SET valeur = '' WHERE cle = 'licence_cle'")
        db.commit()
        flash("La licence a été réinitialisée au statut GRATUIT.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données lors de la réinitialisation : {e}",
              "error")

    return redirect(url_for('admin'))


@app.route("/admin/utilisateurs")
@admin_required
def gestion_utilisateurs():
    db = get_db()
    utilisateurs = db.execute(
        "SELECT id, nom_utilisateur, role, email FROM utilisateurs "
        "ORDER BY nom_utilisateur").fetchall()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("admin_utilisateurs.html",
                           utilisateurs=utilisateurs,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/admin/utilisateurs/modifier_email/<int:id_user>",
           methods=["POST"])
@admin_required
def modifier_email_utilisateur(id_user):
    email = request.form.get('email', '').strip()
    if not email or '@' not in email:
        flash("Veuillez fournir une adresse e-mail valide.", "error")
        return redirect(url_for('gestion_utilisateurs'))

    db = get_db()
    user = db.execute("SELECT nom_utilisateur FROM utilisateurs WHERE id = ?",
                      (id_user, )).fetchone()
    if not user:
        flash("Utilisateur non trouvé.", "error")
        return redirect(url_for('gestion_utilisateurs'))

    try:
        db.execute("UPDATE utilisateurs SET email = ? WHERE id = ?",
                   (email, id_user))
        db.commit()
        flash(
            f"L'adresse e-mail pour '{user['nom_utilisateur']}' a été "
            "mise à jour.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('gestion_utilisateurs'))


@app.route("/admin/utilisateurs/supprimer/<int:id_user>", methods=["POST"])
@admin_required
def supprimer_utilisateur(id_user):
    if id_user == session['user_id']:
        flash("Vous ne pouvez pas supprimer votre propre compte.", "error")
        return redirect(url_for('gestion_utilisateurs'))
    db = get_db()
    user = db.execute("SELECT nom_utilisateur FROM utilisateurs WHERE id = ?",
                      (id_user, )).fetchone()
    if user:
        db.execute("DELETE FROM utilisateurs WHERE id = ?", (id_user, ))
        db.commit()
        flash(f"L'utilisateur '{user['nom_utilisateur']}' a été supprimé.",
              "success")
    else:
        flash("Utilisateur non trouvé.", "error")
    return redirect(url_for('gestion_utilisateurs'))


@app.route("/admin/utilisateurs/promouvoir/<int:id_user>", methods=["POST"])
@admin_required
def promouvoir_utilisateur(id_user):
    if id_user == session['user_id']:
        flash("Action non autorisée sur votre propre compte.", "error")
        return redirect(url_for('gestion_utilisateurs'))
    password = request.form.get('password')
    db = get_db()
    admin_actuel = db.execute(
        "SELECT mot_de_passe FROM utilisateurs WHERE id = ?",
        (session['user_id'], )).fetchone()
    if not admin_actuel or not check_password_hash(
            admin_actuel['mot_de_passe'], password):
        flash(
            "Mot de passe administrateur incorrect. "
            "La passation de pouvoir a échoué.", "error")
        return redirect(url_for('gestion_utilisateurs'))
    try:
        db.execute("UPDATE utilisateurs SET role = 'admin' WHERE id = ?",
                   (id_user, ))
        db.execute("UPDATE utilisateurs SET role = 'utilisateur' WHERE id = ?",
                   (session['user_id'], ))
        db.commit()
        flash(
            "Passation de pouvoir réussie ! "
            "Vous êtes maintenant un utilisateur standard.", "success")
        return redirect(url_for('logout'))
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Une erreur est survenue lors de la passation de pouvoir : {e}",
              "error")
        return redirect(url_for('gestion_utilisateurs'))


@app.route("/admin/utilisateurs/reinitialiser_mdp/<int:id_user>",
           methods=["POST"])
@admin_required
def reinitialiser_mdp(id_user):
    if id_user == session['user_id']:
        flash(
            "Vous ne pouvez pas réinitialiser votre propre mot de passe ici.",
            "error")
        return redirect(url_for('gestion_utilisateurs'))
    nouveau_mdp = request.form.get('nouveau_mot_de_passe')
    if not nouveau_mdp or len(nouveau_mdp) < 4:
        flash(
            "Le nouveau mot de passe est requis et doit contenir "
            "au moins 4 caractères.", "error")
        return redirect(url_for('gestion_utilisateurs'))
    db = get_db()
    user = db.execute("SELECT nom_utilisateur FROM utilisateurs WHERE id = ?",
                      (id_user, )).fetchone()
    if user:
        try:
            db.execute("UPDATE utilisateurs SET mot_de_passe = ? WHERE id = ?",
                       (generate_password_hash(nouveau_mdp,
                                               method='scrypt'), id_user))
            db.commit()
            flash(
                f"Le mot de passe pour l'utilisateur "
                f"'{user['nom_utilisateur']}' a été réinitialisé avec succès.",
                "success")
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données : {e}", "error")
    else:
        flash("Utilisateur non trouvé.", "error")
    return redirect(url_for('gestion_utilisateurs'))


# --- ROUTES POUR LA GESTION DES KITS ---
@app.route("/admin/kits")
@admin_required
def gestion_kits():
    db = get_db()
    kits = db.execute("""
        SELECT k.id, k.nom, k.description, COUNT(ko.id) as count
        FROM kits k
        LEFT JOIN kit_objets ko ON k.id = ko.kit_id
        GROUP BY k.id, k.nom, k.description
        ORDER BY k.nom
        """).fetchall()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("admin_kits.html",
                           kits=kits,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/admin/kits/ajouter", methods=["POST"])
@admin_required
def ajouter_kit():
    nom = request.form.get("nom", "").strip()
    description = request.form.get("description", "").strip()
    if not nom:
        flash("Le nom du kit ne peut pas être vide.", "error")
        return redirect(url_for('gestion_kits'))

    db = get_db()
    try:
        cursor = db.execute(
            "INSERT INTO kits (nom, description) VALUES (?, ?)",
            (nom, description))
        db.commit()
        new_kit_id = cursor.lastrowid
        flash(
            f"Le kit '{nom}' a été créé. "
            "Vous pouvez maintenant y ajouter des objets.", "success")
        return redirect(url_for('modifier_kit', kit_id=new_kit_id))
    except sqlite3.IntegrityError:
        flash(f"Un kit avec le nom '{nom}' existe déjà.", "error")
        return redirect(url_for('gestion_kits'))


@app.route("/admin/kits/modifier/<int:kit_id>", methods=["GET", "POST"])
@admin_required
def modifier_kit(kit_id):
    db = get_db()
    kit = db.execute("SELECT * FROM kits WHERE id = ?", (kit_id, )).fetchone()
    if not kit:
        flash("Kit non trouvé.", "error")
        return redirect(url_for('gestion_kits'))

    if request.method == "POST":
        objet_id_str = request.form.get("objet_id")
        quantite_str = request.form.get("quantite")

        # --- LOGIQUE D'AJOUT D'UN NOUVEL OBJET AU KIT ---
        if objet_id_str and quantite_str:
            try:
                objet_id = int(objet_id_str)
                quantite = int(quantite_str)

                # Vérification du stock disponible
                objet_stock = db.execute("SELECT nom, quantite FROM objets WHERE id = ?", (objet_id,)).fetchone()
                if not objet_stock:
                    flash("Objet non trouvé.", "error")
                    return redirect(url_for('modifier_kit', kit_id=kit_id))

                if quantite > objet_stock['quantite']:
                    flash(f"Quantité invalide pour '{objet_stock['nom']}'. Vous ne pouvez pas ajouter plus que le stock disponible ({objet_stock['quantite']}).", "error")
                    return redirect(url_for('modifier_kit', kit_id=kit_id))

                # Insertion ou mise à jour dans le kit
                existing = db.execute(
                    "SELECT id FROM kit_objets WHERE kit_id = ? AND objet_id = ?",
                    (kit_id, objet_id)).fetchone()
                if existing:
                    db.execute("UPDATE kit_objets SET quantite = ? WHERE id = ?",
                               (quantite, existing['id']))
                else:
                    db.execute(
                        "INSERT INTO kit_objets (kit_id, objet_id, quantite) "
                        "VALUES (?, ?, ?)", (kit_id, objet_id, quantite))
                db.commit()
                flash(f"L'objet '{objet_stock['nom']}' a été ajouté/mis à jour dans le kit.", "success")

            except (ValueError, TypeError):
                flash("Données invalides.", "error")
            
            return redirect(url_for('modifier_kit', kit_id=kit_id))

        # --- LOGIQUE DE MISE À JOUR DES QUANTITÉS EXISTANTES ---
        for key, value in request.form.items():
            if key.startswith("quantite_"):
                try:
                    kit_objet_id = int(key.split("_")[1])
                    new_quantite = int(value)

                    # Vérification du stock pour la mise à jour
                    objet_info = db.execute("""
                        SELECT o.nom, o.quantite FROM kit_objets ko
                        JOIN objets o ON ko.objet_id = o.id
                        WHERE ko.id = ?
                    """, (kit_objet_id,)).fetchone()

                    if not objet_info:
                        continue # Passe au suivant si l'objet n'est pas trouvé

                    if new_quantite > objet_info['quantite']:
                         flash(f"Quantité invalide pour '{objet_info['nom']}'. Vous ne pouvez pas dépasser le stock disponible ({objet_info['quantite']}).", "error")
                    else:
                        db.execute("UPDATE kit_objets SET quantite = ? WHERE id = ?",
                                   (new_quantite, kit_objet_id))
                        flash(f"Quantité pour '{objet_info['nom']}' mise à jour.", "success")
                
                except (ValueError, TypeError):
                    flash("Une quantité fournie est invalide.", "error")
        
        db.commit()
        return redirect(url_for('modifier_kit', kit_id=kit_id))

    # --- RÉCUPÉRATION DES DONNÉES POUR L'AFFICHAGE (GET) ---
    objets_in_kit = db.execute(
        """
        SELECT ko.id, o.nom, o.quantite as stock_disponible, ko.quantite
        FROM kit_objets ko
        JOIN objets o ON ko.objet_id = o.id
        WHERE ko.kit_id = ?
        ORDER BY o.nom
        """, (kit_id, )).fetchall()

    objets_disponibles = db.execute(
        """
        SELECT id, nom, quantite FROM objets
        WHERE id NOT IN (SELECT objet_id FROM kit_objets WHERE kit_id = ?)
        ORDER BY nom
        """, (kit_id, )).fetchall()

    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()

    return render_template("admin_kit_modifier.html",
                           kit=kit,
                           objets_in_kit=objets_in_kit,
                           objets_disponibles=objets_disponibles,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/admin/kits/retirer_objet/<int:kit_objet_id>")
@admin_required
def retirer_objet_kit(kit_objet_id):
    db = get_db()
    kit_objet = db.execute("SELECT kit_id FROM kit_objets WHERE id = ?",
                           (kit_objet_id, )).fetchone()
    if kit_objet:
        kit_id = kit_objet['kit_id']
        db.execute("DELETE FROM kit_objets WHERE id = ?", (kit_objet_id, ))
        db.commit()
        flash("Objet retiré du kit.", "success")
        return redirect(url_for('modifier_kit', kit_id=kit_id))
    flash("Erreur : objet du kit non trouvé.", "error")
    return redirect(url_for('gestion_kits'))


@app.route("/admin/kits/supprimer/<int:kit_id>", methods=["POST"])
@admin_required
def supprimer_kit(kit_id):
    db = get_db()
    kit = db.execute("SELECT nom FROM kits WHERE id = ?",
                     (kit_id, )).fetchone()
    if kit:
        db.execute("DELETE FROM kits WHERE id = ?", (kit_id, ))
        db.commit()
        flash(f"Le kit '{kit['nom']}' a été supprimé.", "success")
    else:
        flash("Kit non trouvé.", "error")
    return redirect(url_for('gestion_kits'))


def generer_rapport_pdf(data, date_debut, date_fin, group_by):
    pdf = PDFWithFooter()
    pdf.alias_nb_pages()
    pdf.add_page(orientation='L')
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Rapport d\'Activite', 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(
        0, 10,
        f"Periode du {datetime.strptime(date_debut, '%Y-%m-%d').strftime('%d/%m/%Y')} "
        f"au {datetime.strptime(date_fin, '%Y-%m-%d').strftime('%d/%m/%Y')}",
        0, 1, 'C')
    pdf.ln(5)

    col_widths = {
        "date": 25,
        "heure": 15,
        "user": 35,
        "action": 60,
        "objet": 60,
        "details": 75
    }
    line_height = 6
    table_width = sum(col_widths.values())

    def draw_header():
        pdf.set_font('Arial', 'B', 9)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(col_widths["date"], 8, 'Date', 1, 0, 'C', 1)
        pdf.cell(col_widths["heure"], 8, 'Heure', 1, 0, 'C', 1)
        pdf.cell(col_widths["user"], 8, 'Utilisateur', 1, 0, 'C', 1)
        pdf.cell(col_widths["action"], 8, 'Action', 1, 0, 'C', 1)
        pdf.cell(col_widths["objet"], 8, 'Objet Concerne', 1, 0, 'C', 1)
        pdf.cell(col_widths["details"], 8, 'Details', 1, 1, 'C', 1)

    draw_header()

    current_group = None
    last_date_str = None

    for i, item in enumerate(data):
        details_for_calc = item['details'].encode('latin-1',
                                                  'replace').decode('latin-1')
        estimated_height = (len(details_for_calc) // 40 + 1) * line_height
        if pdf.get_y() + estimated_height > 190:
            pdf.add_page(orientation='L')
            draw_header()
            last_date_str = None

        if group_by == 'action' and item['action'] != current_group:
            current_group = item['action']
            last_date_str = None
            if i > 0:
                pdf.add_page(orientation='L')
                draw_header()
            pdf.set_font('Arial', 'B', 10)
            pdf.set_fill_color(230, 240, 255)
            pdf.cell(table_width, 8, f"Type d'action : {current_group}", 1, 1,
                     'L', 1)

        pdf.set_font('Arial', '', 8)

        timestamp_dt = datetime.fromisoformat(item['timestamp'])
        current_date_str = timestamp_dt.strftime('%d/%m/%Y')
        heure_str = timestamp_dt.strftime('%H:%M')

        date_str_display = ''
        border_date = 'LR'
        if group_by == 'date' and current_date_str != last_date_str:
            date_str_display = current_date_str
            border_date = 'LTR'

        is_last_in_date_group = (i + 1 >= len(data) or datetime.fromisoformat(
            data[i + 1]['timestamp']).strftime('%d/%m/%Y') != current_date_str)
        if group_by == 'date' and is_last_in_date_group:
            border_date = 'LBR' if date_str_display == '' else 'LRTB'

        utilisateur = item['nom_utilisateur'].encode(
            'latin-1', 'replace').decode('latin-1')
        action = item['action'].encode('latin-1', 'replace').decode('latin-1')
        objet_nom = item['objet_nom'].encode('latin-1',
                                             'replace').decode('latin-1')
        details = item['details'].encode('latin-1',
                                         'replace').decode('latin-1')

        with pdf.unbreakable() as doc:
            doc.multi_cell(col_widths["date"],
                           line_height,
                           date_str_display,
                           border=border_date,
                           align='C',
                           max_line_height=line_height,
                           new_x="RIGHT",
                           new_y="TOP")
            doc.multi_cell(col_widths["heure"],
                           line_height,
                           heure_str,
                           border=1,
                           align='C',
                           max_line_height=line_height,
                           new_x="RIGHT",
                           new_y="TOP")
            doc.multi_cell(col_widths["user"],
                           line_height,
                           utilisateur,
                           border=1,
                           align='C',
                           max_line_height=line_height,
                           new_x="RIGHT",
                           new_y="TOP")
            doc.multi_cell(col_widths["action"],
                           line_height,
                           action,
                           border=1,
                           align='C',
                           max_line_height=line_height,
                           new_x="RIGHT",
                           new_y="TOP")
            doc.multi_cell(col_widths["objet"],
                           line_height,
                           objet_nom,
                           border=1,
                           align='L',
                           max_line_height=line_height,
                           new_x="RIGHT",
                           new_y="TOP")
            doc.multi_cell(col_widths["details"],
                           line_height,
                           details,
                           border=1,
                           align='L',
                           max_line_height=line_height,
                           new_x="LMARGIN",
                           new_y="NEXT")

        last_date_str = current_date_str

    return BytesIO(pdf.output())


def generer_rapport_excel(data, date_debut, date_fin, group_by):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rapport d'Activite"

    title_font = Font(name='Calibri', size=16, bold=True)
    subtitle_font = Font(name='Calibri', size=11, italic=True, color="6c7a89")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568",
                              end_color="4A5568",
                              fill_type="solid")
    group_font = Font(name='Calibri', size=11, bold=True, color="2D3748")
    group_fill = PatternFill(start_color="E2E8F0",
                             end_color="E2E8F0",
                             fill_type="solid")
    even_row_fill = PatternFill(start_color="F7FAFC",
                                end_color="F7FAFC",
                                fill_type="solid")

    center_align = Alignment(horizontal='center',
                             vertical='center',
                             wrap_text=True)
    left_align = Alignment(horizontal='left',
                           vertical='center',
                           wrap_text=True)

    thin_border_side = Side(style='thin', color="4A5568")
    thick_border_side = Side(style='medium', color="000000")

    cell_border = Border(left=thin_border_side,
                         right=thin_border_side,
                         top=thin_border_side,
                         bottom=thin_border_side)

    start_col = 2
    headers = [
        "Date", "Heure", "Utilisateur", "Action", "Objet Concerne", "Details"
    ]
    end_col = start_col + len(headers) - 1

    sheet.merge_cells(start_row=2,
                      start_column=start_col,
                      end_row=2,
                      end_column=end_col)
    title_cell = sheet.cell(row=2, column=start_col)
    title_cell.value = "Rapport d'Activite"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    sheet.merge_cells(start_row=3,
                      start_column=start_col,
                      end_row=3,
                      end_column=end_col)
    subtitle_cell = sheet.cell(row=3, column=start_col)
    subtitle_cell.value = (
        f"Periode du "
        f"{datetime.strptime(date_debut, '%Y-%m-%d').strftime('%d/%m/%Y')} au "
        f"{datetime.strptime(date_fin, '%Y-%m-%d').strftime('%d/%m/%Y')}")
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    header_row = 5
    for i, header_text in enumerate(headers, start=start_col):
        cell = sheet.cell(row=header_row, column=i, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border

    row_index = header_row + 1
    current_group = None
    last_date_str = None
    is_even = False

    for item in data:
        if group_by == 'action' and item['action'] != current_group:
            current_group = item['action']
            last_date_str = None
            cell = sheet.cell(row=row_index,
                              column=start_col,
                              value=f"Type d'action : {current_group}")
            sheet.merge_cells(start_row=row_index,
                              start_column=start_col,
                              end_row=row_index,
                              end_column=end_col)
            for c in range(start_col, end_col + 1):
                sheet.cell(row=row_index, column=c).fill = group_fill
                sheet.cell(row=row_index, column=c).font = group_font
                sheet.cell(row=row_index, column=c).border = cell_border
            row_index += 1
            is_even = False

        timestamp_dt = datetime.fromisoformat(item['timestamp'])
        current_date_str = timestamp_dt.strftime('%d/%m/%Y')

        date_str_display = current_date_str
        if group_by == 'date' and current_date_str == last_date_str:
            date_str_display = ""

        row_data = [
            date_str_display,
            timestamp_dt.strftime('%H:%M'), item['nom_utilisateur'],
            item['action'], item['objet_nom'], item['details']
        ]

        for col_index, value in enumerate(row_data, start=start_col):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = cell_border
            if col_index in [
                    start_col, start_col + 1, start_col + 2, start_col + 3
            ]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            if is_even:
                cell.fill = even_row_fill

        last_date_str = current_date_str
        row_index += 1
        is_even = not is_even

    end_row_index = row_index - 1
    if end_row_index >= header_row:
        for row in sheet.iter_rows(min_row=header_row,
                                   max_row=end_row_index,
                                   min_col=start_col,
                                   max_col=end_col):
            for cell in row:
                new_border = Border(left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom)
                if cell.row == header_row:
                    new_border.top = thick_border_side
                if cell.row == end_row_index:
                    new_border.bottom = thick_border_side
                if cell.column == start_col:
                    new_border.left = thick_border_side
                if cell.column == end_col:
                    new_border.right = thick_border_side
                cell.border = new_border

    column_widths = {'B': 15, 'C': 10, 'D': 25, 'E': 50, 'F': 40, 'G': 60}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=start_col)

    if end_row_index >= header_row:
        sheet.auto_filter.ref = (
            f"{sheet.cell(row=header_row, column=start_col).coordinate}:"
            f"{sheet.cell(row=end_row_index, column=end_col).coordinate}")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@app.route("/fournisseurs")
@login_required
def voir_fournisseurs():
    db = get_db()
    fournisseurs = db.execute(
        "SELECT * FROM fournisseurs ORDER BY nom").fetchall()
    return render_template("fournisseurs.html", fournisseurs=fournisseurs)


@app.route("/admin/fournisseurs", methods=['GET', 'POST'])
@admin_required
def gestion_fournisseurs():
    db = get_db()
    if request.method == 'POST':
        nom = request.form.get('nom', '').strip()
        site_web = request.form.get('site_web', '').strip()
        logo_name = None

        if not nom:
            flash("Le nom du fournisseur est obligatoire.", "error")
            return redirect(url_for('gestion_fournisseurs'))

        if 'logo' in request.files:
            logo = request.files['logo']
            if logo and logo.filename != '':
                filename = secure_filename(logo.filename)
                logo.save(os.path.join('static/images/fournisseurs', filename))
                logo_name = filename

        try:
            db.execute(
                "INSERT INTO fournisseurs (nom, site_web, logo) "
                "VALUES (?, ?, ?)", (nom, site_web or None, logo_name))
            db.commit()
            flash(f"Le fournisseur '{nom}' a été ajouté.", "success")
        except sqlite3.IntegrityError:
            flash(f"Un fournisseur avec le nom '{nom}' existe déjà.", "error")
        except sqlite3.Error as e:
            flash(f"Erreur de base de données : {e}", "error")
        return redirect(url_for('gestion_fournisseurs'))

    fournisseurs = db.execute(
        "SELECT * FROM fournisseurs ORDER BY nom").fetchall()
    return render_template("admin_fournisseurs.html",
                           fournisseurs=fournisseurs)


@app.route("/admin/fournisseurs/supprimer/<int:id>", methods=['POST'])
@admin_required
def supprimer_fournisseur(id):
    db = get_db()
    fournisseur = db.execute("SELECT logo, nom FROM fournisseurs WHERE id = ?",
                             (id, )).fetchone()
    if fournisseur:
        if fournisseur['logo']:
            try:
                os.remove(
                    os.path.join('static/images/fournisseurs',
                                 fournisseur['logo']))
            except OSError:
                pass

        db.execute("DELETE FROM fournisseurs WHERE id = ?", (id, ))
        db.commit()
        flash(f"Le fournisseur '{fournisseur['nom']}' a été supprimé.",
              "success")
    else:
        flash("Fournisseur non trouvé.", "error")
    return redirect(url_for('gestion_fournisseurs'))


@app.route("/admin/fournisseurs/modifier/<int:id>", methods=['POST'])
@admin_required
def modifier_fournisseur(id):
    db = get_db()
    fournisseur_avant = db.execute("SELECT * FROM fournisseurs WHERE id = ?",
                                   (id, )).fetchone()
    if not fournisseur_avant:
        flash("Fournisseur non trouvé.", "error")
        return redirect(url_for('gestion_fournisseurs'))

    nom = request.form.get('nom', '').strip()
    site_web = request.form.get('site_web', '').strip()
    logo_name = fournisseur_avant['logo']

    if not nom:
        flash("Le nom du fournisseur est obligatoire.", "error")
        return redirect(url_for('gestion_fournisseurs'))

    if request.form.get('supprimer_logo'):
        if logo_name:
            try:
                os.remove(os.path.join('static/images/fournisseurs',
                                       logo_name))
            except OSError:
                pass
        logo_name = None
    elif 'logo' in request.files:
        nouveau_logo = request.files['logo']
        if nouveau_logo and nouveau_logo.filename != '':
            if logo_name:
                try:
                    os.remove(
                        os.path.join('static/images/fournisseurs', logo_name))
                except OSError:
                    pass
            filename = secure_filename(nouveau_logo.filename)
            nouveau_logo.save(
                os.path.join('static/images/fournisseurs', filename))
            logo_name = filename

    try:
        db.execute(
            "UPDATE fournisseurs SET nom = ?, site_web = ?, logo = ? "
            "WHERE id = ?", (nom, site_web or None, logo_name, id))
        db.commit()
        flash(f"Le fournisseur '{nom}' a été mis à jour.", "success")
    except sqlite3.IntegrityError:
        flash(f"Un autre fournisseur avec le nom '{nom}' existe déjà.",
              "error")
    except sqlite3.Error as e:
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('gestion_fournisseurs'))


# --- ROUTES POUR LA GESTION DU BUDGET ---
@app.route("/budget/voir")
@login_required
def voir_budget():
    db = get_db()
    now = datetime.now()
    
    # Logique de l'année scolaire : commence en septembre.
    annee_scolaire_actuelle = now.year if now.month >= 9 else now.year - 1
    
    budget_actuel = db.execute(
        "SELECT * FROM budgets WHERE annee = ? AND cloture = 0",
        (annee_scolaire_actuelle, )).fetchone()

    depenses = []
    total_depenses = 0
    solde = 0

    if budget_actuel:
        depenses = db.execute(
            """SELECT d.id, d.contenu, d.montant, d.date_depense,
                      d.est_bon_achat, d.fournisseur_id, f.nom as fournisseur_nom
            FROM depenses d
            LEFT JOIN fournisseurs f ON d.fournisseur_id = f.id
            WHERE d.budget_id = ?
            ORDER BY d.date_depense DESC""",
            (budget_actuel['id'], )).fetchall()

        total_depenses_result = db.execute(
            "SELECT SUM(montant) as total FROM depenses WHERE budget_id = ?",
            (budget_actuel['id'], )).fetchone()
        total_depenses = (total_depenses_result['total']
                          if total_depenses_result['total'] is not None else 0)
        solde = budget_actuel['montant_initial'] - total_depenses

    return render_template("budget_voir.html",
                           budget_actuel=budget_actuel,
                           depenses=depenses,
                           total_depenses=total_depenses,
                           solde=solde)


@app.route("/budget", methods=['GET'])
@admin_required
def budget():
    db = get_db()
    now = datetime.now()
    
    # CORRECTION : Le basculement se fait maintenant en août (mois 8)
    annee_scolaire_actuelle = now.year if now.month >= 8 else now.year - 1

    budgets_archives = db.execute(
        "SELECT annee FROM budgets ORDER BY annee DESC").fetchall()

    annee_a_afficher_str = request.args.get('annee', type=str)
    
    if annee_a_afficher_str:
        annee_a_afficher = int(annee_a_afficher_str)
    else:
        # Par défaut, on affiche l'année scolaire en cours
        annee_a_afficher = annee_scolaire_actuelle

    budget_a_afficher = db.execute("SELECT * FROM budgets WHERE annee = ?",
                                   (annee_a_afficher, )).fetchone()

    # Si aucun budget n'existe pour l'année à afficher (cas du premier lancement), on en crée un vide
    if not budget_a_afficher and not budgets_archives:
        try:
            cursor = db.execute(
                "INSERT INTO budgets (annee, montant_initial, cloture) VALUES (?, ?, 0)",
                (annee_a_afficher, 0.0)
            )
            db.commit()
            budget_id = cursor.lastrowid
            budget_a_afficher = db.execute("SELECT * FROM budgets WHERE id = ?", (budget_id,)).fetchone()
            budgets_archives = db.execute("SELECT annee FROM budgets ORDER BY annee DESC").fetchall()
        except sqlite3.IntegrityError:
            db.rollback()
            budget_a_afficher = db.execute("SELECT * FROM budgets WHERE annee = ?", (annee_a_afficher,)).fetchone()

    depenses = []
    total_depenses = 0
    solde = 0
    cloture_autorisee = False

    if budget_a_afficher:
        depenses = db.execute(
            """SELECT d.id, d.contenu, d.montant, d.date_depense,
                      d.est_bon_achat, d.fournisseur_id, f.nom as fournisseur_nom
               FROM depenses d
               LEFT JOIN fournisseurs f ON d.fournisseur_id = f.id
               WHERE d.budget_id = ?
               ORDER BY d.date_depense DESC""",
            (budget_a_afficher['id'], )).fetchall()

        total_depenses_result = db.execute(
            "SELECT SUM(montant) as total FROM depenses WHERE budget_id = ?",
            (budget_a_afficher['id'], )).fetchone()
        total_depenses = (total_depenses_result['total'] if total_depenses_result['total'] is not None else 0)
        solde = budget_a_afficher['montant_initial'] - total_depenses

        # Logique de sécurité pour la clôture
        annee_fin_budget = budget_a_afficher['annee'] + 1
        date_limite_cloture = date(annee_fin_budget, 6, 1) # 1er Juin de l'année N+1
        if date.today() >= date_limite_cloture:
            cloture_autorisee = True

    budget_actuel_pour_modales = db.execute(
        "SELECT * FROM budgets WHERE annee = ? AND cloture = 0",
        (annee_scolaire_actuelle, )).fetchone()

    annee_proposee_pour_creation = annee_scolaire_actuelle
    if not budget_actuel_pour_modales:
        derniere_annee_budget = db.execute("SELECT MAX(annee) as max_annee FROM budgets").fetchone()
        if derniere_annee_budget and derniere_annee_budget['max_annee'] is not None:
            annee_proposee_pour_creation = derniere_annee_budget['max_annee'] + 1
        else:
            annee_proposee_pour_creation = annee_scolaire_actuelle

    fournisseurs = db.execute(
        "SELECT id, nom FROM fournisseurs ORDER BY nom").fetchall()

    return render_template(
        "budget.html",
        budget_affiche=budget_a_afficher,
        budget_actuel_pour_modales=budget_actuel_pour_modales,
        annee_proposee_pour_creation=annee_proposee_pour_creation,
        depenses=depenses,
        total_depenses=total_depenses,
        solde=solde,
        fournisseurs=fournisseurs,
        budgets_archives=budgets_archives,
        annee_selectionnee=annee_a_afficher,
        cloture_autorisee=cloture_autorisee,
        now=datetime.now)


@app.route("/budget/definir", methods=['POST'])
@admin_required
def definir_budget():
    db = get_db()
    montant = request.form.get('montant_initial')
    annee = request.form.get('annee')

    if not montant or not annee:
        flash("L'année et le montant sont obligatoires.", "error")
        return redirect(url_for('budget'))

    try:
        montant_float = float(montant.replace(',', '.'))
        annee_int = int(annee)

        existing_budget = db.execute(
            "SELECT id FROM budgets WHERE annee = ?", (annee_int,)
        ).fetchone()
        if existing_budget:
            db.execute(
                "UPDATE budgets SET montant_initial = ?, cloture = 0 WHERE id = ?",
                (montant_float, existing_budget['id'])
            )
        else:
            db.execute(
                "INSERT INTO budgets (annee, montant_initial) VALUES (?, ?)",
                (annee_int, montant_float)
            )

        db.commit()
        flash(
            f"Le budget pour l'année scolaire {annee_scolaire_format(annee_int)} a été défini à "
            f"{montant_float:.2f} €.", "success"
        )
    except ValueError:
        flash("Le montant ou l'année saisi(e) est invalide.", "error")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('budget', annee=annee))


@app.route("/budget/ajouter_depense", methods=['POST'])
@admin_required
def ajouter_depense():
    db = get_db()
    budget_id = request.form.get('budget_id')
    fournisseur_id = request.form.get('fournisseur_id')
    contenu = request.form.get('contenu', '').strip()
    montant = request.form.get('montant')
    date_depense = request.form.get('date_depense')
    est_bon_achat = 1 if request.form.get('est_bon_achat') == 'on' else 0

    if not all([budget_id, contenu, montant, date_depense]):
        flash("Tous les champs sont obligatoires pour ajouter une dépense.",
              "error")
        return redirect(url_for('budget'))

    if est_bon_achat:
        fournisseur_id = None
    elif not fournisseur_id:
        flash(
            "Veuillez sélectionner un fournisseur ou cocher la case "
            "'Bon d'achat'.", "error")
        return redirect(url_for('budget'))

    try:
        montant_float = float(montant.replace(',', '.'))
        db.execute(
            """INSERT INTO depenses (budget_id, fournisseur_id, contenu,
               montant, date_depense, est_bon_achat)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (budget_id, fournisseur_id, contenu, montant_float, date_depense,
             est_bon_achat))
        db.commit()
        flash("La dépense a été ajoutée avec succès.", "success")
    except ValueError:
        flash("Le montant saisi est invalide.", "error")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('budget'))


@app.route("/budget/modifier_depense/<int:id>", methods=['POST'])
@admin_required
def modifier_depense(id):
    db = get_db()
    depense = db.execute("SELECT id FROM depenses WHERE id = ?",
                         (id, )).fetchone()
    if not depense:
        flash("Dépense non trouvée.", "error")
        return redirect(url_for('budget'))

    fournisseur_id = request.form.get('fournisseur_id')
    contenu = request.form.get('contenu', '').strip()
    montant = request.form.get('montant')
    date_depense = request.form.get('date_depense')
    est_bon_achat = 1 if request.form.get('est_bon_achat') == 'on' else 0

    if not all([contenu, montant, date_depense]):
        flash("Les champs contenu, montant et date sont obligatoires.",
              "error")
        return redirect(request.referrer or url_for('budget'))

    if est_bon_achat:
        fournisseur_id = None
    elif not fournisseur_id:
        flash(
            "Veuillez sélectionner un fournisseur ou cocher la case "
            "'Bon d'achat'.", "error")
        return redirect(request.referrer or url_for('budget'))

    try:
        montant_float = float(montant.replace(',', '.'))
        db.execute(
            """UPDATE depenses SET fournisseur_id = ?, contenu = ?,
               montant = ?, date_depense = ?, est_bon_achat = ?
               WHERE id = ?""", (fournisseur_id, contenu, montant_float,
                                 date_depense, est_bon_achat, id))
        db.commit()
        flash("La dépense a été modifiée avec succès.", "success")
    except ValueError:
        flash("Le montant saisi est invalide.", "error")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(request.referrer or url_for('budget'))


@app.route("/budget/supprimer_depense/<int:id>", methods=['POST'])
@admin_required
def supprimer_depense(id):
    db = get_db()
    depense = db.execute("SELECT id FROM depenses WHERE id = ?",
                         (id, )).fetchone()
    if depense:
        try:
            db.execute("DELETE FROM depenses WHERE id = ?", (id, ))
            db.commit()
            flash("La dépense a été supprimée avec succès.", "success")
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données : {e}", "error")
    else:
        flash("Dépense non trouvée.", "error")

    return redirect(request.referrer or url_for('budget'))


@app.route("/budget/cloturer", methods=['POST'])
@admin_required
def cloturer_budget():
    budget_id = request.form.get('budget_id')
    db = get_db()
    
    budget = db.execute("SELECT * FROM budgets WHERE id = ?", (budget_id,)).fetchone()

    if not budget:
        flash("Budget non trouvé.", "error")
        return redirect(url_for('budget'))

    # --- VÉRIFICATION DE SÉCURITÉ CÔTÉ SERVEUR ---
    annee_fin_budget = budget['annee'] + 1
    date_limite_cloture = date(annee_fin_budget, 6, 1)
    if date.today() < date_limite_cloture:
        flash(f"La clôture du budget {annee_scolaire_format(budget['annee'])} n'est autorisée qu'à partir du {date_limite_cloture.strftime('%d/%m/%Y')}.", "error")
        return redirect(url_for('budget', annee=budget['annee']))

    if budget['cloture']:
        flash(f"Le budget pour l'année scolaire {annee_scolaire_format(budget['annee'])} est déjà clôturé.", "warning")
        return redirect(url_for('budget'))

    try:
        db.execute("UPDATE budgets SET cloture = 1 WHERE id = ?", (budget_id,))
        db.commit()
        flash(f"Le budget pour l'année scolaire {annee_scolaire_format(budget['annee'])} a été clôturé avec succès.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('budget'))


@app.route("/budget/exporter")
@admin_required
def exporter_budget():
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    format_type = request.args.get('format')

    if not all([date_debut, date_fin, format_type]):
        flash("Tous les champs sont requis pour l'export.", "error")
        return redirect(url_for('budget'))

    db = get_db()
    depenses_data = db.execute(
        """
        SELECT d.date_depense, d.contenu, d.montant, f.nom as fournisseur_nom
        FROM depenses d
        LEFT JOIN fournisseurs f ON d.fournisseur_id = f.id
        WHERE d.date_depense BETWEEN ? AND ?
        ORDER BY d.date_depense ASC
        """, (date_debut, date_fin)).fetchall()

    if not depenses_data:
        flash("Aucune dépense trouvée pour la période sélectionnée.",
              "warning")
        return redirect(url_for('budget'))

    if format_type == 'pdf':
        buffer = generer_budget_pdf(depenses_data, date_debut, date_fin)
        return send_file(buffer,
                         as_attachment=True,
                         download_name='rapport_depenses.pdf',
                         mimetype='application/pdf')
    elif format_type == 'excel':
        buffer = generer_budget_excel(depenses_data, date_debut, date_fin)
        return send_file(
            buffer,
            as_attachment=True,
            download_name='rapport_depenses.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    flash("Format d'exportation non valide.", "error")
    return redirect(url_for('budget'))


# --- ROUTES POUR LA GESTION DES ÉCHÉANCES ---
@app.route("/admin/echeances")
@admin_required
def gestion_echeances():
    db = get_db()
    echeances_brutes = db.execute(
        "SELECT * FROM echeances ORDER BY date_echeance ASC").fetchall()

    echeances_converties = []
    for echeance in echeances_brutes:
        echeance_dict = dict(echeance)
        echeance_dict['date_echeance'] = datetime.strptime(
            echeance['date_echeance'], '%Y-%m-%d').date()
        echeances_converties.append(echeance_dict)

    return render_template("admin_echeances.html",
                           echeances=echeances_converties,
                           date_actuelle=datetime.now().date(),
                           url_ajout=url_for('ajouter_echeance'))


@app.route("/admin/echeances/ajouter", methods=['POST'])
@admin_required
def ajouter_echeance():
    intitule = request.form.get('intitule', '').strip()
    date_echeance = request.form.get('date_echeance')
    details = request.form.get('details', '').strip()

    if not all([intitule, date_echeance]):
        flash("L'intitulé et la date d'échéance sont obligatoires.", "error")
        return redirect(url_for('gestion_echeances'))

    db = get_db()
    try:
        db.execute(
            "INSERT INTO echeances (intitule, date_echeance, details) "
            "VALUES (?, ?, ?)", (intitule, date_echeance, details or None))
        db.commit()
        flash("L'échéance a été ajoutée avec succès.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('gestion_echeances'))


@app.route("/admin/echeances/modifier/<int:id>", methods=['POST'])
@admin_required
def modifier_echeance(id):
    db = get_db()
    echeance = db.execute("SELECT id FROM echeances WHERE id = ?",
                          (id, )).fetchone()
    if not echeance:
        flash("Échéance non trouvée.", "error")
        return redirect(url_for('gestion_echeances'))

    intitule = request.form.get('intitule', '').strip()
    date_echeance = request.form.get('date_echeance')
    details = request.form.get('details', '').strip()
    traite = 1 if request.form.get('traite') == 'on' else 0

    if not all([intitule, date_echeance]):
        flash("L'intitulé et la date d'échéance sont obligatoires.", "error")
        return redirect(url_for('gestion_echeances'))

    try:
        db.execute(
            "UPDATE echeances SET intitule = ?, date_echeance = ?, "
            "details = ?, traite = ? WHERE id = ?",
            (intitule, date_echeance, details or None, traite, id))
        db.commit()
        flash("L'échéance a été modifiée avec succès.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('gestion_echeances'))


@app.route("/admin/echeances/supprimer/<int:id>", methods=['POST'])
@admin_required
def supprimer_echeance(id):
    db = get_db()
    echeance = db.execute("SELECT id FROM echeances WHERE id = ?",
                          (id, )).fetchone()
    if echeance:
        try:
            db.execute("DELETE FROM echeances WHERE id = ?", (id, ))
            db.commit()
            flash("L'échéance a été supprimée avec succès.", "success")
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données : {e}", "error")
    else:
        flash("Échéance non trouvée.", "error")

    return redirect(url_for('gestion_echeances'))


# --- ROUTES POUR LES RAPPORTS ---
@app.route("/admin/rapports", methods=['GET'])
@admin_required
def rapports():
    db = get_db()
    dernieres_actions = db.execute("""
        SELECT h.timestamp, h.action, o.nom as objet_nom, u.nom_utilisateur
        FROM historique h
        JOIN objets o ON h.objet_id = o.id
        JOIN utilisateurs u ON h.utilisateur_id = u.id
        ORDER BY h.timestamp DESC
        LIMIT 5
        """).fetchall()

    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()

    return render_template("rapports.html",
                           dernieres_actions=dernieres_actions,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)

@app.route("/admin/rapports/exporter", methods=['GET'])
@admin_required
def exporter_rapports():
    db = get_db()
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    group_by = request.args.get('group_by')
    format_type = request.args.get('format')

    if not all([date_debut, date_fin, group_by, format_type]):
        flash("Tous les champs sont requis pour générer un rapport.", "error")
        return redirect(url_for('rapports'))

    try:
        date_fin_dt = datetime.strptime(date_fin, '%Y-%m-%d') + timedelta(days=1)
        date_fin_str = date_fin_dt.strftime('%Y-%m-%d')
    except ValueError:
        flash("Format de date invalide.", "error")
        return redirect(url_for('rapports'))

    query = """
        SELECT h.timestamp, h.action, h.details, o.nom as objet_nom,
               u.nom_utilisateur
        FROM historique h
        JOIN objets o ON h.objet_id = o.id
        JOIN utilisateurs u ON h.utilisateur_id = u.id
        WHERE h.timestamp >= ? AND h.timestamp < ?
    """
    order_clause = "ORDER BY h.timestamp ASC"
    if group_by == 'action':
        order_clause = "ORDER BY h.action ASC, h.timestamp ASC"
    query += order_clause

    historique_data = db.execute(query, (date_debut, date_fin_str)).fetchall()

    if not historique_data:
        flash("Aucune donnée d'historique trouvée pour la période sélectionnée.", "warning")
        return redirect(url_for('rapports'))

    if format_type == 'pdf':
        buffer = generer_rapport_pdf(historique_data, date_debut, date_fin, group_by)
        return send_file(buffer,
                         as_attachment=True,
                         download_name='rapport_activite.pdf',
                         mimetype='application/pdf')
    elif format_type == 'excel':
        buffer = generer_rapport_excel(historique_data, date_debut, date_fin, group_by)
        return send_file(
            buffer,
            as_attachment=True,
            download_name='rapport_activite.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    flash("Format d'exportation non valide.", "error")
    return redirect(url_for('rapports'))


@app.route("/objet/<int:objet_id>/telecharger_fds")
@login_required
def telecharger_fds_objet(objet_id):
    db = get_db()
    objet = db.execute(
        "SELECT fds_nom_original, fds_nom_securise FROM objets WHERE id = ?",
        (objet_id, )).fetchone()
    if objet and objet['fds_nom_securise']:
        try:
            return send_file(os.path.join(app.config['FDS_UPLOAD_FOLDER'],
                                          objet['fds_nom_securise']),
                             as_attachment=True,
                             download_name=objet['fds_nom_original'])
        except FileNotFoundError:
            flash("Le fichier FDS n'a pas été trouvé sur le serveur.", "error")
            return redirect(url_for('voir_objet', objet_id=objet_id))
    else:
        flash("Cet objet n'a pas de FDS associée.", "error")
        return redirect(url_for('voir_objet', objet_id=objet_id))


# --- ROUTES D'AUTHENTIFICATION ---
@app.route("/login", methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        db = get_db()
        user = db.execute(
            'SELECT * FROM utilisateurs WHERE nom_utilisateur = ?',
            (username, )).fetchone()
        if user and check_password_hash(user['mot_de_passe'], password):
            session.permanent = (user['role'] != 'admin')
            session['user_id'] = user['id']
            session['username'] = user['nom_utilisateur']
            session['user_role'] = user['role']
            flash(f"Bienvenue, {user['nom_utilisateur']} !", "success")
            return redirect(url_for('index'))
        else:
            flash("Nom d'utilisateur ou mot de passe invalide.", "error")
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if is_setup_needed():
        return redirect(url_for('setup'))
    if request.method == 'POST':
        username = request.form.get('username').strip()
        password = request.form.get('password')
        email = request.form.get('email').strip()
        if not username or not password or not email:
            flash("Tous les champs sont requis.", "error")
            return redirect(url_for('register'))
        db = get_db()
        try:
            db.execute(
                "INSERT INTO utilisateurs (nom_utilisateur, mot_de_passe, "
                "email) VALUES (?, ?, ?)",
                (username, generate_password_hash(password,
                                                  method='scrypt'), email))
            db.commit()
            flash(
                f"Le compte '{username}' a été créé. "
                "Vous pouvez maintenant vous connecter.", "success")
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash(f"Le nom d'utilisateur '{username}' existe déjà.", "error")
            return redirect(url_for('register'))
    return render_template('register.html')


@app.route("/logout")
def logout():
    session.clear()
    flash("Vous avez été déconnecté.", "success")
    return redirect(url_for('login'))


# --- ROUTES D'ACTION ---
@app.route("/ajouter", methods=["POST"])
@login_required
def ajouter():
    type_objet = request.form.get("type")
    nom = request.form.get("nom", "").strip()
    redirect_to = ("gestion_armoires"
                   if type_objet == "armoire" else "gestion_categories")
    if not nom:
        flash("Le nom ne peut pas être vide.", "error")
        return redirect(url_for(redirect_to))
    table_name = "armoires" if type_objet == "armoire" else "categories"
    db = get_db()
    try:
        db.execute(f"INSERT INTO {table_name} (nom) VALUES (?)", (nom, ))
        db.commit()
        flash(f"L'élément '{nom}' a été créé.", "success")
    except sqlite3.IntegrityError:
        flash(f"L'élément '{nom}' existe déjà.", "error")
    return redirect(url_for(redirect_to))


@app.route("/ajouter_objet", methods=["POST"])
@login_required
@limit_objets_required
def ajouter_objet():
    nom = request.form.get("nom", "").strip()
    quantite = request.form.get("quantite")
    seuil = request.form.get("seuil")
    armoire_id = request.form.get("armoire_id")
    categorie_id = request.form.get("categorie_id")
    date_peremption = request.form.get("date_peremption")
    date_peremption_db = date_peremption if date_peremption else None

    image_name = ""
    if 'image' in request.files:
        image = request.files['image']
        if image and image.filename != '':
            filename = secure_filename(image.filename)
            image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            image_name = filename

    fds_nom_original = None
    fds_nom_securise = None
    if 'fds_file' in request.files:
        fds_file = request.files['fds_file']
        if fds_file and fds_file.filename != '':
            fds_nom_original = fds_file.filename
            fds_nom_securise = str(uuid.uuid4()) + '_' + \
                secure_filename(fds_nom_original)
            if not os.path.exists(app.config['FDS_UPLOAD_FOLDER']):
                os.makedirs(app.config['FDS_UPLOAD_FOLDER'])
            fds_file.save(
                os.path.join(app.config['FDS_UPLOAD_FOLDER'],
                             fds_nom_securise))

    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        """INSERT INTO objets (nom, quantite, seuil, armoire_id, categorie_id,
                               image, en_commande, date_peremption, traite,
                               fds_nom_original, fds_nom_securise)
           VALUES (?, ?, ?, ?, ?, ?, 0, ?, 0, ?, ?)""",
        (nom, quantite, seuil, armoire_id, categorie_id, image_name,
         date_peremption_db, fds_nom_original, fds_nom_securise))
    new_objet_id = cursor.lastrowid
    db.commit()
    details_str = f"Créé avec quantité {quantite} et seuil {seuil}."
    enregistrer_action(new_objet_id, "Création", details_str)
    flash(f"L'objet '{nom}' a été ajouté avec succès !", "success")
    return redirect(request.referrer or url_for('index'))


@app.route("/supprimer/<type_objet>/<int:id>", methods=["POST"])
@admin_required
def supprimer(type_objet, id):
    db = get_db()
    redirect_to = ("gestion_armoires"
                   if type_objet == "armoire" else "gestion_categories")
    if type_objet == "armoire":
        if db.execute("SELECT COUNT(id) FROM objets WHERE armoire_id = ?",
                      (id, )).fetchone()[0] > 0:
            flash(
                "Impossible de supprimer. Cette armoire contient encore "
                "des objets.", "error")
            return redirect(url_for(redirect_to))
    table_map = {"armoire": "armoires", "categorie": "categories"}
    if type_objet in table_map:
        table_name = table_map[type_objet]
        nom_element = db.execute(f"SELECT nom FROM {table_name} WHERE id = ?",
                                 (id, )).fetchone()
        db.execute(f"DELETE FROM {table_name} WHERE id = ?", (id, ))
        db.commit()
        if nom_element:
            flash(
                f"L'élément '{nom_element['nom']}' a été supprimé avec succès.",
                "success")
    else:
        flash("Type d'élément à supprimer non valide.", "error")
    return redirect(url_for(redirect_to))


@app.route("/modifier_objet/<int:id_objet>", methods=["POST"])
@login_required
def modifier_objet(id_objet):
    db = get_db()
    objet_avant = db.execute("SELECT * FROM objets WHERE id = ?",
                             (id_objet, )).fetchone()
    if not objet_avant:
        flash("Objet non trouvé.", "error")
        return redirect(request.referrer or url_for('index'))

    nom = request.form.get("nom", "").strip()
    quantite = int(request.form.get("quantite"))
    seuil = int(request.form.get("seuil"))
    armoire_id = int(request.form.get("armoire_id"))
    categorie_id = int(request.form.get("categorie_id"))
    date_peremption = request.form.get("date_peremption")
    date_peremption_db = date_peremption if date_peremption else None

    image_name = objet_avant['image']

    # --- NOUVELLE LOGIQUE POUR LA SUPPRESSION D'IMAGE ---
    if request.form.get('supprimer_image'):
        if image_name:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], image_name))
            except OSError:
                # Le fichier n'existe peut-être plus, on ignore l'erreur
                pass
        image_name = "" # On vide le nom de l'image pour la BDD
    
    if 'image' in request.files:
        nouvelle_image = request.files['image']
        if nouvelle_image and nouvelle_image.filename != '':
            # On supprime l'ancienne image s'il y en avait une
            if image_name:
                try:
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], image_name))
                except OSError:
                    pass
            # On sauvegarde la nouvelle
            filename = secure_filename(nouvelle_image.filename)
            nouvelle_image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            image_name = filename

    fds_nom_original = objet_avant['fds_nom_original']
    fds_nom_securise = objet_avant['fds_nom_securise']
    if 'fds_file' in request.files:
        nouvelle_fds = request.files['fds_file']
        if nouvelle_fds and nouvelle_fds.filename != '':
            if fds_nom_securise:
                try:
                    os.remove(
                        os.path.join(app.config['FDS_UPLOAD_FOLDER'],
                                     fds_nom_securise))
                except OSError:
                    pass

            fds_nom_original = nouvelle_fds.filename
            fds_nom_securise = str(uuid.uuid4()) + '_' + \
                secure_filename(fds_nom_original)
            nouvelle_fds.save(
                os.path.join(app.config['FDS_UPLOAD_FOLDER'],
                             fds_nom_securise))

    details = []
    if objet_avant['fds_nom_original'] != fds_nom_original:
        details.append("FDS modifiée")

    details_str = ", ".join(details)
    if not details_str:
        details_str = "Aucun changement détecté."

    db.execute(
        """
        UPDATE objets SET nom = ?, quantite = ?, seuil = ?, armoire_id = ?,
                         categorie_id = ?, image = ?, date_peremption = ?,
                         fds_nom_original = ?, fds_nom_securise = ?
        WHERE id = ?
        """,
        (nom, quantite, seuil, armoire_id, categorie_id, image_name,
         date_peremption_db, fds_nom_original, fds_nom_securise, id_objet))
    db.commit()

    if details_str != "Aucun changement détecté.":
        enregistrer_action(id_objet, "Modification", details_str)

    flash(f"L'objet '{nom}' a été mis à jour avec succès !", "success")
    return redirect(request.referrer or url_for('index'))

@app.route("/objet/supprimer/<int:id_objet>", methods=["POST"])
@admin_required
def supprimer_objet(id_objet):
    """Supprime un objet, son historique, et ses fichiers associés."""
    db = get_db()
    objet = db.execute(
        "SELECT nom, image, fds_nom_securise FROM objets WHERE id = ?",
        (id_objet,)
    ).fetchone()

    if not objet:
        flash("Objet non trouvé.", "error")
        return redirect(request.referrer or url_for('inventaire'))

    try:
        # Étape 1 : Supprimer les fichiers physiques pour ne pas laisser d'orphelins
        if objet['image']:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], objet['image']))
        if objet['fds_nom_securise']:
            os.remove(os.path.join(app.config['FDS_UPLOAD_FOLDER'], objet['fds_nom_securise']))

        # Étape 2 : Supprimer les enregistrements dépendants
        db.execute("DELETE FROM historique WHERE objet_id = ?", (id_objet,))
        db.execute("DELETE FROM reservations WHERE objet_id = ?", (id_objet,))
        db.execute("DELETE FROM kit_objets WHERE objet_id = ?", (id_objet,))

        # Étape 3 : Supprimer l'objet lui-même
        db.execute("DELETE FROM objets WHERE id = ?", (id_objet,))

        db.commit()
        flash(f"L'objet '{objet['nom']}' et toutes ses données associées ont été supprimés.", "success")

    except OSError:
        # Gère le cas où un fichier n'existe pas, mais on continue la suppression
        flash(f"Un fichier associé à '{objet['nom']}' n'a pas pu être trouvé, mais l'objet a été supprimé de la base de données.", "warning")
        db.commit() # On s'assure que la suppression en BDD a bien lieu
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Une erreur de base de données est survenue : {e}", "error")

    return redirect(request.referrer or url_for('inventaire'))

@app.route("/modifier_armoire", methods=["POST"])
@admin_required
def modifier_armoire():
    data = request.get_json()
    armoire_id, nouveau_nom = data.get("id"), data.get("nom")
    if not all([armoire_id, nouveau_nom, nouveau_nom.strip()]):
        return jsonify(success=False, error="Données invalides"), 400
    db = get_db()
    try:
        db.execute("UPDATE armoires SET nom = ? WHERE id = ?",
                   (nouveau_nom.strip(), armoire_id))
        db.commit()
        return jsonify(success=True, nouveau_nom=nouveau_nom.strip())
    except sqlite3.IntegrityError:
        return jsonify(success=False,
                       error="Ce nom d'armoire existe déjà."), 500


@app.route("/modifier_categorie", methods=["POST"])
@admin_required
def modifier_categorie():
    data = request.get_json()
    categorie_id, nouveau_nom = data.get("id"), data.get("nom")
    if not all([categorie_id, nouveau_nom, nouveau_nom.strip()]):
        return jsonify(success=False, error="Données invalides"), 400
    db = get_db()
    try:
        db.execute("UPDATE categories SET nom = ? WHERE id = ?",
                   (nouveau_nom.strip(), categorie_id))
        db.commit()
        return jsonify(success=True, nouveau_nom=nouveau_nom.strip())
    except sqlite3.IntegrityError:
        return jsonify(success=False,
                       error="Ce nom de catégorie existe déjà."), 500


# --- ROUTES API ---
@app.route("/api/rechercher")
@login_required
def api_rechercher():
    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify([])
    db = get_db()
    resultats = db.execute(
        """SELECT o.id, o.nom, o.armoire_id, a.nom as armoire_nom,
                  c.nom as categorie_nom
           FROM objets o
           JOIN armoires a ON o.armoire_id = a.id
           JOIN categories c ON o.categorie_id = c.id
           WHERE unaccent(LOWER(o.nom)) LIKE unaccent(LOWER(?))
           LIMIT 10""", (f"%{query}%", )).fetchall()
    return jsonify([dict(row) for row in resultats])


@app.route("/api/filtrer_inventaire")
@login_required
def api_filtrer_inventaire():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort_by', 'nom')
    direction = request.args.get('direction', 'asc')
    search_query = request.args.get('q', None)
    armoire_id = request.args.get('armoire', None)
    categorie_id = request.args.get('categorie', None)
    etat = request.args.get('etat', None)

    objets, total_pages = get_paginated_objets(db, page, sort_by, direction,
                                               search_query, armoire_id,
                                               categorie_id, etat)

    pagination = {
        'page': page,
        'total_pages': total_pages,
        'endpoint': 'inventaire',
        'id': None
    }

    table_html = render_template('_table_objets.html',
                                 objets=objets,
                                 date_actuelle=datetime.now(),
                                 sort_by=sort_by,
                                 direction=direction,
                                 pagination=pagination)
    pagination_html = render_template('_pagination.html',
                                      pagination=pagination,
                                      sort_by=sort_by,
                                      direction=direction)

    return jsonify(table_html=table_html, pagination_html=pagination_html)

@app.route("/api/inventaire/")
@login_required
def api_inventaire():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort_by', 'nom')
    direction = request.args.get('direction', 'asc')
    search_query = request.args.get('q', None)
    armoire_id = request.args.get('armoire', None)
    categorie_id = request.args.get('categorie', None)
    etat = request.args.get('etat', None)

    objets, total_pages = get_paginated_objets(db, page, sort_by, direction,
                                               search_query, armoire_id,
                                               categorie_id, etat)
    pagination = {'page': page, 'total_pages': total_pages, 'endpoint': 'inventaire', 'id': None}
    
    # On utilise un nouveau template partiel pour le contenu de la page inventaire
    html = render_template('_inventaire_content.html', objets=objets, pagination=pagination, date_actuelle=datetime.now(), sort_by=sort_by, direction=direction, session=session)
    return jsonify(html=html)

@app.route("/api/deplacer_objets", methods=['POST'])
@admin_required
def deplacer_objets():
    data = request.get_json()
    objet_ids = data.get('objet_ids')
    destination_id = data.get('destination_id')
    type_destination = data.get('type_destination')

    if not all([objet_ids, destination_id, type_destination]):
        return jsonify(success=False, error="Données manquantes."), 400

    db = get_db()
    try:
        field_to_update = ('categorie_id' if type_destination == 'categorie'
                           else 'armoire_id')

        for objet_id in objet_ids:
            db.execute(f"UPDATE objets SET {field_to_update} = ? WHERE id = ?",
                       (destination_id, objet_id))

        db.commit()
        flash(f"{len(objet_ids)} objet(s) déplacé(s) avec succès.", "success")
        return jsonify(success=True)
    except sqlite3.Error as e:
        db.rollback()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/reservations_par_mois/<int:year>/<int:month>")
@login_required
def api_reservations_par_mois(year, month):
    db = get_db()

    start_date_str = f"{year}-{str(month).zfill(2)}-01 00:00:00"

    if month == 12:
        end_date_str = f"{year + 1}-01-01 00:00:00"
    else:
        end_date_str = f"{year}-{str(month + 1).zfill(2)}-01 00:00:00"

    reservations = db.execute(
        """
        SELECT
            DATE(r.debut_reservation) as jour_reservation,
            r.groupe_id,
            r.utilisateur_id,
            u.nom_utilisateur
        FROM reservations r
        JOIN utilisateurs u ON r.utilisateur_id = u.id
        WHERE r.debut_reservation >= ? AND r.debut_reservation < ?
              AND r.groupe_id IS NOT NULL
        GROUP BY jour_reservation, r.groupe_id
        """, (start_date_str, end_date_str)).fetchall()

    results = {}
    for row in reservations:
        date = row['jour_reservation']
        if date not in results:
            results[date] = []

        results[date].append(
            {'is_mine': row['utilisateur_id'] == session['user_id']})
    return jsonify(results)


@app.route("/api/reservation_details/<groupe_id>")
@login_required
def api_reservation_details(groupe_id):
    db = get_db()
    reservations = db.execute(
        """
        SELECT r.quantite_reservee, o.id as objet_id, o.nom as objet_nom, 
               u.nom_utilisateur, r.utilisateur_id, r.kit_id, k.nom as kit_nom,
               r.debut_reservation, r.fin_reservation
        FROM reservations r
        JOIN objets o ON r.objet_id = o.id
        JOIN utilisateurs u ON r.utilisateur_id = u.id
        LEFT JOIN kits k ON r.kit_id = k.id
        WHERE r.groupe_id = ?
        """, (groupe_id, )).fetchall()

    if not reservations:
        return jsonify({'error': 'Réservation non trouvée'}), 404

    details = {
        'kits': {},
        'objets_manuels': [],
        'nom_utilisateur': reservations[0]['nom_utilisateur'],
        'utilisateur_id': reservations[0]['utilisateur_id'],
        'debut_reservation': reservations[0]['debut_reservation'],
        'fin_reservation': reservations[0]['fin_reservation']
    }

    objets_manuels_calcul = {r['objet_id']: dict(r) for r in reservations if r['kit_id'] is None}
    objets_kits_reserves = [r for r in reservations if r['kit_id'] is not None]

    kits_comptes = {}
    for r in objets_kits_reserves:
        if r['kit_id'] not in kits_comptes:
            kits_comptes[r['kit_id']] = {'nom': r['kit_nom'], 'objets_reserves': {}}
        kits_comptes[r['kit_id']]['objets_reserves'][r['objet_id']] = r['quantite_reservee']

    for kit_id, data in kits_comptes.items():
        objets_base_du_kit = db.execute("SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?", (kit_id,)).fetchall()
        if not objets_base_du_kit: continue
        
        id_objet_calcul = None
        quantite_par_kit = 0
        for obj_base in objets_base_du_kit:
            if obj_base['objet_id'] in data['objets_reserves']:
                id_objet_calcul = obj_base['objet_id']
                quantite_par_kit = obj_base['quantite']
                break

        if id_objet_calcul and quantite_par_kit > 0:
            quantite_reelle_reservee = data['objets_reserves'][id_objet_calcul]
            nombre_de_kits = quantite_reelle_reservee // quantite_par_kit
            details['kits'][str(kit_id)] = {'quantite': nombre_de_kits, 'nom': data['nom']}

            for obj_base in objets_base_du_kit:
                if obj_base['objet_id'] in objets_manuels_calcul:
                    objets_manuels_calcul[obj_base['objet_id']]['quantite_reservee'] -= nombre_de_kits * obj_base['quantite']

    for obj_id, data in objets_manuels_calcul.items():
        if data['quantite_reservee'] > 0:
            # --- CORRECTION CI-DESSOUS ---
            # On ajoute l'objet_id et on utilise 'quantite' pour la cohérence
            details['objets_manuels'].append({
                'objet_id': obj_id,
                'quantite_reservee': data['quantite_reservee'],
                'nom': data['objet_nom']
            })

    return jsonify(details)


@app.route("/api/reservation_data/<date>")
@login_required
def api_reservation_data(date):
    db = get_db()

    start_of_day = f"{date} 00:00:00"
    end_of_day = f"{date} 23:59:59"

    objets = db.execute(
        """
        SELECT o.id, o.nom, c.nom as categorie, o.quantite,
               (SELECT COALESCE(SUM(r.quantite_reservee), 0)
                FROM reservations r
                WHERE r.objet_id = o.id AND r.debut_reservation <= ?
                      AND r.fin_reservation >= ?) as deja_reserve_ce_jour
        FROM objets o
        JOIN categories c ON o.categorie_id = c.id
        ORDER BY c.nom, o.nom
        """, (end_of_day, start_of_day)).fetchall()

    grouped_objets = {}
    for row in objets:
        categorie_nom = row['categorie']
        if categorie_nom not in grouped_objets:
            grouped_objets[categorie_nom] = []

        total_reserve_all_time = db.execute(
            "SELECT COALESCE(SUM(quantite_reservee), 0) FROM reservations "
            "WHERE objet_id = ?", (row['id'], )).fetchone()[0]
        quantite_totale = row['quantite'] + total_reserve_all_time

        grouped_objets[categorie_nom].append({
            "id":
            row['id'],
            "nom":
            row['nom'],
            "quantite_totale":
            quantite_totale,
            "quantite_disponible":
            row['quantite']
        })

    kits = db.execute(
        "SELECT id, nom, description FROM kits ORDER BY nom").fetchall()
    kits_details = []
    for kit in kits:
        objets_du_kit = db.execute(
            "SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?",
            (kit['id'], )).fetchall()

        kits_details.append({
            'id': kit['id'],
            'nom': kit['nom'],
            'description': kit['description'],
            'objets': [dict(o) for o in objets_du_kit]
        })

    return jsonify({'objets': grouped_objets, 'kits': kits_details})


def process_and_insert_reservations(db, date, heure_debut, heure_fin, user_id,
                                    groupe_id, reservations_a_faire,
                                    kits_reserves):
    debut_dt = datetime.strptime(f"{date} {heure_debut}", '%Y-%m-%d %H:%M')
    fin_dt = datetime.strptime(f"{date} {heure_fin}", '%Y-%m-%d %H:%M')
    date_formatee = debut_dt.strftime('%d/%m/%Y')

    for res in reservations_a_faire:
        obj_id = int(res['objet_id'])
        quantite_totale = int(res['quantite'])
        
        kit_id_associe = None
        # La clé dans kits_reserves peut être une chaîne, on la compare en int
        for kit_id_str in kits_reserves.keys():
            kit_id = int(kit_id_str)
            is_in_kit = db.execute("SELECT 1 FROM kit_objets WHERE kit_id = ? AND objet_id = ?", (kit_id, obj_id)).fetchone()
            if is_in_kit:
                kit_id_associe = kit_id
                break

        db.execute("UPDATE objets SET quantite = quantite - ? WHERE id = ?",
                   (quantite_totale, obj_id))
        
        db.execute(
            """INSERT INTO reservations (objet_id, quantite_reservee,
               debut_reservation, fin_reservation, utilisateur_id,
               groupe_id, kit_id)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (obj_id, quantite_totale, debut_dt, fin_dt, user_id, groupe_id, kit_id_associe))
        
        action_detail = f"Quantité: {quantite_totale} pour le {date_formatee}"
        enregistrer_action(obj_id, "Réservation", action_detail)


@app.route("/api/reserver", methods=["POST"])
@login_required
def api_reserver():
    data = request.get_json()
    date_str = data.get("date")
    
    # --- SÉCURITÉ : VÉRIFICATION DE LA DATE ---
    try:
        reservation_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        if reservation_date < date.today():
            return jsonify(success=False, error="Il est impossible de faire une réservation pour une date passée."), 400
    except (ValueError, TypeError):
        return jsonify(success=False, error="Format de date invalide."), 400

    heure_debut = data.get("heure_debut")
    heure_fin = data.get("heure_fin")
    reservations_a_faire = data.get("reservations")
    kits_reserves = data.get("kits", {})
    user_id = session['user_id']
    db = get_db()

    try:
        for res in reservations_a_faire:
            stock_actuel = db.execute(
                "SELECT nom, quantite FROM objets WHERE id = ?",
                (res['objet_id'], )).fetchone()
            if not stock_actuel or stock_actuel['quantite'] < res['quantite']:
                nom_objet = stock_actuel['nom'] if stock_actuel else 'Inconnu'
                stock_dispo = stock_actuel['quantite'] if stock_actuel else 0
                return jsonify(success=False,
                               error=f"Stock insuffisant pour '{nom_objet}'. "
                               f"Demandé : {res['quantite']}, "
                               f"Disponible : {stock_dispo}."), 400

        groupe_id = str(uuid.uuid4())

        process_and_insert_reservations(db, date_str, heure_debut, heure_fin,
                                        user_id, groupe_id,
                                        reservations_a_faire, kits_reserves)

        db.commit()
        flash("Matériel réservé avec succès !", "success")
        return jsonify(success=True)
    except sqlite3.Error as e:
        db.rollback()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/modifier_reservation", methods=["POST"])
@login_required
def api_modifier_reservation():
    data = request.get_json()
    date_str = data.get("date")

    # --- SÉCURITÉ : VÉRIFICATION DE LA DATE ---
    try:
        reservation_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        if reservation_date < date.today():
            return jsonify(success=False, error="Il est impossible de déplacer une réservation vers une date passée."), 400
    except (ValueError, TypeError):
        return jsonify(success=False, error="Format de date invalide."), 400

    groupe_id = data.get("groupe_id")
    heure_debut = data.get("heure_debut")
    heure_fin = data.get("heure_fin")
    nouvelles_reservations = data.get("reservations", [])
    kits_reserves = data.get("kits", {})
    user_id = session['user_id']
    db = get_db()

    try:
        ancienne_reservation = db.execute(
            "SELECT objet_id, quantite_reservee FROM reservations "
            "WHERE groupe_id = ?", (groupe_id, )).fetchall()
        if not ancienne_reservation:
            return jsonify(success=False,
                           error="Réservation originale non trouvée."), 404

        for item in ancienne_reservation:
            db.execute(
                "UPDATE objets SET quantite = quantite + ? WHERE id = ?",
                (item['quantite_reservee'], item['objet_id']))
            enregistrer_action(item['objet_id'],
                               "Modification Résa (Restitution)",
                               f"Quantité: {item['quantite_reservee']}")

        db.execute("DELETE FROM reservations WHERE groupe_id = ?",
                   (groupe_id, ))

        for res in nouvelles_reservations:
            stock_actuel = db.execute(
                "SELECT nom, quantite FROM objets WHERE id = ?",
                (res['objet_id'], )).fetchone()
            if not stock_actuel or stock_actuel['quantite'] < res['quantite']:
                db.rollback()
                nom_objet = stock_actuel['nom'] if stock_actuel else 'Inconnu'
                stock_dispo = stock_actuel['quantite'] if stock_actuel else 0
                return jsonify(success=False,
                               error=f"Stock insuffisant pour '{nom_objet}'. "
                               f"Demandé : {res['quantite']}, "
                               f"Disponible : {stock_dispo}."), 400

        process_and_insert_reservations(db, date_str, heure_debut, heure_fin,
                                        user_id, groupe_id,
                                        nouvelles_reservations, kits_reserves)

        db.commit()
        flash("Réservation modifiée avec succès !", "success")
        return jsonify(success=True)
    except sqlite3.Error as e:
        db.rollback()
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/supprimer_reservation", methods=["POST"])
@login_required
def api_supprimer_reservation():
    data = request.get_json()
    groupe_id = data.get("groupe_id")
    db = get_db()
    try:
        reservation_info = db.execute(
            "SELECT utilisateur_id, debut_reservation FROM reservations "
            "WHERE groupe_id = ? LIMIT 1", (groupe_id, )).fetchone()
        if not reservation_info:
            return jsonify(success=False,
                           error="Réservation non trouvée."), 404
        if (session.get('user_role') != 'admin'
                and reservation_info['utilisateur_id'] != session['user_id']):
            return jsonify(
                success=False,
                error="Vous n'avez pas la permission de supprimer cette "
                "réservation."), 403
        items_to_restock = db.execute(
            "SELECT objet_id, quantite_reservee FROM reservations "
            "WHERE groupe_id = ?", (groupe_id, )).fetchall()
        date_formatee = datetime.strptime(
            reservation_info['debut_reservation'],
            '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
        for item in items_to_restock:
            db.execute(
                "UPDATE objets SET quantite = quantite + ? WHERE id = ?",
                (item['quantite_reservee'], item['objet_id']))
            enregistrer_action(
                item['objet_id'], "Annulation Réservation",
                f"Quantité: {item['quantite_reservee']} "
                f"(résa du {date_formatee})")
        db.execute("DELETE FROM reservations WHERE groupe_id = ?",
                   (groupe_id, ))
        db.commit()
        flash("La réservation a été annulée et le stock mis à jour.",
              "success")
        return jsonify(success=True)
    except sqlite3.Error as e:
        db.rollback()
        return jsonify(success=False, error=str(e)), 500

@app.route("/api/valider_panier", methods=["POST"])
@login_required
def api_valider_panier():
    cart_data = request.get_json()
    if not cart_data:
        return jsonify(success=False, error="Le panier est vide."), 400

    db = get_db()
    user_id = session['user_id']

    try:
        today = date.today()
        
        reservations_par_creneau = {}
        for creneau_key, resa_details in cart_data.items():
            try:
                reservation_date = datetime.strptime(resa_details['date'], '%Y-%m-%d').date()
                if reservation_date < today:
                    msg = f"Le panier contient une réservation pour une date passée ({reservation_date.strftime('%d/%m/%Y')}). Veuillez la retirer."
                    return jsonify(success=False, error=msg), 400
            except (ValueError, TypeError):
                return jsonify(success=False, error="Le panier contient une réservation avec une date invalide."), 400

            objets_agreges = {}
            for kit_id, kit_data in resa_details.get('kits', {}).items():
                quantite_kit = kit_data.get('quantite', 0)
                objets_du_kit = db.execute("SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?", (kit_id,)).fetchall()
                for obj_in_kit in objets_du_kit:
                    obj_id_str = str(obj_in_kit['objet_id'])
                    quantite_a_ajouter = obj_in_kit['quantite'] * quantite_kit
                    objets_agreges[obj_id_str] = objets_agreges.get(obj_id_str, 0) + quantite_a_ajouter
            
            for obj_id, obj_data in resa_details.get('objets', {}).items():
                objets_agreges[obj_id] = objets_agreges.get(obj_id, 0) + obj_data.get('quantite', 0)

            reservations_par_creneau[creneau_key] = {
                'details': resa_details,
                'objets_agreges': objets_agreges
            }

        for creneau_key, data in reservations_par_creneau.items():
            for obj_id, quantite_demandee in data['objets_agreges'].items():
                stock_actuel = db.execute("SELECT nom, quantite FROM objets WHERE id = ?", (obj_id,)).fetchone()
                if not stock_actuel or stock_actuel['quantite'] < quantite_demandee:
                    nom_objet = stock_actuel['nom'] if stock_actuel else f"ID {obj_id}"
                    stock_dispo = stock_actuel['quantite'] if stock_actuel else 0
                    error_msg = (f"Stock insuffisant pour '{nom_objet}' le {data['details']['date']} "
                                 f"(demandé: {quantite_demandee}, disponible: {stock_dispo}). La transaction a été annulée.")
                    return jsonify(success=False, error=error_msg), 400

        for creneau_key, data in reservations_par_creneau.items():
            groupe_id = str(uuid.uuid4())
            resa_details = data['details']
            
            # --- BLOC DE LOGIQUE CORRIGÉ ---
            # On utilise un dictionnaire avec une clé composite (objet_id, kit_id) pour agréger correctement
            final_reservations = {}

            # Traiter les kits
            for kit_id_str, kit_data in resa_details.get('kits', {}).items():
                kit_id = int(kit_id_str)
                quantite_kit = kit_data.get('quantite', 0)
                objets_du_kit = db.execute("SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?", (kit_id,)).fetchall()
                for obj_in_kit in objets_du_kit:
                    key = (obj_in_kit['objet_id'], kit_id)
                    final_reservations[key] = final_reservations.get(key, 0) + (obj_in_kit['quantite'] * quantite_kit)

            # Traiter les objets manuels
            for obj_id_str, obj_data in resa_details.get('objets', {}).items():
                obj_id = int(obj_id_str)
                key = (obj_id, None) # Le kit_id est None pour les objets manuels
                final_reservations[key] = final_reservations.get(key, 0) + obj_data.get('quantite', 0)

            # Insertion finale dans la base de données
            debut_dt = datetime.strptime(f"{resa_details['date']} {resa_details['heure_debut']}", '%Y-%m-%d %H:%M')
            fin_dt = datetime.strptime(f"{resa_details['date']} {resa_details['heure_fin']}", '%Y-%m-%d %H:%M')
            date_formatee = debut_dt.strftime('%d/%m/%Y')

            for (obj_id, kit_id), quantite_totale in final_reservations.items():
                if quantite_totale > 0:
                    db.execute("UPDATE objets SET quantite = quantite - ? WHERE id = ?", (quantite_totale, obj_id))
                    db.execute(
                        """INSERT INTO reservations (objet_id, quantite_reservee,
                           debut_reservation, fin_reservation, utilisateur_id,
                           groupe_id, kit_id)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (obj_id, quantite_totale, debut_dt, fin_dt, user_id, groupe_id, kit_id))
                    
                    action_detail = f"Quantité: {quantite_totale} pour le {date_formatee}"
                    enregistrer_action(obj_id, "Réservation", action_detail)
            # --- FIN DU BLOC CORRIGÉ ---

        db.commit()
        flash("Toutes vos réservations ont été confirmées avec succès !", "success")
        return jsonify(success=True)

    except Exception as e:
        db.rollback()
        traceback.print_exc()
        return jsonify(success=False, error=f"Une erreur interne est survenue : {e}"), 500


@app.route("/maj_commande/<int:objet_id>", methods=["POST"])
@login_required
def maj_commande(objet_id):
    data = request.get_json()
    en_commande = 1 if data.get("en_commande") else 0
    db = get_db()
    db.execute("UPDATE objets SET en_commande = ? WHERE id = ?",
               (en_commande, objet_id))
    db.commit()
    return jsonify(success=True)


@app.route("/api/maj_traite/<int:objet_id>", methods=["POST"])
@login_required
def maj_traite(objet_id):
    data = request.get_json()
    traite = 1 if data.get("traite") else 0
    db = get_db()
    db.execute("UPDATE objets SET traite = ? WHERE id = ?", (traite, objet_id))
    db.commit()
    return jsonify(success=True)


@app.route("/api/suggestion_commande/<int:objet_id>")
@admin_required
def api_suggestion_commande(objet_id):
    db = get_db()

    date_limite = datetime.now() - timedelta(days=90)

    result = db.execute(
        """
        SELECT SUM(quantite_reservee)
        FROM reservations
        WHERE objet_id = ? AND debut_reservation >= ?
        """, (objet_id, date_limite)).fetchone()

    consommation = result[0] if result and result[0] is not None else 0

    suggestion = 0
    if consommation > 0:
        suggestion = math.ceil(consommation * 1.5)
    else:
        objet = db.execute("SELECT seuil FROM objets WHERE id = ?",
                           (objet_id, )).fetchone()
        if objet:
            suggestion = objet['seuil'] * 2
        else:
            suggestion = 5

    return jsonify(suggestion=suggestion, consommation=consommation)

# --- ROUTE POUR SERVIR LES IMAGES UPLOADÉES ---
from flask import send_from_directory

@app.route('/uploads/images/<path:filename>')
@login_required
def serve_image(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# --- ROUTES D'IMPORT/EXPORT ET EXÉCUTION ---
@app.route("/telecharger_db")
@admin_required
def telecharger_db():
    db = get_db()
    licence_row = db.execute("SELECT valeur FROM parametres WHERE cle = ?",
                             ('licence_statut', )).fetchone()
    is_pro = licence_row and licence_row['valeur'] == 'PRO'

    if not is_pro:
        flash(
            "Le téléchargement de la base de données est une fonctionnalité "
            "de la version Pro.", "warning")
        return redirect(url_for('admin'))

    return send_file(DATABASE, as_attachment=True)


@app.route("/importer_db", methods=["POST"])
@admin_required
def importer_db():
    if 'fichier' not in request.files:
        flash("Aucun fichier sélectionné.", "error")
        return redirect(url_for('admin'))
    fichier = request.files.get("fichier")
    if not fichier or fichier.filename == '':
        flash("Aucun fichier selecté.", "error")
        return redirect(url_for('admin'))
    if fichier and fichier.filename.endswith(".db"):
        temp_db_path = DATABASE + ".tmp"
        fichier.save(temp_db_path)
        shutil.move(temp_db_path, DATABASE)
        flash("Base de données importée avec succès !", "success")
    else:
        flash("Le fichier fourni n'est pas une base de données valide (.db).",
              "error")
    return redirect(url_for('admin'))


@app.route("/admin/exporter")
@admin_required
def exporter_inventaire():
    db = get_db()
    inventaire_data = db.execute("""
        SELECT o.nom, o.quantite, a.nom AS armoire, c.nom AS categorie,
               o.date_peremption
        FROM objets o JOIN armoires a ON o.armoire_id = a.id
        JOIN categories c ON o.categorie_id = c.id
        ORDER BY c.nom, o.nom
        """).fetchall()
    format_type = request.args.get('format')
    if format_type == 'pdf':
        buffer = generer_pdf(inventaire_data)
        return send_file(buffer,
                         as_attachment=True,
                         download_name='inventaire_labo.pdf',
                         mimetype='application/pdf')
    elif format_type == 'excel':
        buffer = generer_excel(inventaire_data)
        return send_file(
            buffer,
            as_attachment=True,
            download_name='inventaire_labo.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    flash("Format d'exportation non valide.","error")
    return redirect(url_for('admin'))


class PDFWithFooter(FPDF):

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'R')


def generer_pdf(data):
    pdf = PDFWithFooter()
    pdf.alias_nb_pages()
    pdf.add_page(orientation='L')
    pdf.set_font('Arial', 'B', 10)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(50, 10, 'Catégorie', 1, 0, 'C', 1)
    pdf.cell(85, 10, 'Nom de l\'objet', 1, 0, 'C', 1)
    pdf.cell(20, 10, 'Qté', 1, 0, 'C', 1)
    pdf.cell(40, 10, 'Armoire', 1, 0, 'C', 1)
    pdf.cell(30, 10, 'Péremption', 1, 1, 'C', 1)
    pdf.set_font('Arial', '', 9)
    grouped_data = {}
    for item in data:
        cat = item['categorie']
        if cat not in grouped_data:
            grouped_data[cat] = []
        grouped_data[cat].append(item)
    for categorie, items in sorted(grouped_data.items()):
        row_count = len(items)
        start_y = pdf.get_y()
        pdf.multi_cell(
            50, 7 * row_count,
            categorie.encode('latin-1', 'replace').decode('latin-1'), 1, 'C')
        pdf.set_y(start_y)
        pdf.set_x(pdf.get_x() + 50)
        for i, item in enumerate(items):
            date_peremption_str = ""
            if item['date_peremption']:
                try:
                    date_obj = datetime.strptime(item['date_peremption'],
                                                 '%Y-%m-%d')
                    date_peremption_str = date_obj.strftime('%d/%m/%Y')
                except (ValueError, TypeError):
                    date_peremption_str = item['date_peremption']
            pdf.cell(
                85, 7, item['nom'].encode('latin-1',
                                          'replace').decode('latin-1'), 1, 0)
            pdf.cell(20, 7, str(item['quantite']), 1, 0, 'C')
            pdf.cell(
                40, 7, item['armoire'].encode('latin-1',
                                              'replace').decode('latin-1'), 1,
                0)
            pdf.cell(30, 7, date_peremption_str, 1, 1, 'C')
            if i < row_count - 1:
                pdf.set_x(pdf.get_x() + 50)
 
    return BytesIO(pdf.output())


def generer_excel(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventaire"
    sheet.sheet_view.showGridLines = False
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F3B73",
                              end_color="1F3B73",
                              fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    category_font = Font(bold=True)
    category_align = Alignment(vertical='center', horizontal='left', indent=1)
    even_row_fill = PatternFill(start_color="F0F4F8",
                                end_color="F0F4F8",
                                fill_type="solid")
    thin_border_side = Side(style='thin', color="BFBFBF")
    thin_border = Border(left=thin_border_side,
                         right=thin_border_side,
                         top=thin_border_side,
                         bottom=thin_border_side)
    sheet[
        'B2'] = f"Inventaire généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sheet['B2'].font = Font(italic=True, color="6c7a89")
    headers = [
        "Catégorie", "Nom de l'objet", "Quantité", "Armoire",
        "Date de Péremption"
    ]
    start_col = 2
    for i, header in enumerate(headers):
        cell = sheet.cell(row=4, column=start_col + i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    sheet.row_dimensions[4].height = 30
    current_cat = None
    start_merge_row = 5
    row_index = 5
    is_even = False
    for item in data:
        date_peremption_val = None
        if item['date_peremption']:
            try:
                date_peremption_val = datetime.strptime(
                    item['date_peremption'], '%Y-%m-%d')
            except (ValueError, TypeError):
                date_peremption_val = item['date_peremption']
        row_data = [
            item['categorie'], item['nom'], item['quantite'], item['armoire'],
            date_peremption_val
        ]
        for i, value in enumerate(row_data):
            cell = sheet.cell(row=row_index, column=start_col + i, value=value)
            cell.border = thin_border
            cell.alignment = center_align if i > 0 else Alignment(
                vertical='center')
            if isinstance(value, datetime):
                cell.number_format = 'DD/MM/YYYY'
            if is_even:
                cell.fill = even_row_fill
        if item['categorie'] != current_cat:
            if current_cat is not None and start_merge_row < row_index:
                sheet.merge_cells(start_row=start_merge_row,
                                  start_column=start_col,
                                  end_row=row_index - 1,
                                  end_column=start_col)
                cell = sheet.cell(row=start_merge_row, column=start_col)
                cell.font = category_font
                cell.alignment = category_align
            current_cat = item['categorie']
            start_merge_row = row_index
        row_index += 1
        is_even = not is_even
    if current_cat is not None and start_merge_row < row_index - 1:
        sheet.merge_cells(start_row=start_merge_row,
                          start_column=start_col,
                          end_row=row_index - 1,
                          end_column=start_col)
        cell = sheet.cell(row=start_merge_row, column=start_col)
        cell.font = category_font
        cell.alignment = category_align
    column_widths = {'B': 30, 'C': 50, 'D': 15, 'E': 30, 'F': 20}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = 'A5'
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generer_budget_pdf(data, date_debut, date_fin):
    pdf = PDFWithFooter()
    pdf.alias_nb_pages()
    pdf.add_page(orientation='P')

    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Rapport des Depenses', 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(
        0, 10,
        f"Periode du {datetime.strptime(date_debut, '%Y-%m-%d').strftime('%d/%m/%Y')} "
        f"au {datetime.strptime(date_fin, '%Y-%m-%d').strftime('%d/%m/%Y')}",
        0, 1, 'C')
    pdf.ln(10)

    pdf.set_font('Arial', 'B', 10)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(25, 8, 'Date', 1, 0, 'C', 1)
    pdf.cell(50, 8, 'Fournisseur', 1, 0, 'C', 1)
    pdf.cell(
        85, 8, 'Contenu de la commande'.encode('latin-1',
                                               'replace').decode('latin-1'), 1,
        0, 'C', 1)
    pdf.cell(30, 8, 'Montant (EUR)', 1, 1, 'C', 1)

    pdf.set_font('Arial', '', 9)
    total_depenses = 0
    fill = False
    for item in data:
        pdf.set_fill_color(240, 240, 240)

        date_str = datetime.strptime(item['date_depense'],
                                     '%Y-%m-%d').strftime('%d/%m/%Y')
        fournisseur = (item['fournisseur_nom']
                       or 'N/A').encode('latin-1', 'replace').decode('latin-1')
        contenu = item['contenu'].encode('latin-1',
                                         'replace').decode('latin-1')
        montant = item['montant']
        total_depenses += montant

        pdf.cell(25, 7, date_str, 1, 0, 'C', fill)
        pdf.cell(50, 7, fournisseur, 1, 0, 'C', fill)
        pdf.cell(85, 7, contenu, 1, 0, 'L', fill)
        pdf.cell(30, 7, f"{montant:.2f}", 1, 1, 'R', fill)

        fill = not fill

    pdf.set_font('Arial', 'B', 10)
    total_text = 'Total des depenses'.encode('latin-1',
                                             'replace').decode('latin-1')
    pdf.cell(160, 8, total_text, 1, 0, 'R')
    pdf.cell(30, 8, f"{total_depenses:.2f}", 1, 1, 'R')

    buffer = BytesIO()
    buffer.write(pdf.output())
    buffer.seek(0)
    return buffer


def generer_budget_excel(data, date_debut, date_fin):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rapport des Depenses"

    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568",
                              end_color="4A5568",
                              fill_type="solid")
    total_font = Font(name='Calibri', size=11, bold=True)

    center_align = Alignment(horizontal='center',
                             vertical='center',
                             wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border_side = Side(style='thin', color="BFBFBF")
    cell_border = Border(left=thin_border_side,
                         right=thin_border_side,
                         top=thin_border_side,
                         bottom=thin_border_side)
    even_row_fill = PatternFill(start_color="F0F4F8",
                                end_color="F0F4F8",
                                fill_type="solid")

    sheet.merge_cells('B2:E2')
    sheet['B2'] = 'Rapport des Dépenses'
    sheet['B2'].font = title_font
    sheet['B2'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('B3:E3')
    sheet['B3'] = (
        f"Période du "
        f"{datetime.strptime(date_debut, '%Y-%m-%d').strftime('%d/%m/%Y')} au "
        f"{datetime.strptime(date_fin, '%Y-%m-%d').strftime('%d/%m/%Y')}")
    sheet['B3'].font = Font(name='Calibri',
                            size=11,
                            italic=True,
                            color="6c7a89")
    sheet['B3'].alignment = Alignment(horizontal='center')

    headers = ["Date", "Fournisseur", "Contenu de la commande", "Montant (€)"]
    for i, header_text in enumerate(headers, start=2):
        cell = sheet.cell(row=5, column=i, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border

    row_index = 6
    total_depenses = 0
    is_even = False
    for item in data:
        date_val = datetime.strptime(item['date_depense'], '%Y-%m-%d')
        montant = item['montant']
        total_depenses += montant

        sheet.cell(row=row_index, column=2,
                   value=date_val).number_format = 'DD/MM/YYYY'
        sheet.cell(row=row_index,
                   column=3,
                   value=item['fournisseur_nom'] or 'N/A')
        sheet.cell(row=row_index, column=4, value=item['contenu'])
        sheet.cell(row=row_index, column=5,
                   value=montant).number_format = '#,##0.00'

        for col_idx in range(2, 6):
            cell = sheet.cell(row=row_index, column=col_idx)
            cell.border = cell_border
            if col_idx == 5:
                cell.alignment = right_align
            else:
                cell.alignment = center_align

            if is_even:
                cell.fill = even_row_fill

        is_even = not is_even
        row_index += 1

    total_cell_label = sheet.cell(row=row_index,
                                  column=4,
                                  value="Total des dépenses")
    total_cell_label.font = total_font
    total_cell_label.alignment = right_align
    total_cell_label.border = cell_border

    total_cell_value = sheet.cell(row=row_index,
                                  column=5,
                                  value=total_depenses)
    total_cell_value.font = total_font
    total_cell_value.number_format = '#,##0.00'
    total_cell_value.alignment = right_align
    total_cell_value.border = cell_border

    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 50
    sheet.column_dimensions['E'].width = 15

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static', 'icons'),
                               'favicon.ico', mimetype='image/vnd.microsoft.icon')

if __name__ == "__main__":
    import webbrowser
    from threading import Timer

    def open_browser():
        webbrowser.open_new("http://127.0.0.1:5000")

    Timer(1, open_browser).start()
    app.run(debug=False, threaded=True)