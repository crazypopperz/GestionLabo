# -----------------------------------------------------------------------------
# 1. IMPORTS DE LA BIBLIOTHÈQUE STANDARD PYTHON
# -----------------------------------------------------------------------------
import hashlib
import logging
import math
import shutil
import sys
import traceback
import os
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO
from logging.handlers import RotatingFileHandler

# -----------------------------------------------------------------------------
# 2. IMPORTS DES BIBLIOTHÈQUES TIERCES (PIP)
# -----------------------------------------------------------------------------
from flask import (Flask, flash, g, jsonify, redirect, render_template, request,
                   send_file, send_from_directory, session, url_for)
from flask_wtf.csrf import CSRFProtect
from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.security import check_password_hash, generate_password_hash

# -----------------------------------------------------------------------------
# 3. IMPORTS DES MODULES LOCAUX
# -----------------------------------------------------------------------------
from db import get_db
from db import init_app as init_db_app
from utils import (admin_required, get_alerte_info, is_setup_needed,
                   limit_objets_required, login_required)
from views.auth import auth_bp
from views.inventaire import inventaire_bp
from views.admin import admin_bp

class PDFWithFooter(FPDF):
    def footer(self):
        self.set_y(-15)
        # CORRECTION : Utilisation de la police de base et de la nouvelle syntaxe
        self.set_font('Helvetica', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='R')

# --- CONFIGURATION DE L'APPLICATION ---
app = Flask(__name__)
app.config.from_object('config')
init_db_app(app)
app.config['SECRET_KEY'] = os.environ.get('GMLCL_SECRET_KEY', 'une-cle-temporaire-pour-le-developpement-a-changer')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
csrf = CSRFProtect(app)

# ENREGISTREMENT DES BLUEPRINTS
app.register_blueprint(auth_bp)
app.register_blueprint(inventaire_bp)
USER_DATA_PATH = os.path.join(os.environ.get('APPDATA'), 'GMLCL')
os.makedirs(USER_DATA_PATH, exist_ok=True)
app.config['UPLOAD_FOLDER'] = os.path.join(USER_DATA_PATH, 'uploads', 'images')
app.config['FDS_UPLOAD_FOLDER'] = os.path.join(USER_DATA_PATH, 'uploads', 'fds')
DATABASE = os.path.join(USER_DATA_PATH, 'base.db')
app.config['DATABASE'] = DATABASE
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['FDS_UPLOAD_FOLDER'], exist_ok=True)
app.register_blueprint(admin_bp)

if not app.debug:
    log_file_path = os.path.join(USER_DATA_PATH, 'app.log')
    logging.basicConfig(
        level=logging.ERROR,
        format='%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]',
        handlers=[
            RotatingFileHandler(
                log_file_path, maxBytes=1024000, backupCount=5
            )
        ]
    )
    app.logger.info('Logging configuré pour écrire dans %s', log_file_path)

ITEMS_PER_PAGE = 10
CLE_PRO_SECRETE = os.environ.get('GMLCL_PRO_KEY', 'valeur-par-defaut-si-non-definie')


# --- FILTRES JINJA2 PERSONNALISÉS ---
def format_datetime(value, fmt='%d/%m/%Y %H:%M'):
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except (ValueError, TypeError):
            try:
                value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S.%f')
            except (ValueError, TypeError):
                try:
                    value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    return value
    if isinstance(value, (datetime, date)):
        return value.strftime(fmt)
    return value

app.jinja_env.filters['strftime'] = format_datetime

def format_datetime_fr(value, fmt):
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            try:
                value = datetime.fromisoformat(value)
            except (ValueError, TypeError):
                return value
    if not isinstance(value, (datetime, date)):
        return value
    jours = ["lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"]
    mois = ["janvier", "février", "mars", "avril", "mai", "juin", "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
    format_fr = fmt.replace('%A', jours[value.weekday()].capitalize())
    format_fr = format_fr.replace('%B', mois[value.month - 1])
    return value.strftime(format_fr)

app.jinja_env.filters['strftime_fr'] = format_datetime_fr

def annee_scolaire_format(year):
    if isinstance(year, int):
        return f"{year}-{year + 1}"
    return year

app.jinja_env.filters['annee_scolaire'] = annee_scolaire_format

# --- GESTION DE L'INITIALISATION AU PREMIER LANCEMENT ---
@app.before_request
def check_setup():
    if not os.path.exists(DATABASE):
        return
    allowed_endpoints = ['static', 'setup', 'login', 'register']
    if request.endpoint and request.endpoint not in allowed_endpoints:
        if is_setup_needed(app):
            return redirect(url_for('auth.setup'))

# --- FONCTIONS COMMUNES ET PROCESSEUR DE CONTEXTE ---
@app.context_processor
def inject_alert_info():
    # Si l'application n'est pas encore configurée ou si l'utilisateur n'est pas connecté,
    # on renvoie des valeurs par défaut sans interroger la base de données.
    if 'user_id' not in session or is_setup_needed(app):
        return {'alertes_total': 0, 'alertes_stock': 0, 'alertes_peremption': 0}
    
    # Ce bloc ne s'exécute que si l'app est configurée ET l'utilisateur connecté.
    try:
        db = get_db()
        return get_alerte_info(db)
    except sqlite3.Error:
        # En cas d'erreur de base de données inattendue, on évite de planter l'application.
        return {'alertes_total': '!', 'alertes_stock': '!', 'alertes_peremption': '!'}


@app.context_processor
def inject_licence_info():
    """
    Injecte le statut de la licence dans le contexte de tous les templates.
    Rend la variable 'licence' disponible globalement.
    """
    licence_info = {'statut': 'FREE', 'is_pro': False, 'instance_id': 'N/A'}

    # On ne fait rien si la session n'est pas active ou si l'app n'est pas configurée.
    if 'user_id' not in session or is_setup_needed(app):
        return {'licence': licence_info}

    try:
        db = get_db()
        params = db.execute(
            "SELECT cle, valeur FROM parametres "
            "WHERE cle IN ('licence_statut', 'instance_id')").fetchall()
        params_dict = {row['cle']: row['valeur'] for row in params}

        if params_dict.get('licence_statut') == 'PRO':
            licence_info['statut'] = 'PRO'
            licence_info['is_pro'] = True

        if params_dict.get('instance_id'):
            licence_info['instance_id'] = params_dict.get('instance_id')

    except sqlite3.Error as e:
        app.logger.warning(
            f"Impossible de lire les informations de licence. Erreur : {e}"
        )
    
    return {'licence': licence_info}

# --- ROUTES PRINCIPALES ---

@app.route("/jour/<string:date_str>")
@login_required
def vue_jour(date_str):
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash("Format de date invalide.", "error")
        return redirect(url_for('calendrier'))

    db = get_db()
    start_of_day = datetime.strptime(f"{date_str} 00:00:00", '%Y-%m-%d %H:%M:%S')
    end_of_day = datetime.strptime(f"{date_str} 23:59:59", '%Y-%m-%d %H:%M:%S')

    reservations_brutes = db.execute(
        """
        SELECT
            r.groupe_id, r.debut_reservation, r.fin_reservation,
            u.nom_utilisateur
        FROM reservations r
        JOIN utilisateurs u ON r.utilisateur_id = u.id
        WHERE datetime(r.debut_reservation) <= ? AND datetime(r.fin_reservation) > ?
        GROUP BY r.groupe_id, r.debut_reservation, r.fin_reservation, u.nom_utilisateur
        ORDER BY r.debut_reservation
        """, (end_of_day.strftime('%Y-%m-%d %H:%M:%S'), start_of_day.strftime('%Y-%m-%d %H:%M:%S'))).fetchall()

    # --- NOUVELLE LOGIQUE : Préparation des données pour la vue en liste ---
    reservations_par_heure = {hour: {'starts': [], 'continues': []} for hour in range(24)}

    for resa in reservations_brutes:
        debut_dt = datetime.fromisoformat(resa['debut_reservation'])
        fin_dt = datetime.fromisoformat(resa['fin_reservation'])

        # On ne traite que les heures visibles (8h à 20h)
        start_hour = max(8, debut_dt.hour)
        end_hour = min(20, fin_dt.hour if fin_dt.minute > 0 else fin_dt.hour - 1)

        # Si la réservation commence bien ce jour-là
        if debut_dt.date() == date_obj:
            reservations_par_heure[debut_dt.hour]['starts'].append(dict(resa))
        
        # Pour les heures suivantes, on ajoute un bloc "continuation"
        for hour in range(start_hour + 1, end_hour + 1):
            if hour >= 8 and hour <= 20:
                # On vérifie que le bloc n'est pas déjà présent (cas de plusieurs objets pour un même groupe_id)
                if not any(d.get('groupe_id') == resa['groupe_id'] for d in reservations_par_heure[hour]['continues']):
                    reservations_par_heure[hour]['continues'].append(dict(resa))

    return render_template("vue_jour.html",
                           date_concernee=date_obj,
                           reservations_par_heure=reservations_par_heure)


# --- ROUTES SÉCURISÉES ET DE GESTION ---
@app.route("/calendrier")
@login_required
def calendrier():
    db = get_db()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("calendrier.html",
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)

@app.route("/panier")
@login_required
def panier():
    return render_template("panier.html")

@app.route("/alertes")
@login_required
def alertes():
    db = get_db()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # CORRECTION: Calcul du stock disponible pour la liste des alertes de stock
    objets_stock_query = """
        SELECT o.id, o.nom, o.quantite_physique, o.seuil, a.nom AS armoire,
            c.nom AS categorie, o.image, o.en_commande,
            o.date_peremption, o.traite,
            (o.quantite_physique - COALESCE(SUM(r.quantite_reservee), 0)) as quantite_disponible
        FROM objets o 
        JOIN armoires a ON o.armoire_id = a.id
        JOIN categories c ON o.categorie_id = c.id
        LEFT JOIN reservations r ON o.id = r.objet_id AND r.fin_reservation > ?
        GROUP BY o.id
        HAVING quantite_disponible <= o.seuil 
        ORDER BY o.nom
    """
    objets_stock = db.execute(objets_stock_query, (now_str,)).fetchall()

    date_limite = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
    
    # CORRECTION: Calcul du stock disponible pour la liste des péremptions
    objets_peremption_query = """
        SELECT o.id, o.nom, o.quantite_physique, o.seuil, a.nom AS armoire,
               c.nom AS categorie, o.image, o.en_commande, o.date_peremption,
               o.traite,
               (o.quantite_physique - COALESCE(SUM(r.quantite_reservee), 0)) as quantite_disponible
        FROM objets o 
        JOIN armoires a ON o.armoire_id = a.id
        JOIN categories c ON o.categorie_id = c.id
        LEFT JOIN reservations r ON o.id = r.objet_id AND r.fin_reservation > ?
        WHERE o.date_peremption IS NOT NULL AND o.date_peremption < ?
        GROUP BY o.id
        ORDER BY o.date_peremption ASC
    """
    objets_peremption = db.execute(objets_peremption_query, (now_str, date_limite, )).fetchall()
    
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    
    return render_template("alertes.html",
                           objets_stock=objets_stock,
                           objets_peremption=objets_peremption,
                           date_actuelle=datetime.now(),
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@app.route("/a-propos")
@login_required
def a_propos():
    return render_template("a_propos.html")
    
@app.route("/admin/activer_licence", methods=["POST"])
@admin_required
def activer_licence():
    licence_cle_fournie = request.form.get('licence_cle', '').strip()
    db = get_db()

    instance_id_row = db.execute(
        "SELECT valeur FROM parametres WHERE cle = 'instance_id'").fetchone()
    if not instance_id_row:
        flash(
            "Erreur critique : Identifiant d'instance manquant. "
            "Impossible de vérifier la licence.", "error")
        return redirect(url_for('admin.admin'))

    instance_id = instance_id_row['valeur']

    chaine_a_verifier = f"{instance_id}-{CLE_PRO_SECRETE}"
    cle_valide_calculee = hashlib.sha256(
        chaine_a_verifier.encode('utf-8')).hexdigest()

    if licence_cle_fournie == cle_valide_calculee[:16]:
        try:
            db.execute("UPDATE parametres SET valeur = 'PRO' "
                       "WHERE cle = 'licence_statut'")
            db.execute(
                "UPDATE parametres SET valeur = ? WHERE cle = 'licence_cle'",
                (licence_cle_fournie, ))
            db.commit()
            flash(
                "Licence Pro activée avec succès ! Toutes les "
                "fonctionnalités sont maintenant débloquées.", "success")
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données lors de l'activation : {e}",
                  "error")
    else:
        flash(
            "La clé de licence fournie est invalide ou ne correspond pas à "
            "cette installation.", "error")

    return redirect(url_for('admin.admin'))


@app.route("/admin/reset_licence", methods=["POST"])
@admin_required
def reset_licence():
    admin_password = request.form.get('admin_password')
    db = get_db()

    admin_user = db.execute(
        "SELECT mot_de_passe FROM utilisateurs WHERE id = ?",
        (session['user_id'], )).fetchone()

    if not admin_user or not check_password_hash(admin_user['mot_de_passe'],
                                                 admin_password):
        flash(
            "Mot de passe administrateur incorrect. "
            "La réinitialisation a été annulée.", "error")
        return redirect(url_for('admin.admin'))

    try:
        db.execute(
            "UPDATE parametres SET valeur = 'FREE' WHERE cle = 'licence_statut'"
        )
        db.execute(
            "UPDATE parametres SET valeur = '' WHERE cle = 'licence_cle'")
        db.commit()
        flash("La licence a été réinitialisée au statut GRATUIT.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données lors de la réinitialisation : {e}",
              "error")

    return redirect(url_for('admin.admin'))


@app.route("/budget/exporter")
@admin_required
def exporter_budget():
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    format_type = request.args.get('format')

    if not all([date_debut, date_fin, format_type]):
        flash("Tous les champs sont requis pour l'export.", "error")
        return redirect(url_for('budget'))

    db = get_db()
    depenses_data = db.execute(
        """
        SELECT d.date_depense, d.contenu, d.montant, f.nom as fournisseur_nom
        FROM depenses d
        LEFT JOIN fournisseurs f ON d.fournisseur_id = f.id
        WHERE d.date_depense BETWEEN ? AND ?
        ORDER BY d.date_depense ASC
        """, (date_debut, date_fin)).fetchall()

    if not depenses_data:
        flash("Aucune dépense trouvée pour la période sélectionnée.",
              "warning")
        return redirect(url_for('budget'))

    if format_type == 'pdf':
        buffer = generer_budget_pdf(depenses_data, date_debut, date_fin)
        return send_file(buffer,
                         as_attachment=True,
                         download_name='rapport_depenses.pdf',
                         mimetype='application/pdf')
    elif format_type == 'excel':
        buffer = generer_budget_excel(depenses_data, date_debut, date_fin)
        return send_file(
            buffer,
            as_attachment=True,
            download_name='rapport_depenses.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    flash("Format d'exportation non valide.", "error")
    return redirect(url_for('budget'))
    

# --- ROUTES POUR LES RAPPORTS ---
@app.route("/objet/<int:objet_id>/telecharger_fds")
@login_required
def telecharger_fds_objet(objet_id):
    db = get_db()
    objet = db.execute(
        "SELECT fds_nom_original, fds_nom_securise FROM objets WHERE id = ?",
        (objet_id, )).fetchone()
    if objet and objet['fds_nom_securise']:
        try:
            return send_file(os.path.join(app.config['FDS_UPLOAD_FOLDER'],
                                          objet['fds_nom_securise']),
                             as_attachment=True,
                             download_name=objet['fds_nom_original'])
        except FileNotFoundError:
            flash("Le fichier FDS n'a pas été trouvé sur le serveur.", "error")
            return redirect(url_for('inventaire.voir_objet', objet_id=objet_id))
    else:
        flash("Cet objet n'a pas de FDS associée.", "error")
        return redirect(url_for('inventaire.voir_objet', objet_id=objet_id))


# --- ROUTES API ---
@app.route("/api/rechercher")
@login_required
def api_rechercher():
    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify([])
    db = get_db()
    resultats = db.execute(
        """SELECT o.id, o.nom, o.armoire_id, a.nom as armoire_nom,
                  c.nom as categorie_nom
           FROM objets o
           JOIN armoires a ON o.armoire_id = a.id
           JOIN categories c ON o.categorie_id = c.id
           WHERE unaccent(LOWER(o.nom)) LIKE unaccent(LOWER(?))
           LIMIT 10""", (f"%{query}%", )).fetchall()
    return jsonify([dict(row) for row in resultats])


@app.route("/api/reservations_par_mois/<int:year>/<int:month>")
@login_required
def api_reservations_par_mois(year, month):
    db = get_db()

    start_date_str = f"{year}-{str(month).zfill(2)}-01 00:00:00"

    if month == 12:
        end_date_str = f"{year + 1}-01-01 00:00:00"
    else:
        end_date_str = f"{year}-{str(month + 1).zfill(2)}-01 00:00:00"

    reservations = db.execute(
        """
        SELECT
            DATE(r.debut_reservation) as jour_reservation,
            r.groupe_id,
            r.utilisateur_id,
            u.nom_utilisateur
        FROM reservations r
        JOIN utilisateurs u ON r.utilisateur_id = u.id
        WHERE r.debut_reservation >= ? AND r.debut_reservation < ?
              AND r.groupe_id IS NOT NULL
        GROUP BY jour_reservation, r.groupe_id
        """, (start_date_str, end_date_str)).fetchall()

    results = {}
    for row in reservations:
        date = row['jour_reservation']
        if date not in results:
            results[date] = []

        results[date].append(
            {'is_mine': row['utilisateur_id'] == session['user_id']})
    return jsonify(results)


@app.route("/api/reservation_details/<groupe_id>")
@login_required
def api_reservation_details(groupe_id):
    db = get_db()
    reservations = db.execute(
        """
        SELECT r.quantite_reservee, o.id as objet_id, o.nom as objet_nom, 
               u.nom_utilisateur, r.utilisateur_id, r.kit_id, k.nom as kit_nom,
               r.debut_reservation, r.fin_reservation
        FROM reservations r
        JOIN objets o ON r.objet_id = o.id
        JOIN utilisateurs u ON r.utilisateur_id = u.id
        LEFT JOIN kits k ON r.kit_id = k.id
        WHERE r.groupe_id = ?
        """, (groupe_id, )).fetchall()

    if not reservations:
        return jsonify({'error': 'Réservation non trouvée'}), 404

    details = {
        'kits': {},
        'objets_manuels': [],
        'nom_utilisateur': reservations[0]['nom_utilisateur'],
        'utilisateur_id': reservations[0]['utilisateur_id'],
        'debut_reservation': reservations[0]['debut_reservation'],
        'fin_reservation': reservations[0]['fin_reservation']
    }

    objets_manuels_bruts = [dict(r) for r in reservations if r['kit_id'] is None]
    objets_kits_reserves = [r for r in reservations if r['kit_id'] is not None]

    kits_comptes = {}
    for r in objets_kits_reserves:
        if r['kit_id'] not in kits_comptes:
            kits_comptes[r['kit_id']] = {'nom': r['kit_nom'], 'objets_reserves': {}}
        kits_comptes[r['kit_id']]['objets_reserves'][r['objet_id']] = r['quantite_reservee']

    for kit_id, data in kits_comptes.items():
        objets_base_du_kit = db.execute("SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?", (kit_id,)).fetchall()
        if not objets_base_du_kit: continue
        
        id_objet_calcul, quantite_par_kit = next(((obj['objet_id'], obj['quantite']) for obj in objets_base_du_kit if obj['objet_id'] in data['objets_reserves']), (None, 0))

        if id_objet_calcul and quantite_par_kit > 0:
            quantite_reelle_reservee = data['objets_reserves'][id_objet_calcul]
            nombre_de_kits = quantite_reelle_reservee // quantite_par_kit
            details['kits'][str(kit_id)] = {'quantite': nombre_de_kits, 'nom': data['nom']}

    objets_manuels_agreges = {}
    for item in objets_manuels_bruts:
        obj_id = item['objet_id']
        nom = item['objet_nom'] 
        quantite = item['quantite_reservee']
        
        if obj_id not in objets_manuels_agreges:
            objets_manuels_agreges[obj_id] = {'nom': nom, 'quantite_reservee': 0}
        objets_manuels_agreges[obj_id]['quantite_reservee'] += quantite

    for obj_id, data in objets_manuels_agreges.items():
        details['objets_manuels'].append({
            'objet_id': obj_id,
            'nom': data['nom'],
            'quantite_reservee': data['quantite_reservee']
        })

    return jsonify(details)


@app.route("/api/reservation_data/<date>")
@login_required
def api_reservation_data(date):
    db = get_db()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    objets_bruts = db.execute(
        f"""
        SELECT 
            o.id, o.nom, c.nom as categorie, o.quantite_physique,
            COALESCE((SELECT SUM(r.quantite_reservee) 
                      FROM reservations r 
                      WHERE r.objet_id = o.id AND r.fin_reservation > '{now_str}'), 0) as total_reserve
        FROM objets o
        JOIN categories c ON o.categorie_id = c.id
        ORDER BY c.nom, o.nom
        """
    ).fetchall()
    
    grouped_objets = {}
    objets_map = {}
    for row in objets_bruts:
        categorie_nom = row['categorie']
        if categorie_nom not in grouped_objets:
            grouped_objets[categorie_nom] = []
        
        quantite_disponible = row['quantite_physique'] - row['total_reserve']
        
        obj_data = {
            "id": row['id'],
            "nom": row['nom'],
            "quantite_totale": row['quantite_physique'],
            "quantite_disponible": quantite_disponible if quantite_disponible >= 0 else 0
        }
        grouped_objets[categorie_nom].append(obj_data)
        objets_map[row['id']] = obj_data

    kits = db.execute("SELECT id, nom, description FROM kits ORDER BY nom").fetchall()
    kits_details = []
    for kit in kits:
        objets_du_kit = db.execute("SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?", (kit['id'],)).fetchall()
        
        disponibilite_kit = 9999
        if not objets_du_kit:
            disponibilite_kit = 0
        else:
            for obj_in_kit in objets_du_kit:
                objet_data = objets_map.get(obj_in_kit['objet_id'])
                if not objet_data or obj_in_kit['quantite'] == 0:
                    disponibilite_kit = 0
                    break
                
                kits_possibles = math.floor(objet_data['quantite_disponible'] / obj_in_kit['quantite'])
                if kits_possibles < disponibilite_kit:
                    disponibilite_kit = kits_possibles

        kits_details.append({
            'id': kit['id'],
            'nom': kit['nom'],
            'description': kit['description'],
            'objets': [dict(o) for o in objets_du_kit],
            'disponibilite': disponibilite_kit if disponibilite_kit != 9999 else 0
        })

    return jsonify({'objets': grouped_objets, 'kits': kits_details})


# Cette route est obsolète mais nous la laissons vide pour ne pas créer d'erreur 404 si elle est appelée
@app.route("/api/reserver", methods=["POST"])
@login_required
def api_reserver():
    return jsonify(success=False, error="Cette route est obsolète."), 400


@app.route("/api/modifier_reservation", methods=["POST"])
@login_required
def api_modifier_reservation():
    data = request.get_json()
    groupe_id = data.get("groupe_id")
    db = get_db()
    
    try:
        # On utilise une transaction pour s'assurer que tout réussit ou tout échoue
        db.execute("BEGIN")
        
        # Étape 1: Supprimer l'ancienne réservation
        db.execute("DELETE FROM reservations WHERE groupe_id = ?", (groupe_id,))
        
        # Étape 2: Valider et insérer la nouvelle réservation comme un "mini-panier"
        creneau_key = f"{data['date']}_{data['heure_debut']}_{data['heure_fin']}"
        mini_cart = { creneau_key: data }
        
        response = api_valider_panier_interne(mini_cart, groupe_id_existant=groupe_id) 
        
        if response.get('success'):
            db.commit()
            flash("Réservation modifiée avec succès !", "success")
            return jsonify(success=True)
        else:
            db.rollback() # Annule la suppression si la nouvelle réservation échoue
            return jsonify(success=False, error=response.get('error', 'Erreur inconnue lors de la modification')), 400

    except Exception as e:
        db.rollback()
        traceback.print_exc()
        return jsonify(success=False, error=f"Une erreur interne est survenue : {e}"), 500

def api_valider_panier_interne(cart_data, groupe_id_existant=None):
    db = get_db()
    user_id = session['user_id']
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        objets_requis = {}
        for creneau_key, resa_details in cart_data.items():
            for kit_id, kit_data in resa_details.get('kits', {}).items():
                objets_du_kit = db.execute("SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?", (kit_id,)).fetchall()
                for obj_in_kit in objets_du_kit:
                    objets_requis[obj_in_kit['objet_id']] = objets_requis.get(obj_in_kit['objet_id'], 0) + (obj_in_kit['quantite'] * kit_data.get('quantite', 0))
            for obj_id, obj_data in resa_details.get('objets', {}).items():
                objets_requis[int(obj_id)] = objets_requis.get(int(obj_id), 0) + obj_data.get('quantite', 0)

        for obj_id, quantite_demandee in objets_requis.items():
            if quantite_demandee <= 0: continue
            stock_info = db.execute("""SELECT o.nom, (o.quantite_physique - COALESCE((SELECT SUM(r.quantite_reservee) FROM reservations r WHERE r.objet_id = o.id AND r.fin_reservation > ?), 0)) as quantite_disponible
                                        FROM objets o WHERE o.id = ?""", (now_str, objet_id,)).fetchone()
            if not stock_info or stock_info['quantite_disponible'] < quantite_demandee:
                return {'success': False, 'error': f"Stock insuffisant pour '{stock_info['nom']}'"}

        for creneau_key, resa_details in cart_data.items():
            groupe_id = groupe_id_existant or str(uuid.uuid4())
            debut_dt = datetime.strptime(f"{resa_details['date']} {resa_details['heure_debut']}", '%Y-%m-%d %H:%M')
            fin_dt = datetime.strptime(f"{resa_details['date']} {resa_details['heure_fin']}", '%Y-%m-%d %H:%M')
            
            final_reservations = {}
            for kit_id_str, kit_data in resa_details.get('kits', {}).items():
                objets_du_kit = db.execute("SELECT objet_id, quantite FROM kit_objets WHERE kit_id = ?", (kit_id_str,)).fetchall()
                for obj_in_kit in objets_du_kit:
                    key = (obj_in_kit['objet_id'], int(kit_id_str))
                    final_reservations[key] = final_reservations.get(key, 0) + (obj_in_kit['quantite'] * kit_data.get('quantite', 0))
            for obj_id_str, obj_data in resa_details.get('objets', {}).items():
                key = (int(obj_id_str), None)
                final_reservations[key] = final_reservations.get(key, 0) + obj_data.get('quantite', 0)

            for (obj_id, kit_id), quantite_totale in final_reservations.items():
                if quantite_totale > 0:
                    db.execute(
                        """INSERT INTO reservations (objet_id, quantite_reservee, debut_reservation, fin_reservation, utilisateur_id, groupe_id, kit_id)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (obj_id, quantite_totale, debut_dt, fin_dt, user_id, groupe_id, kit_id))
                    action = "Modification Réservation" if groupe_id_existant else "Réservation"
                    enregistrer_action(obj_id, action, f"Quantité: {quantite_totale} pour le {debut_dt.strftime('%d/%m/%Y')}")
        
        return {'success': True}
    except Exception as e:
        traceback.print_exc()
        return {'success': False, 'error': f"Erreur interne: {e}"}

@app.route("/api/valider_panier", methods=["POST"])
@login_required
def api_valider_panier():
    cart_data = request.get_json()
    if not cart_data:
        return jsonify(success=False, error="Le panier est vide."), 400

    db = get_db()
    try:
        # On utilise une transaction pour s'assurer que l'ensemble du panier est validé ou rien du tout.
        db.execute("BEGIN")
        response = api_valider_panier_interne(cart_data)
        
        if response.get('success'):
            db.commit()
            flash("Toutes vos réservations ont été confirmées avec succès !", "success")
            return jsonify(success=True)
        else:
            db.rollback()
            return jsonify(success=False, error=response.get('error')), 400
            
    except Exception as e:
        db.rollback()
        traceback.print_exc()
        return jsonify(success=False, error=f"Une erreur interne est survenue : {e}"), 500

@app.route("/api/supprimer_reservation", methods=["POST"])
@login_required
def api_supprimer_reservation():
    data = request.get_json()
    groupe_id = data.get("groupe_id")
    db = get_db()
    try:
        reservation_info = db.execute("SELECT utilisateur_id FROM reservations WHERE groupe_id = ? LIMIT 1", (groupe_id, )).fetchone()
        if not reservation_info:
            return jsonify(success=False, error="Réservation non trouvée."), 404
        if (session.get('user_role') != 'admin' and reservation_info['utilisateur_id'] != session['user_id']):
            return jsonify(success=False, error="Vous n'avez pas la permission de supprimer cette réservation."), 403
        
        db.execute("DELETE FROM reservations WHERE groupe_id = ?", (groupe_id,))
        db.commit()
        
        flash("La réservation a été annulée.", "success")
        return jsonify(success=True)
    except sqlite3.Error as e:
        db.rollback()
        return jsonify(success=False, error=str(e)), 500





@app.route("/api/suggestion_commande/<int:objet_id>")
@admin_required
def api_suggestion_commande(objet_id):
    db = get_db()

    date_limite = datetime.now() - timedelta(days=90)

    result = db.execute(
        """
        SELECT SUM(quantite_reservee)
        FROM reservations
        WHERE objet_id = ? AND debut_reservation >= ?
        """, (objet_id, date_limite)).fetchone()

    consommation = result[0] if result and result[0] is not None else 0

    suggestion = 0
    if consommation > 0:
        suggestion = math.ceil(consommation * 1.5)
    else:
        objet = db.execute("SELECT seuil FROM objets WHERE id = ?",
                           (objet_id, )).fetchone()
        if objet:
            suggestion = objet['seuil'] * 2
        else:
            suggestion = 5

    return jsonify(suggestion=suggestion, consommation=consommation)

# --- ROUTE POUR SERVIR LES IMAGES UPLOADÉES ---
from flask import send_from_directory

@app.route('/uploads/images/<path:filename>')
@login_required
def serve_image(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# --- ROUTES D'IMPORT/EXPORT ET EXÉCUTION ---
@app.route("/telecharger_db")
@admin_required
def telecharger_db():
    db = get_db()
    licence_row = db.execute("SELECT valeur FROM parametres WHERE cle = ?",
                             ('licence_statut', )).fetchone()
    is_pro = licence_row and licence_row['valeur'] == 'PRO'

    if not is_pro:
        flash(
            "Le téléchargement de la base de données est une fonctionnalité "
            "de la version Pro.", "warning")
        return redirect(url_for('admin.admin'))

    return send_file(DATABASE, as_attachment=True)


@app.route("/importer_db", methods=["POST"])
@admin_required
def importer_db():
    if 'fichier' not in request.files:
        flash("Aucun fichier sélectionné.", "error")
        return redirect(url_for('admin.admin'))
    fichier = request.files.get("fichier")
    if not fichier or fichier.filename == '':
        flash("Aucun fichier selecté.", "error")
        return redirect(url_for('admin.admin'))
    if fichier and fichier.filename.endswith(".db"):
        temp_db_path = DATABASE + ".tmp"
        fichier.save(temp_db_path)
        shutil.move(temp_db_path, DATABASE)
        flash("Base de données importée avec succès !", "success")
    else:
        flash("Le fichier fourni n'est pas une base de données valide (.db).",
              "error")
    return redirect(url_for('admin.admin'))


@app.route("/admin/exporter")
@admin_required
def generer_pdf(data):
    # ... (le début de la fonction est identique)
    for categorie, items in sorted(grouped_data.items()):
        # ...
        for i, item in enumerate(items):
            # ...
            pdf.cell(85, 7, item['nom'].encode('latin-1', 'replace').decode('latin-1'), 1, 0)
            # CORRECTION: Utiliser la bonne clé pour la quantité
            pdf.cell(20, 7, str(item['quantite_disponible']), 1, 0, 'C')
            pdf.cell(40, 7, item['armoire'].encode('latin-1', 'replace').decode('latin-1'), 1, 0)
            # ...
    return BytesIO(pdf.output())

def generer_excel(data):
    # ... (le début de la fonction est identique)
    for item in data:
        # ...
        # CORRECTION: Utiliser la bonne clé pour la quantité
        row_data = [
            item['categorie'], item['nom'], item['quantite_disponible'], item['armoire'],
            date_peremption_val
        ]
        for i, value in enumerate(row_data):
            cell = sheet.cell(row=row_index, column=start_col + i, value=value)
            cell.border = thin_border
            cell.alignment = center_align if i > 0 else Alignment(
                vertical='center')
            if isinstance(value, datetime):
                cell.number_format = 'DD/MM/YYYY'
            if is_even:
                cell.fill = even_row_fill
        if item['categorie'] != current_cat:
            if current_cat is not None and start_merge_row < row_index:
                sheet.merge_cells(start_row=start_merge_row,
                                  start_column=start_col,
                                  end_row=row_index - 1,
                                  end_column=start_col)
                cell = sheet.cell(row=start_merge_row, column=start_col)
                cell.font = category_font
                cell.alignment = category_align
            current_cat = item['categorie']
            start_merge_row = row_index
        row_index += 1
        is_even = not is_even
    if current_cat is not None and start_merge_row < row_index - 1:
        sheet.merge_cells(start_row=start_merge_row,
                          start_column=start_col,
                          end_row=row_index - 1,
                          end_column=start_col)
        cell = sheet.cell(row=start_merge_row, column=start_col)
        cell.font = category_font
        cell.alignment = category_align
    column_widths = {'B': 30, 'C': 50, 'D': 15, 'E': 30, 'F': 20}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = 'A5'
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generer_budget_pdf(data, date_debut, date_fin):
    pdf = PDFWithFooter()
    pdf.alias_nb_pages()
    pdf.add_page(orientation='P')
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'Rapport des Depenses', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(
        0, 10,
        f"Periode du {datetime.strptime(date_debut, '%Y-%m-%d').strftime('%d/%m/%Y')} "
        f"au {datetime.strptime(date_fin, '%Y-%m-%d').strftime('%d/%m/%Y')}",
        new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(10)

    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(25, 8, 'Date', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(50, 8, 'Fournisseur', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(85, 8, 'Contenu de la commande'.encode('latin-1', 'replace').decode('latin-1'), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(30, 8, 'Montant (EUR)', border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C', fill=True)

    pdf.set_font('Helvetica', '', 9)
    total_depenses = 0
    fill = False
    for item in data:
        pdf.set_fill_color(240, 240, 240)
        date_str = datetime.strptime(item['date_depense'], '%Y-%m-%d').strftime('%d/%m/%Y')
        fournisseur = (item['fournisseur_nom'] or 'N/A').encode('latin-1', 'replace').decode('latin-1')
        contenu = item['contenu'].encode('latin-1', 'replace').decode('latin-1')
        montant = item['montant']
        total_depenses += montant

        pdf.cell(25, 7, date_str, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=fill)
        pdf.cell(50, 7, fournisseur, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=fill)
        pdf.cell(85, 7, contenu, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='L', fill=fill)
        pdf.cell(30, 7, f"{montant:.2f}", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R', fill=fill)
        fill = not fill

    pdf.set_font('Helvetica', 'B', 10)
    total_text = 'Total des depenses'.encode('latin-1', 'replace').decode('latin-1')
    pdf.cell(160, 8, total_text, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='R')
    pdf.cell(30, 8, f"{total_depenses:.2f}", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')

    return BytesIO(pdf.output())


def generer_budget_excel(data, date_debut, date_fin):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rapport des Depenses"

    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568",
                              end_color="4A5568",
                              fill_type="solid")
    total_font = Font(name='Calibri', size=11, bold=True)

    center_align = Alignment(horizontal='center',
                             vertical='center',
                             wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border_side = Side(style='thin', color="BFBFBF")
    cell_border = Border(left=thin_border_side,
                         right=thin_border_side,
                         top=thin_border_side,
                         bottom=thin_border_side)
    even_row_fill = PatternFill(start_color="F0F4F8",
                                end_color="F0F4F8",
                                fill_type="solid")

    sheet.merge_cells('B2:E2')
    sheet['B2'] = 'Rapport des Dépenses'
    sheet['B2'].font = title_font
    sheet['B2'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('B3:E3')
    sheet['B3'] = (
        f"Période du "
        f"{datetime.strptime(date_debut, '%Y-%m-%d').strftime('%d/%m/%Y')} au "
        f"{datetime.strptime(date_fin, '%Y-%m-%d').strftime('%d/%m/%Y')}")
    sheet['B3'].font = Font(name='Calibri',
                            size=11,
                            italic=True,
                            color="6c7a89")
    sheet['B3'].alignment = Alignment(horizontal='center')

    headers = ["Date", "Fournisseur", "Contenu de la commande", "Montant (€)"]
    for i, header_text in enumerate(headers, start=2):
        cell = sheet.cell(row=5, column=i, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border

    row_index = 6
    total_depenses = 0
    is_even = False
    for item in data:
        date_val = datetime.strptime(item['date_depense'], '%Y-%m-%d')
        montant = item['montant']
        total_depenses += montant

        sheet.cell(row=row_index, column=2,
                   value=date_val).number_format = 'DD/MM/YYYY'
        sheet.cell(row=row_index,
                   column=3,
                   value=item['fournisseur_nom'] or 'N/A')
        sheet.cell(row=row_index, column=4, value=item['contenu'])
        sheet.cell(row=row_index, column=5,
                   value=montant).number_format = '#,##0.00'

        for col_idx in range(2, 6):
            cell = sheet.cell(row=row_index, column=col_idx)
            cell.border = cell_border
            if col_idx == 5:
                cell.alignment = right_align
            else:
                cell.alignment = center_align

            if is_even:
                cell.fill = even_row_fill

        is_even = not is_even
        row_index += 1

    total_cell_label = sheet.cell(row=row_index,
                                  column=4,
                                  value="Total des dépenses")
    total_cell_label.font = total_font
    total_cell_label.alignment = right_align
    total_cell_label.border = cell_border

    total_cell_value = sheet.cell(row=row_index,
                                  column=5,
                                  value=total_depenses)
    total_cell_value.font = total_font
    total_cell_value.number_format = '#,##0.00'
    total_cell_value.alignment = right_align
    total_cell_value.border = cell_border

    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 50
    sheet.column_dimensions['E'].width = 15

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static', 'icons'),
                               'favicon.ico', mimetype='image/vnd.microsoft.icon')


if __name__ == "__main__":
    import webbrowser
    from threading import Timer

    def open_browser():
        webbrowser.open_new("http://127.0.0.1:5000")

    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        Timer(1, open_browser).start()

    app.run(debug=True, threaded=True)