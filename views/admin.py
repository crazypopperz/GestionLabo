# ============================================================
# IMPORTS
# ============================================================

# Imports depuis la bibliothèque standard
import hashlib
import os
from datetime import datetime, date, timedelta

# Imports depuis les bibliothèques tierces (Flask, etc.)
from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, session, jsonify, send_file)
from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

# Imports depuis nos propres modules
from db import get_db
from utils import admin_required, login_required
# On importera d'autres fonctions de utils au besoin

# ============================================================
# CRÉATION DU BLUEPRINT POUR L'ADMINISTRATION
# ============================================================
# On utilise url_prefix pour que toutes les routes de ce blueprint
# commencent automatiquement par /admin
admin_bp = Blueprint(
    'admin', 
    __name__,
    template_folder='../templates',
    url_prefix='/admin'
)

# ============================================================
# LES FONCTIONS DE ROUTES ADMIN
# ============================================================
@admin_bp.route("/admin")
@admin_required
def admin():
    db = get_db()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("admin.html",
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)

#=== GESTION UTILISATEURS ===
@admin_bp.route("/utilisateurs")
@admin_required
def gestion_utilisateurs():
    db = get_db()
    utilisateurs = db.execute(
        "SELECT id, nom_utilisateur, role, email FROM utilisateurs "
        "ORDER BY nom_utilisateur").fetchall()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("admin_utilisateurs.html",
                           utilisateurs=utilisateurs,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


#=== MODIFIER EMAIL ===
@admin_bp.route("/utilisateurs/modifier_email/<int:id_user>",
           methods=["POST"])
@admin_required
def modifier_email_utilisateur(id_user):
    email = request.form.get('email', '').strip()
    if not email or '@' not in email:
        flash("Veuillez fournir une adresse e-mail valide.", "error")
        return redirect(url_for('admin.gestion_utilisateurs'))

    db = get_db()
    user = db.execute("SELECT nom_utilisateur FROM utilisateurs WHERE id = ?",
                      (id_user, )).fetchone()
    if not user:
        flash("Utilisateur non trouvé.", "error")
        return redirect(url_for('admin.gestion_utilisateurs'))

    try:
        db.execute("UPDATE utilisateurs SET email = ? WHERE id = ?",
                   (email, id_user))
        db.commit()
        flash(
            f"L'adresse e-mail pour '{user['nom_utilisateur']}' a été "
            "mise à jour.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('admin.gestion_utilisateurs'))

#=== SUPPRIMER UTILISATEUR ===
@admin_bp.route("/utilisateurs/supprimer/<int:id_user>", methods=["POST"])
@admin_required
def supprimer_utilisateur(id_user):
    if id_user == session['user_id']:
        flash("Vous ne pouvez pas supprimer votre propre compte.", "error")
        return redirect(url_for('admin.gestion_utilisateurs'))
    db = get_db()
    user = db.execute("SELECT nom_utilisateur FROM utilisateurs WHERE id = ?",
                      (id_user, )).fetchone()
    if user:
        db.execute("DELETE FROM utilisateurs WHERE id = ?", (id_user, ))
        db.commit()
        flash(f"L'utilisateur '{user['nom_utilisateur']}' a été supprimé.",
              "success")
    else:
        flash("Utilisateur non trouvé.", "error")
    return redirect(url_for('admin.gestion_utilisateurs'))

#=== PROMOUVOIR UTILISATEUR ===
@admin_bp.route("/utilisateurs/promouvoir/<int:id_user>", methods=["POST"])
@admin_required
def promouvoir_utilisateur(id_user):
    if id_user == session['user_id']:
        flash("Action non autorisée sur votre propre compte.", "error")
        return redirect(url_for('admin.gestion_utilisateurs'))
    password = request.form.get('password')
    db = get_db()
    admin_actuel = db.execute(
        "SELECT mot_de_passe FROM utilisateurs WHERE id = ?",
        (session['user_id'], )).fetchone()
    if not admin_actuel or not check_password_hash(
            admin_actuel['mot_de_passe'], password):
        flash(
            "Mot de passe administrateur incorrect. "
            "La passation de pouvoir a échoué.", "error")
        return redirect(url_for('admin.gestion_utilisateurs'))
    try:
        db.execute("UPDATE utilisateurs SET role = 'admin' WHERE id = ?",
                   (id_user, ))
        db.execute("UPDATE utilisateurs SET role = 'utilisateur' WHERE id = ?",
                   (session['user_id'], ))
        db.commit()
        flash(
            "Passation de pouvoir réussie ! "
            "Vous êtes maintenant un utilisateur standard.", "success")
        return redirect(url_for('auth.logout'))
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Une erreur est survenue lors de la passation de pouvoir : {e}",
              "error")
        return redirect(url_for('admin.gestion_utilisateurs'))

#=== REINITIALISER MDP ===
@admin_bp.route("/utilisateurs/reinitialiser_mdp/<int:id_user>",
           methods=["POST"])
@admin_required
def reinitialiser_mdp(id_user):
    if id_user == session['user_id']:
        flash(
            "Vous ne pouvez pas réinitialiser votre propre mot de passe ici.",
            "error")
        return redirect(url_for('admin.gestion_utilisateurs'))
    nouveau_mdp = request.form.get('nouveau_mot_de_passe')
    if not nouveau_mdp or len(nouveau_mdp) < 4:
        flash(
            "Le nouveau mot de passe est requis et doit contenir "
            "au moins 4 caractères.", "error")
        return redirect(url_for('admin.gestion_utilisateurs'))
    db = get_db()
    user = db.execute("SELECT nom_utilisateur FROM utilisateurs WHERE id = ?",
                      (id_user, )).fetchone()
    if user:
        try:
            db.execute("UPDATE utilisateurs SET mot_de_passe = ? WHERE id = ?",
                       (generate_password_hash(nouveau_mdp,
                                               method='scrypt'), id_user))
            db.commit()
            flash(
                f"Le mot de passe pour l'utilisateur "
                f"'{user['nom_utilisateur']}' a été réinitialisé avec succès.",
                "success")
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données : {e}", "error")
    else:
        flash("Utilisateur non trouvé.", "error")
    return redirect(url_for('admin.gestion_utilisateurs'))


#===============================
# GESTION KITS
#===============================
@admin_bp.route("/kits")
@admin_required
def gestion_kits():
    db = get_db()
    kits = db.execute("""
        SELECT k.id, k.nom, k.description, COUNT(ko.id) as count
        FROM kits k
        LEFT JOIN kit_objets ko ON k.id = ko.kit_id
        GROUP BY k.id, k.nom, k.description
        ORDER BY k.nom
        """).fetchall()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("admin_kits.html",
                           kits=kits,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)

#=== AJOUTER KIT ===
@admin_bp.route("/kits/ajouter", methods=["POST"])
@admin_required
def ajouter_kit():
    nom = request.form.get("nom", "").strip()
    description = request.form.get("description", "").strip()
    if not nom:
        flash("Le nom du kit ne peut pas être vide.", "error")
        return redirect(url_for('admin.gestion_kits'))

    db = get_db()
    try:
        cursor = db.execute(
            "INSERT INTO kits (nom, description) VALUES (?, ?)",
            (nom, description))
        db.commit()
        new_kit_id = cursor.lastrowid
        flash(
            f"Le kit '{nom}' a été créé. "
            "Vous pouvez maintenant y ajouter des objets.", "success")
        return redirect(url_for('admin.modifier_kit', kit_id=new_kit_id))
    except sqlite3.IntegrityError:
        flash(f"Un kit avec le nom '{nom}' existe déjà.", "error")
        return redirect(url_for('admin.gestion_kits'))

#=== MODIFIER KIT ===
@admin_bp.route("/kits/modifier/<int:kit_id>", methods=["GET", "POST"])
@admin_required
def modifier_kit(kit_id):
    db = get_db()
    kit = db.execute("SELECT * FROM kits WHERE id = ?", (kit_id, )).fetchone()
    if not kit:
        flash("Kit non trouvé.", "error")
        return redirect(url_for('admin.gestion_kits'))

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if request.method == "POST":
        objet_id_str = request.form.get("objet_id")
        quantite_str = request.form.get("quantite")

        if objet_id_str and quantite_str:
            try:
                objet_id = int(objet_id_str)
                quantite = int(quantite_str)
                stock_info = db.execute(
                    f"""
                    SELECT o.nom, (o.quantite_physique - COALESCE((SELECT SUM(r.quantite_reservee) FROM reservations r WHERE r.objet_id = o.id AND r.fin_reservation > '{now_str}'), 0)) as quantite_disponible
                    FROM objets o WHERE o.id = ?
                    """, (objet_id,)
                ).fetchone()

                if not stock_info:
                    flash("Objet non trouvé.", "error")
                    return redirect(url_for('admin.modifier_kit', kit_id=kit_id))

                if quantite > stock_info['quantite_disponible']:
                    flash(f"Quantité invalide pour '{stock_info['nom']}'. Vous ne pouvez pas ajouter plus que le stock disponible ({stock_info['quantite_disponible']}).", "error")
                    return redirect(url_for('admin.modifier_kit', kit_id=kit_id))

                existing = db.execute("SELECT id FROM kit_objets WHERE kit_id = ? AND objet_id = ?", (kit_id, objet_id)).fetchone()
                if existing:
                    db.execute("UPDATE kit_objets SET quantite = ? WHERE id = ?", (quantite, existing['id']))
                else:
                    db.execute("INSERT INTO kit_objets (kit_id, objet_id, quantite) VALUES (?, ?, ?)", (kit_id, objet_id, quantite))
                db.commit()
                flash(f"L'objet '{stock_info['nom']}' a été ajouté/mis à jour dans le kit.", "success")

            except (ValueError, TypeError):
                flash("Données invalides.", "error")
            
            return redirect(url_for('admin.modifier_kit', kit_id=kit_id))

        for key, value in request.form.items():
            if key.startswith("quantite_"):
                try:
                    kit_objet_id = int(key.split("_")[1])
                    new_quantite = int(value)

                    objet_info = db.execute(
                        f"""
                        SELECT o.nom, o.id as objet_id, (o.quantite_physique - COALESCE((SELECT SUM(r.quantite_reservee) FROM reservations r WHERE r.objet_id = o.id AND r.fin_reservation > '{now_str}'), 0)) as quantite_disponible
                        FROM kit_objets ko JOIN objets o ON ko.objet_id = o.id
                        WHERE ko.id = ?
                        """, (kit_objet_id,)
                    ).fetchone()

                    if not objet_info: continue

                    if new_quantite > objet_info['quantite_disponible']:
                         flash(f"Quantité invalide pour '{objet_info['nom']}'. Vous ne pouvez pas dépasser le stock disponible ({objet_info['quantite_disponible']}).", "error")
                    else:
                        db.execute("UPDATE kit_objets SET quantite = ? WHERE id = ?", (new_quantite, kit_objet_id))
                        flash(f"Quantité pour '{objet_info['nom']}' mise à jour.", "success")
                
                except (ValueError, TypeError):
                    flash("Une quantité fournie est invalide.", "error")
        
        db.commit()
        return redirect(url_for('admin.modifier_kit', kit_id=kit_id))

    objets_in_kit = db.execute(
        f"""
        SELECT ko.id, o.nom, 
               (o.quantite_physique - COALESCE((SELECT SUM(r.quantite_reservee) FROM reservations r WHERE r.objet_id = o.id AND r.fin_reservation > '{now_str}'), 0)) as stock_disponible, 
               ko.quantite
        FROM kit_objets ko
        JOIN objets o ON ko.objet_id = o.id
        WHERE ko.kit_id = ?
        ORDER BY o.nom
        """, (kit_id, )).fetchall()

    objets_disponibles = db.execute(
        f"""
        SELECT id, nom, (quantite_physique - COALESCE((SELECT SUM(r.quantite_reservee) FROM reservations r WHERE r.objet_id = objets.id AND r.fin_reservation > '{now_str}'), 0)) as quantite_disponible 
        FROM objets
        WHERE id NOT IN (SELECT objet_id FROM kit_objets WHERE kit_id = ?)
        ORDER BY nom
        """, (kit_id, )).fetchall()

    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()

    return render_template("admin_kit_modifier.html",
                           kit=kit,
                           objets_in_kit=objets_in_kit,
                           objets_disponibles=objets_disponibles,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


#=== RETIRER OBJET DUN KIT ===
@admin_bp.route("/kits/retirer_objet/<int:kit_objet_id>", methods=['GET', 'POST'])
@admin_required
def retirer_objet_kit(kit_objet_id):
    db = get_db()
    kit_objet = db.execute("SELECT kit_id FROM kit_objets WHERE id = ?",
                           (kit_objet_id, )).fetchone()
    if kit_objet:
        kit_id = kit_objet['kit_id']
        db.execute("DELETE FROM kit_objets WHERE id = ?", (kit_objet_id, ))
        db.commit()
        flash("Objet retiré du kit.", "success")
        return redirect(url_for('admin.modifier_kit', kit_id=kit_id))
    flash("Erreur : objet du kit non trouvé.", "error")
    return redirect(url_for('admin.gestion_kits'))


#=== SUPPRIMER KIT ===
@admin_bp.route("/kits/supprimer/<int:kit_id>", methods=["POST"])
@admin_required
def supprimer_kit(kit_id):
    db = get_db()
    kit = db.execute("SELECT nom FROM kits WHERE id = ?",
                     (kit_id, )).fetchone()
    if kit:
        db.execute("DELETE FROM kits WHERE id = ?", (kit_id, ))
        db.commit()
        flash(f"Le kit '{kit['nom']}' a été supprimé.", "success")
    else:
        flash("Kit non trouvé.", "error")
    return redirect(url_for('admin.gestion_kits'))
    
    
#===================================
# GESTION FOURNISSEURS 
#===================================
@admin_bp.route("/fournisseurs", methods=['GET', 'POST'])
@admin_required
def gestion_fournisseurs():
    db = get_db()
    if request.method == 'POST':
        nom = request.form.get('nom', '').strip()
        site_web = request.form.get('site_web', '').strip()
        logo_name = None

        if not nom:
            flash("Le nom du fournisseur est obligatoire.", "error")
            return redirect(url_for('admin.gestion_fournisseurs'))

        if 'logo' in request.files:
            logo = request.files['logo']
            if logo and logo.filename != '':
                filename = secure_filename(logo.filename)
                logo.save(os.path.join('static/images/fournisseurs', filename))
                logo_name = filename

        try:
            db.execute(
                "INSERT INTO fournisseurs (nom, site_web, logo) "
                "VALUES (?, ?, ?)", (nom, site_web or None, logo_name))
            db.commit()
            flash(f"Le fournisseur '{nom}' a été ajouté.", "success")
        except sqlite3.IntegrityError:
            flash(f"Un fournisseur avec le nom '{nom}' existe déjà.", "error")
        except sqlite3.Error as e:
            flash(f"Erreur de base de données : {e}", "error")
        return redirect(url_for('admin.gestion_fournisseurs'))

    fournisseurs = db.execute(
        "SELECT * FROM fournisseurs ORDER BY nom").fetchall()
    return render_template("admin_fournisseurs.html",
                           fournisseurs=fournisseurs)


@admin_bp.route("/fournisseurs/supprimer/<int:id>", methods=['POST'])
@admin_required
def supprimer_fournisseur(id):
    db = get_db()
    fournisseur = db.execute("SELECT logo, nom FROM fournisseurs WHERE id = ?",
                             (id, )).fetchone()
    if fournisseur:
        if fournisseur['logo']:
            try:
                os.remove(
                    os.path.join('static/images/fournisseurs',
                                 fournisseur['logo']))
            except OSError:
                pass

        db.execute("DELETE FROM fournisseurs WHERE id = ?", (id, ))
        db.commit()
        flash(f"Le fournisseur '{fournisseur['nom']}' a été supprimé.",
              "success")
    else:
        flash("Fournisseur non trouvé.", "error")
    return redirect(url_for('admin.gestion_fournisseurs'))


@admin_bp.route("/fournisseurs/modifier/<int:id>", methods=['POST'])
@admin_required
def modifier_fournisseur(id):
    db = get_db()
    fournisseur_avant = db.execute("SELECT * FROM fournisseurs WHERE id = ?",
                                   (id, )).fetchone()
    if not fournisseur_avant:
        flash("Fournisseur non trouvé.", "error")
        return redirect(url_for('admin.gestion_fournisseurs'))

    nom = request.form.get('nom', '').strip()
    site_web = request.form.get('site_web', '').strip()
    logo_name = fournisseur_avant['logo']

    if not nom:
        flash("Le nom du fournisseur est obligatoire.", "error")
        return redirect(url_for('admin.gestion_fournisseurs'))

    if request.form.get('supprimer_logo'):
        if logo_name:
            try:
                os.remove(os.path.join('static/images/fournisseurs',
                                       logo_name))
            except OSError:
                pass
        logo_name = None
    elif 'logo' in request.files:
        nouveau_logo = request.files['logo']
        if nouveau_logo and nouveau_logo.filename != '':
            if logo_name:
                try:
                    os.remove(
                        os.path.join('static/images/fournisseurs', logo_name))
                except OSError:
                    pass
            filename = secure_filename(nouveau_logo.filename)
            nouveau_logo.save(
                os.path.join('static/images/fournisseurs', filename))
            logo_name = filename

    try:
        db.execute(
            "UPDATE fournisseurs SET nom = ?, site_web = ?, logo = ? "
            "WHERE id = ?", (nom, site_web or None, logo_name, id))
        db.commit()
        flash(f"Le fournisseur '{nom}' a été mis à jour.", "success")
    except sqlite3.IntegrityError:
        flash(f"Un autre fournisseur avec le nom '{nom}' existe déjà.",
              "error")
    except sqlite3.Error as e:
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('admin.gestion_fournisseurs'))
    
@admin_bp.route("/fournisseurs")
@login_required
def voir_fournisseurs():
    db = get_db()
    fournisseurs = db.execute(
        "SELECT * FROM fournisseurs ORDER BY nom").fetchall()
    return render_template("fournisseurs.html", fournisseurs=fournisseurs)
    

#==================================
# GESTION DU BUDGET
#==================================
@admin_bp.route("/budget", methods=['GET'])
@admin_required
def budget():
    db = get_db()
    now = datetime.now()
    
    # CORRECTION : Le basculement se fait maintenant en août (mois 8)
    annee_scolaire_actuelle = now.year if now.month >= 8 else now.year - 1

    budgets_archives = db.execute(
        "SELECT annee FROM budgets ORDER BY annee DESC").fetchall()

    annee_a_afficher_str = request.args.get('annee', type=str)
    
    if annee_a_afficher_str:
        annee_a_afficher = int(annee_a_afficher_str)
    else:
        # Par défaut, on affiche l'année scolaire en cours
        annee_a_afficher = annee_scolaire_actuelle

    budget_a_afficher = db.execute("SELECT * FROM budgets WHERE annee = ?",
                                   (annee_a_afficher, )).fetchone()

    # Si aucun budget n'existe pour l'année à afficher (cas du premier lancement), on en crée un vide
    if not budget_a_afficher and not budgets_archives:
        try:
            cursor = db.execute(
                "INSERT INTO budgets (annee, montant_initial, cloture) VALUES (?, ?, 0)",
                (annee_a_afficher, 0.0)
            )
            db.commit()
            budget_id = cursor.lastrowid
            budget_a_afficher = db.execute("SELECT * FROM budgets WHERE id = ?", (budget_id,)).fetchone()
            budgets_archives = db.execute("SELECT annee FROM budgets ORDER BY annee DESC").fetchall()
        except sqlite3.IntegrityError:
            db.rollback()
            budget_a_afficher = db.execute("SELECT * FROM budgets WHERE annee = ?", (annee_a_afficher,)).fetchone()

    depenses = []
    total_depenses = 0
    solde = 0
    cloture_autorisee = False

    if budget_a_afficher:
        depenses = db.execute(
            """SELECT d.id, d.contenu, d.montant, d.date_depense,
                      d.est_bon_achat, d.fournisseur_id, f.nom as fournisseur_nom
               FROM depenses d
               LEFT JOIN fournisseurs f ON d.fournisseur_id = f.id
               WHERE d.budget_id = ?
               ORDER BY d.date_depense DESC""",
            (budget_a_afficher['id'], )).fetchall()

        total_depenses_result = db.execute(
            "SELECT SUM(montant) as total FROM depenses WHERE budget_id = ?",
            (budget_a_afficher['id'], )).fetchone()
        total_depenses = (total_depenses_result['total'] if total_depenses_result['total'] is not None else 0)
        solde = budget_a_afficher['montant_initial'] - total_depenses

        # Logique de sécurité pour la clôture
        annee_fin_budget = budget_a_afficher['annee'] + 1
        date_limite_cloture = date(annee_fin_budget, 6, 1) # 1er Juin de l'année N+1
        if date.today() >= date_limite_cloture:
            cloture_autorisee = True

    budget_actuel_pour_modales = db.execute(
        "SELECT * FROM budgets WHERE annee = ? AND cloture = 0",
        (annee_scolaire_actuelle, )).fetchone()

    annee_proposee_pour_creation = annee_scolaire_actuelle
    if not budget_actuel_pour_modales:
        derniere_annee_budget = db.execute("SELECT MAX(annee) as max_annee FROM budgets").fetchone()
        if derniere_annee_budget and derniere_annee_budget['max_annee'] is not None:
            annee_proposee_pour_creation = derniere_annee_budget['max_annee'] + 1
        else:
            annee_proposee_pour_creation = annee_scolaire_actuelle

    fournisseurs = db.execute(
        "SELECT id, nom FROM fournisseurs ORDER BY nom").fetchall()

    return render_template(
        "budget.html",
        budget_affiche=budget_a_afficher,
        budget_actuel_pour_modales=budget_actuel_pour_modales,
        annee_proposee_pour_creation=annee_proposee_pour_creation,
        depenses=depenses,
        total_depenses=total_depenses,
        solde=solde,
        fournisseurs=fournisseurs,
        budgets_archives=budgets_archives,
        annee_selectionnee=annee_a_afficher,
        cloture_autorisee=cloture_autorisee,
        now=datetime.now)


@admin_bp.route("/budget/definir", methods=['POST'])
@admin_required
def definir_budget():
    db = get_db()
    montant = request.form.get('montant_initial')
    annee = request.form.get('annee')

    if not montant or not annee:
        flash("L'année et le montant sont obligatoires.", "error")
        return redirect(url_for('admin.budget'))

    try:
        montant_float = float(montant.replace(',', '.'))
        annee_int = int(annee)

        existing_budget = db.execute(
            "SELECT id FROM budgets WHERE annee = ?", (annee_int,)
        ).fetchone()
        if existing_budget:
            db.execute(
                "UPDATE budgets SET montant_initial = ?, cloture = 0 WHERE id = ?",
                (montant_float, existing_budget['id'])
            )
        else:
            db.execute(
                "INSERT INTO budgets (annee, montant_initial) VALUES (?, ?)",
                (annee_int, montant_float)
            )

        db.commit()
        flash(
            f"Le budget pour l'année scolaire {annee_scolaire_format(annee_int)} a été défini à "
            f"{montant_float:.2f} €.", "success"
        )
    except ValueError:
        flash("Le montant ou l'année saisi(e) est invalide.", "error")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('admin.budget', annee=annee))


@admin_bp.route("/budget/ajouter_depense", methods=['POST'])
@admin_required
def ajouter_depense():
    db = get_db()
    budget_id = request.form.get('budget_id')
    fournisseur_id = request.form.get('fournisseur_id')
    contenu = request.form.get('contenu', '').strip()
    montant = request.form.get('montant')
    date_depense = request.form.get('date_depense')
    est_bon_achat = 1 if request.form.get('est_bon_achat') == 'on' else 0

    if not all([budget_id, contenu, montant, date_depense]):
        flash("Tous les champs sont obligatoires pour ajouter une dépense.",
              "error")
        return redirect(url_for('admin.budget'))

    if est_bon_achat:
        fournisseur_id = None
    elif not fournisseur_id:
        flash(
            "Veuillez sélectionner un fournisseur ou cocher la case "
            "'Bon d'achat'.", "error")
        return redirect(url_for('admin.budget'))

    try:
        montant_float = float(montant.replace(',', '.'))
        db.execute(
            """INSERT INTO depenses (budget_id, fournisseur_id, contenu,
               montant, date_depense, est_bon_achat)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (budget_id, fournisseur_id, contenu, montant_float, date_depense,
             est_bon_achat))
        db.commit()
        flash("La dépense a été ajoutée avec succès.", "success")
    except ValueError:
        flash("Le montant saisi est invalide.", "error")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('admin.budget'))


@admin_bp.route("/budget/modifier_depense/<int:id>", methods=['POST'])
@admin_required
def modifier_depense(id):
    db = get_db()
    depense = db.execute("SELECT id FROM depenses WHERE id = ?",
                         (id, )).fetchone()
    if not depense:
        flash("Dépense non trouvée.", "error")
        return redirect(url_for('admin.budget'))

    fournisseur_id = request.form.get('fournisseur_id')
    contenu = request.form.get('contenu', '').strip()
    montant = request.form.get('montant')
    date_depense = request.form.get('date_depense')
    est_bon_achat = 1 if request.form.get('est_bon_achat') == 'on' else 0

    if not all([contenu, montant, date_depense]):
        flash("Les champs contenu, montant et date sont obligatoires.",
              "error")
        return redirect(request.referrer or url_for('admin.budget'))

    if est_bon_achat:
        fournisseur_id = None
    elif not fournisseur_id:
        flash(
            "Veuillez sélectionner un fournisseur ou cocher la case "
            "'Bon d'achat'.", "error")
        return redirect(request.referrer or url_for('admin.budget'))

    try:
        montant_float = float(montant.replace(',', '.'))
        db.execute(
            """UPDATE depenses SET fournisseur_id = ?, contenu = ?,
               montant = ?, date_depense = ?, est_bon_achat = ?
               WHERE id = ?""", (fournisseur_id, contenu, montant_float,
                                 date_depense, est_bon_achat, id))
        db.commit()
        flash("La dépense a été modifiée avec succès.", "success")
    except ValueError:
        flash("Le montant saisi est invalide.", "error")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(request.referrer or url_for('admin.budget'))


@admin_bp.route("/budget/supprimer_depense/<int:id>", methods=['POST'])
@admin_required
def supprimer_depense(id):
    db = get_db()
    depense = db.execute("SELECT id FROM depenses WHERE id = ?",
                         (id, )).fetchone()
    if depense:
        try:
            db.execute("DELETE FROM depenses WHERE id = ?", (id, ))
            db.commit()
            flash("La dépense a été supprimée avec succès.", "success")
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données : {e}", "error")
    else:
        flash("Dépense non trouvée.", "error")

    return redirect(request.referrer or url_for('admin.budget'))


@admin_bp.route("/budget/cloturer", methods=['POST'])
@admin_required
def cloturer_budget():
    budget_id = request.form.get('budget_id')
    db = get_db()
    
    budget = db.execute("SELECT * FROM budgets WHERE id = ?", (budget_id,)).fetchone()

    if not budget:
        flash("Budget non trouvé.", "error")
        return redirect(url_for('admin.budget'))

    # --- VÉRIFICATION DE SÉCURITÉ CÔTÉ SERVEUR ---
    annee_fin_budget = budget['annee'] + 1
    date_limite_cloture = date(annee_fin_budget, 6, 1)
    if date.today() < date_limite_cloture:
        flash(f"La clôture du budget {annee_scolaire_format(budget['annee'])} n'est autorisée qu'à partir du {date_limite_cloture.strftime('%d/%m/%Y')}.", "error")
        return redirect(url_for('admin.budget', annee=budget['annee']))

    if budget['cloture']:
        flash(f"Le budget pour l'année scolaire {annee_scolaire_format(budget['annee'])} est déjà clôturé.", "warning")
        return redirect(url_for('admin.budget'))

    try:
        db.execute("UPDATE budgets SET cloture = 1 WHERE id = ?", (budget_id,))
        db.commit()
        flash(f"Le budget pour l'année scolaire {annee_scolaire_format(budget['annee'])} a été clôturé avec succès.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('admin.budget'))

@admin_bp.route("/budget/voir")
@login_required
def voir_budget():
    db = get_db()
    now = datetime.now()
    
    # Logique de l'année scolaire : commence en septembre.
    annee_scolaire_actuelle = now.year if now.month >= 9 else now.year - 1
    
    budget_actuel = db.execute(
        "SELECT * FROM budgets WHERE annee = ? AND cloture = 0",
        (annee_scolaire_actuelle, )).fetchone()

    depenses = []
    total_depenses = 0
    solde = 0

    if budget_actuel:
        depenses = db.execute(
            """SELECT d.id, d.contenu, d.montant, d.date_depense,
                      d.est_bon_achat, d.fournisseur_id, f.nom as fournisseur_nom
            FROM depenses d
            LEFT JOIN fournisseurs f ON d.fournisseur_id = f.id
            WHERE d.budget_id = ?
            ORDER BY d.date_depense DESC""",
            (budget_actuel['id'], )).fetchall()

        total_depenses_result = db.execute(
            "SELECT SUM(montant) as total FROM depenses WHERE budget_id = ?",
            (budget_actuel['id'], )).fetchone()
        total_depenses = (total_depenses_result['total']
                          if total_depenses_result['total'] is not None else 0)
        solde = budget_actuel['montant_initial'] - total_depenses

    return render_template("budget_voir.html",
                           budget_actuel=budget_actuel,
                           depenses=depenses,
                           total_depenses=total_depenses,
                           solde=solde)


#=======================================
# GESTION DES ARMOIRES ET CATEGORIES
#=======================================
@admin_bp.route("/ajouter", methods=["POST"])
@login_required
def ajouter():
    type_objet = request.form.get("type")
    nom = request.form.get("nom", "").strip()
    redirect_to = ("admin.gestion_armoires"
                   if type_objet == "armoire" else "admin.gestion_categories")
    if not nom:
        flash("Le nom ne peut pas être vide.", "error")
        return redirect(url_for(redirect_to))
    table_name = "armoires" if type_objet == "armoire" else "categories"
    db = get_db()
    try:
        db.execute(f"INSERT INTO {table_name} (nom) VALUES (?)", (nom, ))
        db.commit()
        flash(f"L'élément '{nom}' a été créé.", "success")
    except sqlite3.IntegrityError:
        flash(f"L'élément '{nom}' existe déjà.", "error")
    return redirect(url_for(redirect_to))

@admin_bp.route("/supprimer/<type_objet>/<int:id>", methods=["POST"])
@admin_required
def supprimer(type_objet, id):
    db = get_db()
    redirect_to = ("admin.gestion_armoires"
                   if type_objet == "armoire" else "admin.gestion_categories")
    if type_objet == "armoire":
        if db.execute("SELECT COUNT(id) FROM objets WHERE armoire_id = ?",
                      (id, )).fetchone()[0] > 0:
            flash(
                "Impossible de supprimer. Cette armoire contient encore "
                "des objets.", "error")
            return redirect(url_for(redirect_to))
    table_map = {"armoire": "armoires", "categorie": "categories"}
    if type_objet in table_map:
        table_name = table_map[type_objet]
        nom_element = db.execute(f"SELECT nom FROM {table_name} WHERE id = ?",
                                 (id, )).fetchone()
        db.execute(f"DELETE FROM {table_name} WHERE id = ?", (id, ))
        db.commit()
        if nom_element:
            flash(
                f"L'élément '{nom_element['nom']}' a été supprimé avec succès.",
                "success")
    else:
        flash("Type d'élément à supprimer non valide.", "error")
    return redirect(url_for(redirect_to))
    

@adm_bp.route("/gestion_armoires")
@login_required
def gestion_armoires():
    db = get_db()
    armoires = db.execute("""
        SELECT a.id, a.nom, COUNT(o.id) as count FROM armoires a
        LEFT JOIN objets o ON a.id = o.armoire_id
        GROUP BY a.id, a.nom ORDER BY a.nom
        """).fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()
    return render_template("gestion_armoires.html",
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)


@admin.bp.route("/gestion_categories")
@login_required
def gestion_categories():
    db = get_db()
    categories = db.execute("""
        SELECT c.id, c.nom, COUNT(o.id) as count FROM categories c
        LEFT JOIN objets o ON c.id = o.categorie_id
        GROUP BY c.id, c.nom ORDER BY c.nom
        """).fetchall()
    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    return render_template("gestion_categories.html",
                           categories=categories,
                           armoires=armoires,
                           now=datetime.now)
                           

@admin_bp.route("/modifier_armoire", methods=["POST"])
@admin_required
def modifier_armoire():
    data = request.get_json()
    armoire_id, nouveau_nom = data.get("id"), data.get("nom")
    if not all([armoire_id, nouveau_nom, nouveau_nom.strip()]):
        return jsonify(success=False, error="Données invalides"), 400
    db = get_db()
    try:
        db.execute("UPDATE armoires SET nom = ? WHERE id = ?",
                   (nouveau_nom.strip(), armoire_id))
        db.commit()
        return jsonify(success=True, nouveau_nom=nouveau_nom.strip())
    except sqlite3.IntegrityError:
        return jsonify(success=False,
                       error="Ce nom d'armoire existe déjà."), 500


@admin_bp.route("/modifier_categorie", methods=["POST"])
@admin_required
def modifier_categorie():
    data = request.get_json()
    categorie_id, nouveau_nom = data.get("id"), data.get("nom")
    if not all([categorie_id, nouveau_nom, nouveau_nom.strip()]):
        return jsonify(success=False, error="Données invalides"), 400
    db = get_db()
    try:
        db.execute("UPDATE categories SET nom = ? WHERE id = ?",
                   (nouveau_nom.strip(), categorie_id))
        db.commit()
        return jsonify(success=True, nouveau_nom=nouveau_nom.strip())
    except sqlite3.IntegrityError:
        return jsonify(success=False,
                       error="Ce nom de catégorie existe déjà."), 500


#========================================
# GESTION RAPPORTS ET EXPORTS
#========================================
@admin_bp.route("/rapports", methods=['GET'])
@admin_required
def rapports():
    db = get_db()
    dernieres_actions = db.execute("""
        SELECT h.timestamp, h.action, o.nom as objet_nom, u.nom_utilisateur
        FROM historique h
        JOIN objets o ON h.objet_id = o.id
        JOIN utilisateurs u ON h.utilisateur_id = u.id
        ORDER BY h.timestamp DESC
        LIMIT 5
        """).fetchall()

    armoires = db.execute("SELECT * FROM armoires ORDER BY nom").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY nom").fetchall()

    return render_template("rapports.html",
                           dernieres_actions=dernieres_actions,
                           armoires=armoires,
                           categories=categories,
                           now=datetime.now)

@admin_bp.route("/rapports/exporter", methods=['GET'])
@admin_required
def exporter_rapports():
    db = get_db()
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    group_by = request.args.get('group_by')
    format_type = request.args.get('format')

    if not all([date_debut, date_fin, group_by, format_type]):
        flash("Tous les champs sont requis pour générer un rapport.", "error")
        return redirect(url_for('admin.rapports'))

    try:
        date_fin_dt = datetime.strptime(date_fin, '%Y-%m-%d') + timedelta(days=1)
        date_fin_str = date_fin_dt.strftime('%Y-%m-%d')
    except ValueError:
        flash("Format de date invalide.", "error")
        return redirect(url_for('admin.rapports'))

    query = """
        SELECT h.timestamp, h.action, h.details, o.nom as objet_nom,
               u.nom_utilisateur
        FROM historique h
        JOIN objets o ON h.objet_id = o.id
        JOIN utilisateurs u ON h.utilisateur_id = u.id
        WHERE h.timestamp >= ? AND h.timestamp < ?
    """
    order_clause = "ORDER BY h.timestamp ASC"
    if group_by == 'action':
        order_clause = "ORDER BY h.action ASC, h.timestamp ASC"
    query += order_clause

    historique_data = db.execute(query, (date_debut, date_fin_str)).fetchall()

    if not historique_data:
        flash("Aucune donnée d'historique trouvée pour la période sélectionnée.", "warning")
        return redirect(url_for('admin.rapports'))

    if format_type == 'pdf':
        buffer = generer_rapport_pdf(historique_data, date_debut, date_fin, group_by)
        return send_file(buffer,
                         as_attachment=True,
                         download_name='rapport_activite.pdf',
                         mimetype='application/pdf')
    elif format_type == 'excel':
        buffer = generer_rapport_excel(historique_data, date_debut, date_fin, group_by)
        return send_file(
            buffer,
            as_attachment=True,
            download_name='rapport_activite.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    flash("Format d'exportation non valide.", "error")
    return redirect(url_for('admin.rapports'))

def generer_rapport_pdf(data, date_debut, date_fin, group_by):
    pdf = PDFWithFooter()
    pdf.alias_nb_pages()
    pdf.add_page(orientation='L')
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'Rapport d\'Activite', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(
        0, 10,
        f"Periode du {datetime.strptime(date_debut, '%Y-%m-%d').strftime('%d/%m/%Y')} "
        f"au {datetime.strptime(date_fin, '%Y-%m-%d').strftime('%d/%m/%Y')}",
        new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(5)

    col_widths = {"date": 25, "heure": 15, "user": 35, "action": 60, "objet": 60, "details": 75}
    table_width = sum(col_widths.values())
    line_height = 7 # Hauteur de ligne fixe

    def draw_header():
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(col_widths["date"], 8, 'Date', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        pdf.cell(col_widths["heure"], 8, 'Heure', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        pdf.cell(col_widths["user"], 8, 'Utilisateur', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        pdf.cell(col_widths["action"], 8, 'Action', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        pdf.cell(col_widths["objet"], 8, 'Objet Concerne', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
        pdf.cell(col_widths["details"], 8, 'Details', border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C', fill=True)

    draw_header()

    current_group = None
    
    for item in data:
        if pdf.get_y() > 190: # Marge de sécurité pour le saut de page
            pdf.add_page(orientation='L')
            draw_header()
            if current_group:
                pdf.set_font('Helvetica', 'B', 10)
                pdf.set_fill_color(230, 240, 255)
                pdf.cell(table_width, 8, f"Type d'action : {current_group}", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L', fill=True)

        if group_by == 'action' and item['action'] != current_group:
            current_group = item['action']
            pdf.set_font('Helvetica', 'B', 10)
            pdf.set_fill_color(230, 240, 255)
            pdf.cell(table_width, 8, f"Type d'action : {current_group}", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L', fill=True)

        pdf.set_font('Helvetica', '', 8)
        timestamp_dt = datetime.fromisoformat(item['timestamp'])
        
        pdf.cell(col_widths["date"], line_height, timestamp_dt.strftime('%d/%m/%Y'), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        pdf.cell(col_widths["heure"], line_height, timestamp_dt.strftime('%H:%M'), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        pdf.cell(col_widths["user"], line_height, item['nom_utilisateur'].encode('latin-1', 'replace').decode('latin-1'), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        pdf.cell(col_widths["action"], line_height, item['action'].encode('latin-1', 'replace').decode('latin-1'), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        pdf.cell(col_widths["objet"], line_height, item['objet_nom'].encode('latin-1', 'replace').decode('latin-1'), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='L')
        pdf.cell(col_widths["details"], line_height, item['details'].encode('latin-1', 'replace').decode('latin-1'), border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')

    return BytesIO(pdf.output())


def generer_rapport_excel(data, date_debut, date_fin, group_by):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rapport d'Activite"

    title_font = Font(name='Calibri', size=16, bold=True)
    subtitle_font = Font(name='Calibri', size=11, italic=True, color="6c7a89")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568",
                              end_color="4A5568",
                              fill_type="solid")
    group_font = Font(name='Calibri', size=11, bold=True, color="2D3748")
    group_fill = PatternFill(start_color="E2E8F0",
                             end_color="E2E8F0",
                             fill_type="solid")
    even_row_fill = PatternFill(start_color="F7FAFC",
                                end_color="F7FAFC",
                                fill_type="solid")

    center_align = Alignment(horizontal='center',
                             vertical='center',
                             wrap_text=True)
    left_align = Alignment(horizontal='left',
                           vertical='center',
                           wrap_text=True)

    thin_border_side = Side(style='thin', color="4A5568")
    thick_border_side = Side(style='medium', color="000000")

    cell_border = Border(left=thin_border_side,
                         right=thin_border_side,
                         top=thin_border_side,
                         bottom=thin_border_side)

    start_col = 2
    headers = [
        "Date", "Heure", "Utilisateur", "Action", "Objet Concerne", "Details"
    ]
    end_col = start_col + len(headers) - 1

    sheet.merge_cells(start_row=2,
                      start_column=start_col,
                      end_row=2,
                      end_column=end_col)
    title_cell = sheet.cell(row=2, column=start_col)
    title_cell.value = "Rapport d'Activite"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    sheet.merge_cells(start_row=3,
                      start_column=start_col,
                      end_row=3,
                      end_column=end_col)
    subtitle_cell = sheet.cell(row=3, column=start_col)
    subtitle_cell.value = (
        f"Periode du "
        f"{datetime.strptime(date_debut, '%Y-%m-%d').strftime('%d/%m/%Y')} au "
        f"{datetime.strptime(date_fin, '%Y-%m-%d').strftime('%d/%m/%Y')}")
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    header_row = 5
    for i, header_text in enumerate(headers, start=start_col):
        cell = sheet.cell(row=header_row, column=i, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border

    row_index = header_row + 1
    current_group = None
    last_date_str = None
    is_even = False

    for item in data:
        if group_by == 'action' and item['action'] != current_group:
            current_group = item['action']
            last_date_str = None
            cell = sheet.cell(row=row_index,
                              column=start_col,
                              value=f"Type d'action : {current_group}")
            sheet.merge_cells(start_row=row_index,
                              start_column=start_col,
                              end_row=row_index,
                              end_column=end_col)
            for c in range(start_col, end_col + 1):
                sheet.cell(row=row_index, column=c).fill = group_fill
                sheet.cell(row=row_index, column=c).font = group_font
                sheet.cell(row=row_index, column=c).border = cell_border
            row_index += 1
            is_even = False

        timestamp_dt = datetime.fromisoformat(item['timestamp'])
        current_date_str = timestamp_dt.strftime('%d/%m/%Y')

        date_str_display = current_date_str
        if group_by == 'date' and current_date_str == last_date_str:
            date_str_display = ""

        row_data = [
            date_str_display,
            timestamp_dt.strftime('%H:%M'), item['nom_utilisateur'],
            item['action'], item['objet_nom'], item['details']
        ]

        for col_index, value in enumerate(row_data, start=start_col):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = cell_border
            if col_index in [
                    start_col, start_col + 1, start_col + 2, start_col + 3
            ]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            if is_even:
                cell.fill = even_row_fill

        last_date_str = current_date_str
        row_index += 1
        is_even = not is_even

    end_row_index = row_index - 1
    if end_row_index >= header_row:
        for row in sheet.iter_rows(min_row=header_row,
                                   max_row=end_row_index,
                                   min_col=start_col,
                                   max_col=end_col):
            for cell in row:
                new_border = Border(left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom)
                if cell.row == header_row:
                    new_border.top = thick_border_side
                if cell.row == end_row_index:
                    new_border.bottom = thick_border_side
                if cell.column == start_col:
                    new_border.left = thick_border_side
                if cell.column == end_col:
                    new_border.right = thick_border_side
                cell.border = new_border

    column_widths = {'B': 15, 'C': 10, 'D': 25, 'E': 50, 'F': 40, 'G': 60}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=start_col)

    if end_row_index >= header_row:
        sheet.auto_filter.ref = (
            f"{sheet.cell(row=header_row, column=start_col).coordinate}:"
            f"{sheet.cell(row=end_row_index, column=end_col).coordinate}")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@admin.route("/exporter_inventaire")
@admin_required
def exporter_inventaire():
    db = get_db()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    inventaire_data = db.execute("""
        SELECT 
            o.nom, 
            c.nom AS categorie,
            a.nom AS armoire,
            (o.quantite_physique - COALESCE(SUM(r.quantite_reservee), 0)) as quantite_disponible,
            o.seuil,
            o.date_peremption
        FROM objets o 
        JOIN armoires a ON o.armoire_id = a.id
        JOIN categories c ON o.categorie_id = c.id
        LEFT JOIN reservations r ON o.id = r.objet_id AND r.fin_reservation > ?
        GROUP BY o.id, o.nom, c.nom, a.nom, o.seuil, o.date_peremption
        ORDER BY c.nom, o.nom
    """, (now_str,)).fetchall()
        
    format_type = request.args.get('format')
    if format_type == 'pdf':
        buffer = generer_inventaire_pdf(inventaire_data)
        return send_file(buffer, as_attachment=True, download_name='inventaire_complet.pdf', mimetype='application/pdf')
    elif format_type == 'excel':
        buffer = generer_inventaire_excel(inventaire_data)
        return send_file(buffer, as_attachment=True, download_name='inventaire_complet.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    flash("Format d'exportation non valide.","error")
    return redirect(url_for('admin.admin'))


def generer_inventaire_pdf(data):
    pdf = PDFWithFooter()
    pdf.alias_nb_pages()
    pdf.add_page(orientation='L')
    
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'Inventaire Complet du Laboratoire', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 10, f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(10)

    # En-têtes du tableau
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(50, 8, 'Catégorie', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(80, 8, 'Nom de l\'objet', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(25, 8, 'Qté Dispo', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(20, 8, 'Seuil', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(50, 8, 'Armoire', border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(30, 8, 'Péremption', border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C', fill=True)
    
    # Regrouper données par catégorie
    grouped_data = {}
    for item in data:
        cat = item['categorie']
        if cat not in grouped_data:
            grouped_data[cat] = []
        grouped_data[cat].append(item)

    pdf.set_font('Helvetica', '', 8)
    for categorie, items in sorted(grouped_data.items()):
        row_count = len(items)
        start_y = pdf.get_y()
        start_x = pdf.get_x()
        for i, item in enumerate(items):
            date_peremption_str = ""
            if item['date_peremption']:
                try:
                    date_obj = datetime.strptime(item['date_peremption'], '%Y-%m-%d')
                    date_peremption_str = date_obj.strftime('%d/%m/%Y')
                except (ValueError, TypeError):
                    date_peremption_str = item['date_peremption']
                    
            pdf.set_x(start_x + 50)
            
            pdf.cell(80, 7, item['nom'].encode('latin-1', 'replace').decode('latin-1'), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='L')
            pdf.cell(25, 7, str(item['quantite_disponible']), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
            pdf.cell(20, 7, str(item['seuil']), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
            pdf.cell(50, 7, item['armoire'].encode('latin-1', 'replace').decode('latin-1'), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
            pdf.cell(30, 7, date_peremption_str, border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')

        end_y = pdf.get_y()
        total_height = end_y - start_y

        pdf.set_y(start_y)
        pdf.set_x(start_x)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.multi_cell(50, total_height, categorie.encode('latin-1', 'replace').decode('latin-1'), border=1, align='C')
        pdf.set_font('Helvetica', '', 8)

        pdf.set_y(end_y)
        
    return BytesIO(pdf.output())

def generer_inventaire_excel(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventaire Complet"

    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    category_align = Alignment(horizontal='left', vertical='center')

    sheet.merge_cells('A1:F1')
    sheet['A1'] = 'Inventaire Complet du Laboratoire'
    sheet['A1'].font = title_font
    sheet['A1'].alignment = center_align
    sheet.merge_cells('A2:F2')
    sheet['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"

    headers = ["Catégorie", "Nom de l'objet", "Quantité Disponible", "Seuil", "Armoire", "Date de Péremption"]
    for i, header_text in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=i, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    row_index = 5
    current_category = None
    start_merge_row = 5

    for item in data:
        if item['categorie'] != current_category:
            if current_category is not None:
                if start_merge_row < row_index - 1:
                    sheet.merge_cells(start_row=start_merge_row, start_column=1, end_row=row_index - 1, end_column=1)
                    sheet.cell(row=start_merge_row, column=1).alignment = category_align
            
            current_category = item['categorie']
            start_merge_row = row_index

        sheet.cell(row=row_index, column=1, value=item['categorie'])
        sheet.cell(row=row_index, column=2, value=item['nom']).alignment = left_align
        sheet.cell(row=row_index, column=3, value=item['quantite_disponible']).alignment = center_align
        sheet.cell(row=row_index, column=4, value=item['seuil']).alignment = center_align
        
        armoire_cell = sheet.cell(row=row_index, column=5, value=item['armoire'])
        armoire_cell.alignment = center_align
        
        date_cell = sheet.cell(row=row_index, column=6)
        if item['date_peremption']:
            date_cell.value = datetime.strptime(item['date_peremption'], '%Y-%m-%d')
            date_cell.number_format = 'DD/MM/YYYY'
        date_cell.alignment = center_align
        
        row_index += 1

    if current_category is not None and start_merge_row < row_index - 1:
        sheet.merge_cells(start_row=start_merge_row, start_column=1, end_row=row_index - 1, end_column=1)
        sheet.cell(row=start_merge_row, column=1).alignment = category_align

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 20
    
    sheet.freeze_panes = 'A5'
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


#===============================================
# GESTION ECHEANCES
#===============================================
@admin_bp.route("/echeances")
@admin_required
def gestion_echeances():
    db = get_db()
    echeances_brutes = db.execute(
        "SELECT * FROM echeances ORDER BY date_echeance ASC").fetchall()

    echeances_converties = []
    for echeance in echeances_brutes:
        echeance_dict = dict(echeance)
        echeance_dict['date_echeance'] = datetime.strptime(
            echeance['date_echeance'], '%Y-%m-%d').date()
        echeances_converties.append(echeance_dict)

    return render_template("admin_echeances.html",
                           echeances=echeances_converties,
                           date_actuelle=datetime.now().date(),
                           url_ajout=url_for('admin.ajouter_echeance'))


@admin_bp.route("/echeances/ajouter", methods=['POST'])
@admin_required
def ajouter_echeance():
    intitule = request.form.get('intitule', '').strip()
    date_echeance = request.form.get('date_echeance')
    details = request.form.get('details', '').strip()

    if not all([intitule, date_echeance]):
        flash("L'intitulé et la date d'échéance sont obligatoires.", "error")
        return redirect(url_for('admin.gestion_echeances'))

    db = get_db()
    try:
        db.execute(
            "INSERT INTO echeances (intitule, date_echeance, details) "
            "VALUES (?, ?, ?)", (intitule, date_echeance, details or None))
        db.commit()
        flash("L'échéance a été ajoutée avec succès.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('admin.gestion_echeances'))


@admin_bp.route("/echeances/modifier/<int:id>", methods=['POST'])
@admin_required
def modifier_echeance(id):
    db = get_db()
    echeance = db.execute("SELECT id FROM echeances WHERE id = ?",
                          (id, )).fetchone()
    if not echeance:
        flash("Échéance non trouvée.", "error")
        return redirect(url_for('admin.gestion_echeances'))

    intitule = request.form.get('intitule', '').strip()
    date_echeance = request.form.get('date_echeance')
    details = request.form.get('details', '').strip()
    traite = 1 if request.form.get('traite') == 'on' else 0

    if not all([intitule, date_echeance]):
        flash("L'intitulé et la date d'échéance sont obligatoires.", "error")
        return redirect(url_for('admin.gestion_echeances'))

    try:
        db.execute(
            "UPDATE echeances SET intitule = ?, date_echeance = ?, "
            "details = ?, traite = ? WHERE id = ?",
            (intitule, date_echeance, details or None, traite, id))
        db.commit()
        flash("L'échéance a été modifiée avec succès.", "success")
    except sqlite3.Error as e:
        db.rollback()
        flash(f"Erreur de base de données : {e}", "error")

    return redirect(url_for('admin.gestion_echeances'))


@admin_bp.route("/echeances/supprimer/<int:id>", methods=['POST'])
@admin_required
def supprimer_echeance(id):
    db = get_db()
    echeance = db.execute("SELECT id FROM echeances WHERE id = ?",
                          (id, )).fetchone()
    if echeance:
        try:
            db.execute("DELETE FROM echeances WHERE id = ?", (id, ))
            db.commit()
            flash("L'échéance a été supprimée avec succès.", "success")
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Erreur de base de données : {e}", "error")
    else:
        flash("Échéance non trouvée.", "error")

    return redirect(url_for('admin.gestion_echeances'))